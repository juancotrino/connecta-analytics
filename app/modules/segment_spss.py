import os
import io
import re
import itertools
from io import BytesIO
import tempfile
import zipfile
from copy import copy
from datetime import datetime
from pytz import timezone

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import GradientStop, GradientFillProperties

import pyreadstat
import numpy as np
import pandas as pd
from scipy.stats import chi2_contingency, pearsonr
# import matplotlib.pyplot as plt
# import seaborn as sns

from app.cloud import CloudStorageClient
from app.modules.authenticator import get_inverted_scales_keywords
from app.modules.utils import get_temp_file

time_zone = timezone('America/Bogota')

scales = {
    'T2B':{
        'regular': {5: 1, 4: 1, 3: 0, 2: 0, 1: 0},
        'inverted': {1: 1, 2: 1, 3: 0, 4: 0, 5: 0},
        'just_right': {1: 0, 2: 0, 3: 1, 4: 0, 5: 0}
    },
    'TB': {
        'regular': {5: 1, 4: 0, 3: 0, 2: 0, 1: 0},
        'inverted': {1: 1, 2: 0, 3: 0, 4: 0, 5: 0},
        'just_right': {1: 0, 2: 0, 3: 1, 4: 0, 5: 0}
    }
}

# Function to perform chi-square test and return chi2, p-value, and Cramer's V
def chi_square_test(column1, column2, correction: bool = False):
    contingency_table = pd.crosstab(column1, column2.fillna(0))
    chi2, p_value, _, _ = chi2_contingency(contingency_table, correction=correction)
    n = contingency_table.sum().sum()
    phi2 = chi2 / n
    r, k = contingency_table.shape
    if correction:
        phi2corr = max(0, phi2 - ((k - 1) * (r - 1)) / (n - 1))
        rcorr = r - ((r - 1) ** 2) / (n - 1)
        kcorr = k - ((k - 1) ** 2) / (n - 1)
        return chi2, p_value, np.sqrt(phi2corr / min((kcorr - 1), (rcorr - 1)))
    else:
        return chi2, p_value, np.sqrt(phi2 / (k - 1)) if k > 1 else 0

def process_chi2(
    file_path: str,
    cross_variable: str,
    chi2_mode: str,
    inverted_scales_keywords: list,
    correction: bool = False,
):

    data, study_metadata = pyreadstat.read_sav(
        file_path,
        apply_value_formats=False
    )

    variables_data = study_metadata.variable_value_labels

    filtered_variables = [variable for variable, scale in variables_data.items() if len(scale) == 5]
    filtered_variables_data = {k: v for k, v in variables_data.items() if k in filtered_variables}

    just_right_variables = [code for code, tag in filtered_variables_data.items() if 'just' in ''.join(tag.values()).lower()]

    inverted_variables = [
        code for code, tag in filtered_variables_data.items()
        if any(
            keyword in ''.join(tag.values()).lower()
            for keyword in inverted_scales_keywords
        ) and (
            code not in just_right_variables
        )
    ]

    filtered_data: pd.DataFrame = data[data.columns[data.columns.isin(filtered_variables)]]
    filtered_data_bool = filtered_data.copy()

    variables_scale_info = pd.DataFrame(list(study_metadata.column_names_to_labels.items()), columns=['code', 'text']).replace({None: np.nan})

    variables_scale_info['scale_type'] = np.where(
        variables_scale_info['code'].isin(just_right_variables),
        'just_right',
        np.where(
            variables_scale_info['code'].isin(inverted_variables),
            'inverted',
            'regular'
        )
    )
    variables_scale_info['scale_type'] = variables_scale_info['scale_type'].replace({None: np.nan})

    for column in filtered_data:
        scale_type = variables_scale_info[variables_scale_info['code'] == column]['scale_type'].values[0]
        filtered_data_bool.loc[:, column] = filtered_data.loc[:, column].map(scales[chi2_mode][scale_type])

    # Dictionary to store results
    results = {}
    # Perform chi-square test for each column
    for column in filtered_data_bool.columns:
        if all(filtered_data_bool[column]):
            continue
        if column != cross_variable:
            chi2, p_value, cramer_v = chi_square_test(filtered_data_bool[cross_variable], filtered_data_bool[column], correction)
            results[column] = {'Chi-square': chi2, 'P-value': p_value, 'Cramer\'s V': cramer_v}

    # Create a DataFrame from the results
    results_df = pd.DataFrame.from_dict(results, orient='index')

    results_df = results_df.reset_index(names='code').merge(variables_scale_info, on='code')

    if not results_df.empty:
        return results_df[['text', 'Cramer\'s V', 'P-value', 'scale_type']].sort_values(by=['Cramer\'s V'], ascending=True)
    else:
        return pd.DataFrame()

def get_correlation_data(
    file_path: str,
    correlation_variables: list[str]
):

    data: pd.DataFrame = pyreadstat.read_sav(
        file_path,
        apply_value_formats=False
    )[0]

    return data[correlation_variables]

def process_correlation(data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:

    # Initialize empty dataframes to store results
    correlation_df = pd.DataFrame(index=data.columns, columns=data.columns)
    p_value_df = pd.DataFrame(index=data.columns, columns=data.columns)

    # Calculate Pearson's correlation coefficient and p-values
    for col1 in data.columns:
        if data[col1].notnull().sum() == 0:  # Skip column if it has no non-null values
            continue
        for col2 in data.columns:
            if col2 == col1 or data[col2].notnull().sum() == 0:  # Skip column if it has no non-null values or if it's the same column
                correlation_df.loc[col1, col2] = 1.0 if col1 == col2 else np.nan
                p_value_df.loc[col1, col2] = 0.0 if col1 == col2 else np.nan
                continue

            # Drop rows with missing values in either column
            valid_rows = data[[col1, col2]].dropna()
            if valid_rows.shape[0] > 1:  # Ensure there are at least two data points
                # Check for sufficient variance in the data
                if np.var(valid_rows[col1]) == 0 or np.var(valid_rows[col2]) == 0:
                    correlation_df.loc[col1, col2] = np.nan
                    p_value_df.loc[col1, col2] = np.nan
                else:
                    corr, p_value = pearsonr(valid_rows[col1], valid_rows[col2])
                    correlation_df.loc[col1, col2] = corr
                    p_value_df.loc[col1, col2] = p_value
            else:
                correlation_df.loc[col1, col2] = np.nan
                p_value_df.loc[col1, col2] = np.nan

    # Convert the results to numeric type
    correlation_df = correlation_df.astype(float)
    p_value_df = p_value_df.astype(float)

    return correlation_df, p_value_df

def copy_axis_style(source_axis, invisible: bool):
    target_axis = copy(source_axis)
    target_axis.majorGridlines = copy_chart_lines(source_axis.majorGridlines, invisible)
    return target_axis

def copy_chart_lines(source_lines, invisible: bool):
    if invisible:
        return None
    else:
        target_lines = openpyxl.chart.line_chart.ChartLines()
        if source_lines:
            if source_lines.spPr is not None:
                target_lines.spPr = copy(source_lines.spPr)
        return target_lines

def copy_chart_style(source_chart, target_chart):
    target_chart.type = "bar"
    target_chart.legend = None

    for src_serie, tgt_serie in zip(source_chart.series, target_chart.series):
        tgt_serie.graphicalProperties = copy(src_serie.graphicalProperties)
        tgt_serie.tx = copy(src_serie.tx)
        tgt_serie.invertIfNegative = copy(src_serie.invertIfNegative)
        tgt_serie.dLbls = copy(src_serie.dLbls)

    # Copy axis styles
    target_chart.x_axis = copy_axis_style(source_chart.x_axis, invisible=True)
    target_chart.y_axis = copy_axis_style(source_chart.y_axis, invisible=False)

# Define function to get the color based on value
def get_color(value):
    match value:
        case x if x == 0:
            return 'FFEB84'
        case x if 0 < x < 0.2:
            return 'E2E383'  # light green
        case x if 0.2 <= x < 0.4:
            return 'C2DA81'  # green
        case x if 0.4 <= x < 0.7:
            return '94CD7F'  # lime green
        case x if 0.7 <= x < 0.9:
            return '74C37C'  # green
        case x if 0.9 <= x < 1.0:
            return '65BE7C'  # dark green
        case x if x == 1:
            return '63BE7B'  # dark green
        case _:
            return 'FFFFFF'  # white

def create_chart(worksheet, source_chart, dataframe, chart_title, chart_destination):
    # If you have the data in a DataFrame
    data = Reference(worksheet, min_col=2, min_row=2, max_row=len(dataframe) + 1, max_col=2)
    categories = Reference(worksheet, min_col=1, min_row=2, max_row=len(dataframe) + 1)

    new_chart = BarChart()

    new_chart.add_data(data)
    new_chart.set_categories(categories)
    new_chart.title = chart_title

    # Set the size of the chart
    new_chart.width = 25  # set width to 25 units
    new_chart.height = 20  # set height to 20 units

    # Copy the style from source to target chart
    copy_chart_style(source_chart, new_chart)

    positions = [0, 74000, 83000, 100000]
    for i, value in enumerate(dataframe.iloc[:, 1], start=0):  # assuming values are in the second column
        color = get_color(value)
        #gsLst = [GradientStop(pos=pos, srgbClr='FFFFFF') if pos == 0 else GradientStop(pos=pos, srgbClr=color) for pos in positions]
        #spPr = GraphicalProperties(gradFill=GradientFillProperties(gsLst=gsLst))
        spPr = GraphicalProperties(solidFill=color)
        dp = DataPoint(idx=i, spPr=spPr)
        new_chart.series[0].dPt.append(dp)

    # print(new_chart.series[0].data_points[0])
    worksheet.add_chart(new_chart, chart_destination)

def read_sav_metadata(file_name: str) -> pd.DataFrame:
    metadata =  pyreadstat.read_sav(
        file_name,
        apply_value_formats=False
    )[1]

    variable_info = pd.DataFrame([metadata.column_names_to_labels, metadata.variable_value_labels])
    variable_info = variable_info.transpose()
    variable_info.index.name = 'name'
    variable_info.columns = ('label', 'values')
    variable_info = variable_info.replace({None: ''})
    variable_info['label'] = variable_info['label'].astype(str)
    variable_info['values'] = variable_info['values'].astype(str)

    return variable_info

def set_time_zone_to_file(zip_file: io.BytesIO, file: str):
    info = zip_file.getinfo(os.path.basename(file))
    current_time = datetime.now(time_zone)
    info.date_time = (
        current_time.year,
        current_time.month,
        current_time.day,
        current_time.hour,
        current_time.minute,
        current_time.second
    )

def format_ws(sheet):
    """
    Adjusts the row height of all rows in the given sheet based on their content.
    """
    max_row = sheet.max_row
    max_column = sheet.max_column
    sheet.column_dimensions["A"].width = 200 / 8.43
    # sheet.row_dimensions[1].height = 100
    # Apply bold formatting to the first two columns
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for cell in [row[0]]:  # First two columns
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Set fixed column width for all columns
    fixed_column_width = 100 / 8.43
    for col in range(2, max_column + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = fixed_column_width

    # Set fixed row height for all rows
    fixed_row_height = 80 / 1.33
    for row in range(1, max_row + 1):
        sheet.row_dimensions[row].height = fixed_row_height

    # Apply the formatting to all cells excluding the first row and first column
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=max_row, max_col=max_column):
        for cell in row:
            if cell.value < 0.01 and cell.value > 0:
                cell.font = Font(color='000BDB')
            elif cell.value >= 0.01 and cell.value < 0.05:
                cell.font = Font(color='FFC179')

            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Clear existing content in the worksheet
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.value = None

# def generate_graph_analysis(df: pd.DataFrame):
#     def corrdot(*args, **kwargs):
#         corr_r = args[0].corr(args[1], 'pearson')
#         corr_text = f"{corr_r:2.2f}".replace("0.", ".")
#         ax = plt.gca()
#         ax.set_axis_off()
#         marker_size = abs(corr_r) * 10000
#         ax.scatter([.5], [.5], marker_size, [corr_r], alpha=0.6, cmap="coolwarm",
#                 vmin=-1, vmax=1, transform=ax.transAxes)
#         font_size = abs(corr_r) * 40 + 5
#         ax.annotate(corr_text, [.5, .5,],  xycoords="axes fraction",
#                     ha='center', va='center', fontsize=font_size)

#     sns.set_theme(style='white', font_scale=1.6)
#     # iris = sns.load_dataset('iris')
#     g = sns.PairGrid(df, aspect=1.4, diag_sharey=False)
#     g.map_lower(sns.regplot, lowess=True, ci=False, line_kws={'color': 'black'})
#     g.map_diag(sns.histplot, line_kws={'color': 'black'}, kde=True)
#     g.map_upper(corrdot)

#     return g

def create_zip(zip_name: str, files: dict):
    temp_dir = tempfile.gettempdir()
    zip_path = os.path.join(temp_dir, zip_name)
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for file_name, file_temp_name in files.items():
            zipf.write(file_temp_name, file_name)
            os.unlink(file_temp_name)
            set_time_zone_to_file(zipf, file_name)
    return zip_path

def upload_to_gcs(source_file_name: str, destination_blob_name: str):
    gcs = CloudStorageClient('connecta-temp-data')
    return gcs.upload_to_gcs(source_file_name, destination_blob_name)

def delete_gcs(blob_name: str):
    gcs = CloudStorageClient('connecta-temp-data')
    gcs.delete_from_gcs(blob_name)

def add_segment_conditions(df: pd.DataFrame, spss_file: BytesIO):
    temp_file_name = get_temp_file(spss_file)
    data , study_metadata = pyreadstat.read_sav(
            temp_file_name,
            apply_value_formats=False
        )
    for i in range(len(df)-1,-1,-1):
        row_original=df.iloc[i]
        condition=row_original[2]
        dict_row={}
        if not condition is None and '#' in condition:
            variables_to_segment=re.findall(r'#(.*?)#',condition)
            for var in variables_to_segment:
                tuples_ref=[]
                flag_total=False
                if re.search(r'\*$',var):
                    var=var[:-1]
                    flag_total=True
                refdict=study_metadata.variable_value_labels[var]
                refs_unique=list(map(int,data[var].dropna().unique()))
                refs_unique.sort(reverse=True)
                for refindex in refs_unique:
                    tuples_ref.append((refindex,refdict[refindex],var))
                if flag_total:
                    tuples_ref.append(("T","TOTAL",var))
                dict_row[var]=tuples_ref

            combinations= list(itertools.product(*dict_row.values()))
            for comb in combinations:
                df = pd.concat([df.iloc[:i+1], pd.DataFrame([row_original]), df.iloc[i+1:]]).reset_index(drop=True)
                row_duplicate=df.iloc[i+1]
                new_condition=condition
                new_name=row_original[0]
                first=True
                for tuple in comb:
                    if first and not row_original[0] is None and not row_original[0] =="":
                        new_name+="_"
                    elif not first:
                        new_name+="_"
                    first=False
                    new_name+=tuple[1]
                    var_to_replace="#"+tuple[2]+"#"
                    if not var_to_replace in new_condition:
                        var_to_replace="#"+tuple[2]+"*#"
                    if str(tuple[0])=="T":
                        new_value_to_var=""
                    else:
                        new_value_to_var="`"+tuple[2]+"`=="+str(tuple[0])
                    new_condition =new_condition.replace(var_to_replace,new_value_to_var)

                if not row_original[4] is None and not row_original[4] =="":
                    new_name+="_"+row_original[4]
                row_duplicate[0]=new_name
                row_duplicate[2]=new_condition
            df=df.drop(i)
    return df

def segment_spss(jobs: pd.DataFrame, spss_file: BytesIO, transform_inverted_scales: bool):
    # print('Started execution')
    temp_file_name = get_temp_file(spss_file)

    survey_data, metadata = pyreadstat.read_sav(
        temp_file_name,
        apply_value_formats=False
    )

    survey_data: pd.DataFrame = survey_data.dropna(how='all')

    if transform_inverted_scales:
        inverted_scales_keywords = get_inverted_scales_keywords()
    else:
        inverted_scales_keywords = []

    files = {}

    jobs_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    jobscommmands=jobs
    jobs = add_segment_conditions(jobs,spss_file)
    with pd.ExcelWriter(jobs_temp_file.name, engine='openpyxl') as writer:
        jobs.to_excel(writer,sheet_name='scenarios', index=False)
        jobscommmands.to_excel(writer, sheet_name='commands', index=False)
    files[f"scenarios_{spss_file.name.split('.')[0]}.xlsx"] = jobs_temp_file.name


    if jobs['cross_variable'].any():

        chi2_xlsx_temp_file = tempfile.NamedTemporaryFile(delete=False)

        # Create a new Excel file
        chi2_xlsx_file_name = f"chi2_{spss_file.name.split('.')[0].replace('Base ', '')}.xlsx"
        chi2_wb = openpyxl.Workbook()
        chi2_wb.remove(chi2_wb.active)

        # Read the existing Excel file
        existing_file = f"static/templates/chart_template.xlsx"
        wb_existing = openpyxl.load_workbook(existing_file)
        ws_existing = wb_existing.active

        # Find the chart in the existing worksheet
        source_chart = None
        for chart_obj in ws_existing._charts:
            if isinstance(chart_obj, BarChart):
                source_chart = chart_obj
                break

        if source_chart is None:
            print("Chart not found in the specified location.")
            exit()

    if jobs['correlation_variables'].any():

        corr_xlsx_temp_file = tempfile.NamedTemporaryFile(delete=False)

        # Create a new Excel file
        corr_xlsx_file_name = f"correlations_{spss_file.name.split('.')[0].replace('Base ', '')}.xlsx"
        corr_wb = openpyxl.Workbook()
        corr_wb.remove(corr_wb.active)

    warning_empty=""

    for _, job in jobs.iterrows():
        if job['variables']:
            if ',' in job['variables']:
                variables = [variable.strip() for variable in job['variables'].split(',')]
            else:
                variables = [variable.strip() for variable in job['variables'].split('\n')]
        else:
            variables = survey_data.columns.to_list()

        if job['condition']:
            try:
                filtered_data = survey_data[variables].query(job['condition'])
            except SyntaxError as e:
                raise e
        else:
            filtered_data = survey_data[variables]

        if filtered_data.empty:
            warning_empty+=(f'Conditions produced an empty database for scenario: {job["scenario_name"]}\n')
            continue

        # Write filtered data to a temporary sav file
        sav_file_name = f"{spss_file.name.split('.')[0].replace('Base ', '')}_{job['scenario_name']}.sav"
        sav_temp_file = tempfile.NamedTemporaryFile(delete=False)

        pyreadstat.write_sav(
            filtered_data,
            sav_temp_file.name,
            column_labels={k: v for k, v in metadata.column_names_to_labels.items() if k in variables},
            variable_value_labels={k: v for k, v in metadata.variable_value_labels.items() if k in variables},
        )

        if job['cross_variable']:
            correction = False
            chi2_df = process_chi2(
                sav_temp_file.name,
                job['cross_variable'],
                job['chi2_mode'],
                inverted_scales_keywords,
                correction,
            )

            # Create a new workbook object and select the active worksheet
            chi2_wb.create_sheet(job['scenario_name'])
            chi2_ws = chi2_wb[job['scenario_name']]

            # Write DataFrame content to the worksheet
            for r in dataframe_to_rows(chi2_df, index=False, header=True):
                chi2_ws.append(r)

            # Create chart in the worksheet
            create_chart(chi2_ws, source_chart, chi2_df, 'Intención de compra', "E3")

        if job['correlation_variables']:
            if ',' in job['variables']:
                corr_variables = [variable.strip() for variable in job['correlation_variables'].split(',')]
            else:
                corr_variables = [variable.strip() for variable in job['correlation_variables'].split('\n')]

            correlation_data = get_correlation_data(sav_temp_file.name, corr_variables)
            correlation_df, p_value_df = process_correlation(correlation_data)

            correlation_df = correlation_df.rename(
                columns=metadata.column_names_to_labels,
                index=metadata.column_names_to_labels
            ).reset_index(names='')

            p_value_df = p_value_df.rename(
                columns=metadata.column_names_to_labels,
                index=metadata.column_names_to_labels
            ).reset_index(names='')

            # Create a new workbook object and select the active worksheet
            corr_wb.create_sheet(job['scenario_name'])
            corr_ws = corr_wb[job['scenario_name']]

            # Write DataFrame content to the worksheet
            for r in dataframe_to_rows(p_value_df, index=False, header=True):
                corr_ws.append(r)

            format_ws(corr_ws)

            start_row = 1
            for r_idx, r in enumerate(dataframe_to_rows(correlation_df, index=False, header=True), start=start_row):
                for c_idx, value in enumerate(r, start=1):
                    corr_ws.cell(row=r_idx, column=c_idx, value=value)

            # pdf_graph_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')

            # graph = generate_graph_analysis(correlation_data)
            # graph.savefig(pdf_graph_temp_file.name)

            # zip_file.write(pdf_graph_temp_file.name, arcname=f'{job["scenario_name"]}_correlation_statistical_graph.pdf')

        # Add the temporary sav file to the zip file with custom arcname
        files[sav_file_name] = sav_temp_file.name

        # Close and delete the temporary sav file
        sav_temp_file.close()

    if jobs['cross_variable'].any():
        # Save the new Excel file
        chi2_wb.save(chi2_xlsx_temp_file)
        chi2_wb.close()

        files[chi2_xlsx_file_name] = chi2_xlsx_temp_file.name
        chi2_xlsx_temp_file.close()

    if jobs['correlation_variables'].any():
        # Save the new Excel file
        corr_wb.save(corr_xlsx_temp_file)
        corr_wb.close()

        files[corr_xlsx_file_name] = corr_xlsx_temp_file.name
        corr_xlsx_temp_file.close()

    return files, warning_empty
