from collections import Counter, defaultdict
from io import BytesIO
import re
import pandas as pd
from copy import copy
from unidecode import unidecode
from itertools import combinations, permutations

from openpyxl import Workbook, load_workbook
from difflib import SequenceMatcher
from openpyxl.utils import get_column_letter, range_boundaries,quote_sheetname
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.cell.cell import MergedCell

from app.modules.segment_spss import get_temp_file
from app.modules.text_function import processSavMulti
from app.modules.utils import write_temp_excel
from app.modules.processing import apply_red_and_blue_color_to_letter,calculate_differences,processing

import pyreadstat


def getPreProcessCode(spss_file: BytesIO, xlsx_file: BytesIO,xlsx_file_LC: BytesIO):
    try:
        file_xlsx = get_temp_file(xlsx_file)
        inverseVarsList = pd.read_excel(
            file_xlsx, usecols="A,E", skiprows=3, names=["vars", "inverses"]
        ).dropna()
        inverseVarsList = inverseVarsList[inverseVarsList["inverses"] == "I"].iloc[:, 0]
        scaleVarsList = pd.read_excel(
            file_xlsx, usecols="A,D", skiprows=3, names=["vars", "scale"]
        ).dropna()

        preprocesscode = ""
        preprocesscode += processSavMulti(spss_file)[1] + processSavMulti(spss_file)[0]
        preprocesscode += getGroupCreateMultisCode(spss_file)
        if not inverseVarsList.empty:
            preprocesscode += getInverseCodeVars(spss_file, inverseVarsList)
        if not scaleVarsList.empty:
            preprocesscode += getScaleCodeVars(spss_file, scaleVarsList)
        preprocesscode += "\nCOMPUTE TOTAL=1.\nVARIABLE LABELS TOTAL 'TOTAL'.\nVALUE LABELS TOTAL 1 \"TOTAL\".\nEXECUTE.\n"
        preprocesscode += getCloneCodeVars(spss_file, xlsx_file)
        preprocesscode += getPreProcessAbiertas(spss_file, xlsx_file,xlsx_file_LC)
        return preprocesscode
    except Exception:
        return "Error with plantilla hwen try to get preprocess code"


def checkPreProcessCodeUnique(spss_file: BytesIO, xlsx_file: BytesIO):
    file_xlsx = get_temp_file(xlsx_file)
    inverseVarsList = pd.read_excel(
        file_xlsx, usecols="A,E", skiprows=3, names=["vars", "inverses"]
    ).dropna()
    inverseVarsList = inverseVarsList[inverseVarsList["inverses"] == "I"].iloc[:, 0]
    scaleVarsList = pd.read_excel(
        file_xlsx, usecols="A,D", skiprows=3, names=["vars", "scale"]
    ).dropna()
    flag1 = False
    flag2 = False
    if not inverseVarsList.empty:
        flag1 = checkInverseCodeVars(spss_file, inverseVarsList)
    if not scaleVarsList.empty:
        flag2 = getScaleCodeVars(spss_file, scaleVarsList) != ""
    return flag1, flag2


def getPreProcessCode2(spss_file: BytesIO):
    preprocesscode = processSavMulti(spss_file)[1] + processSavMulti(spss_file)[0]
    return preprocesscode


def getProcessCode2(
    spss_file: BytesIO, xlsx_file: BytesIO, xlsx_file_LC: BytesIO, checkinclude=False, rutaarchivo=""
):
    result, warning2 = getProcessCode(spss_file, xlsx_file, xlsx_file_LC, checkinclude)
    file_xlsx = get_temp_file(xlsx_file)
    nombrehoja = (
        pd.read_excel(file_xlsx, usecols="O", skiprows=3, names=["name"])
        .dropna()
        .iloc[0, 0]
    )
    try:
        sufijo = (
            pd.read_excel(file_xlsx, usecols="P", skiprows=3, names=["name"])
            .dropna()
            .iloc[0, 0]
        )
        sufijo = " " + str(sufijo)
    except Exception:
        sufijo = ""
    warning = ""
    if warning2 != "":
        warning += "/" + nombrehoja + "/ " + warning2
    result += (
        "\nOUTPUT MODIFY\n  /SELECT ALL EXCEPT (TABLES)\n  /DELETEOBJECT DELETE = YES."
    )
    result += (
        "\nOUTPUT EXPORT\n  /CONTENTS  EXPORT=VISIBLE  LAYERS=VISIBLE  MODELVIEWS=PRINTSETTING\n  /XLSX  DOCUMENTFILE='"
        + rutaarchivo
        + "'\n     OPERATION=CREATESHEET  SHEET='"
        + nombrehoja
        + sufijo
        + "'\n     LOCATION=LASTCOLUMN  NOTESCAPTIONS=NO.\n"
        + "OUTPUT CLOSE NAME=*.\nEXECUTE.\n"
    )
    temp_file_name = get_temp_file(spss_file)
    data, study_metadata = pyreadstat.read_sav(
        temp_file_name, apply_value_formats=False
    )
    result += (
        "\n*___TOTAL____________________________________________________________________________\n ____________________________________________________________________________________\n ______"
        + nombrehoja
        + sufijo
        + "______________________________________________________________________________.\n"
    )
    varsList = (
        pd.read_excel(file_xlsx, usecols="M", skiprows=3, names=["varsSegment"])
        .dropna()["varsSegment"]
        .tolist()
    )
    varsList_segment = (
        pd.read_excel(file_xlsx, usecols="N", skiprows=3, names=["vars_Segment"])
        .dropna()["vars_Segment"]
        .tolist()
    )

    if not varsList_segment:
        for var in varsList:
            refdict = study_metadata.variable_value_labels[var]
            refs_unique = data[var].dropna().unique()
            refs_unique.sort()
            for refindex in refs_unique:
                name_var = limpiar_texto(refdict[refindex])
                name_dataset = name_var
                name_sheet = (
                    nombrehoja
                    + " "
                    + unidecode(refdict[refindex]).replace(".", "")
                    + sufijo
                )
                if len(name_sheet) > 30:
                    name_sheet = (
                        nombrehoja
                        + " "
                        + unidecode(refdict[refindex]).replace(".", "")[:10]
                        + sufijo
                    )
                result += "DATASET ACTIVATE REF_" + name_dataset + ".\n"
                condition = data[var] == refindex
                result_preg, _ = getProcessCode(
                    spss_file, xlsx_file, xlsx_file_LC, checkinclude, condition=condition
                )
                result += result_preg
                result += "\nOUTPUT MODIFY\n  /SELECT ALL EXCEPT (TABLES)\n  /DELETEOBJECT DELETE = YES."
                result += (
                    "\nOUTPUT EXPORT\n  /CONTENTS  EXPORT=VISIBLE  LAYERS=VISIBLE  MODELVIEWS=PRINTSETTING\n  /XLSX  DOCUMENTFILE='"
                    + rutaarchivo
                    + "'\n     OPERATION=CREATESHEET  SHEET='"
                    + name_sheet
                    + "'\n     LOCATION=LASTCOLUMN  NOTESCAPTIONS=NO.\n"
                    + "OUTPUT CLOSE NAME=*.\n"
                )
                result += "DATASET CLOSE REF_" + name_dataset + ".\n"
            result += (
                "\n*____________________________________________________________________________________\n ______"
                + var
                + "______________________________________________________________________________\n ______"
                + nombrehoja
                + sufijo
                + "______________________________________________________________________________.\nEXECUTE.\n"
            )
    else:
        for var_segment in varsList_segment:
            refdict_segment = study_metadata.variable_value_labels[var_segment]
            refs_segment_unique = data[var_segment].dropna().unique()
            refs_segment_unique.sort()
            for refindex_segment in refs_segment_unique:
                name_var_segment = limpiar_texto(refdict_segment[refindex_segment])
                for var in varsList:
                    refdict = study_metadata.variable_value_labels[var]
                    refs_unique = data[var].dropna().unique()
                    refs_unique.sort()
                    for refindex in refs_unique:
                        name_var = limpiar_texto(refdict[refindex])
                        name_dataset = name_var + "_" + name_var_segment
                        name_sheet = (
                            nombrehoja
                            + " "
                            + unidecode(refdict[refindex]).replace(".", "")
                            + " "
                            + unidecode(refdict_segment[refindex_segment]).replace(".", "")
                            + sufijo
                        )
                        if len(name_sheet) > 30:
                            name_sheet = (
                                nombrehoja
                                + " "
                                + unidecode(refdict[refindex]).replace(".", "")[:10]
                                + " "
                                + unidecode(refdict_segment[refindex_segment]).replace(".", "")[:10]
                                + sufijo
                            )
                        result += "DATASET ACTIVATE REF_" + name_dataset + ".\n"
                        condition1 = data[var] == refindex
                        condition2 = data[var_segment] == refindex_segment
                        condition = condition1 & condition2
                        result_preg, _ = getProcessCode(
                            spss_file, xlsx_file, xlsx_file_LC, checkinclude, condition=condition
                        )
                        result += result_preg
                        result += "\nOUTPUT MODIFY\n  /SELECT ALL EXCEPT (TABLES)\n  /DELETEOBJECT DELETE = YES."
                        result += (
                            "\nOUTPUT EXPORT\n  /CONTENTS  EXPORT=VISIBLE  LAYERS=VISIBLE  MODELVIEWS=PRINTSETTING\n  /XLSX  DOCUMENTFILE='"
                            + rutaarchivo
                            + "'\n     OPERATION=CREATESHEET  SHEET='"
                            + name_sheet
                            + "'\n     LOCATION=LASTCOLUMN  NOTESCAPTIONS=NO.\n"
                            + "OUTPUT CLOSE NAME=*.\n"
                        )
                        result += "DATASET CLOSE REF_" + name_dataset + ".\n"
                    result += (
                        "\n*____________________________________________________________________________________"
                        + "\n ______"
                        + var
                        + "______________________________________________________________________________"
                        + "\n ______"
                        + nombrehoja
                        + sufijo
                        + "_"
                        + name_var_segment
                        + "______________________________________________________________________________.\nEXECUTE.\n"
                    )

    result += """*
                ⠸⣷⣦⠤⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣀⣠⣤⠀⠀⠀
                ⠀⠙⣿⡄⠈⠑⢄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣀⠔⠊⠉⣿⡿⠁⠀⠀⠀
                ⠀⠀⠈⠣⡀⠀⠀⠑⢄⠀⠀⠀⠀⠀⠀⠀⠀⠀⡠⠊⠁⠀⠀⣰⠟⠀⠀⠀⠀⠀
                ⠀⠀⠀⠀⠈⠢⣄⠀⡈⠒⠊⠉⠁⠀⠈⠉⠑⠚⠀⠀⣀⠔⢊⣠⠤⠒⠊⡽⠀⠀
                ⠀⠀⠀⠀⠀⠀⠀⡽⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠩⡔⠊⠁⠀⠀⠀⠀⠀⠀⠇
                ⠀⠀⠀⠀⠀⠀⠀⡇⢠⡤⢄⠀⠀⠀⠀⠀⡠⢤⣄⠀⡇⠀⠀⠀⠀⠀⠀⠀⢰⠀
                ⠀⠀⠀⠀⠀⠀⢀⠇⠹⠿⠟⠀⠀⠤⠀⠀⠻⠿⠟⠀⣇⠀⠀⡀⠠⠄⠒⠊⠁⠀
                ⠀⠀⠀⠀⠀⠀⢸⣿⣿⡆⠀⠰⠤⠖⠦⠴⠀⢀⣶⣿⣿⠀⠙⢄⠀⠀⠀⠀⠀⠀
                ⠀⠀⠀⠀⠀⠀⠀⢻⣿⠃⠀⠀⠀⠀⠀⠀⠀⠈⠿⡿⠛⢄⠀⠀⠱⣄⠀⠀⠀⠀
                ⠀⠀⠀⠀⠀⠀⠀⢸⠈⠓⠦⠀⣀⣀⣀⠀⡠⠴⠊⠹⡞⣁⠤⠒⠉⠀⠀⠀⠀⠀
                ⠀⠀⠀⠀⠀⠀⣠⠃⠀⠀⠀⠀⡌⠉⠉⡤⠀⠀⠀⠀⢻⠿⠆⠀⠀⠀⠀⠀⠀⠀
                ⠀⠀⠀⠀⠀⠰⠁⡀⠀⠀⠀⠀⢸⠀⢰⠃⠀⠀⠀⢠⠀⢣⠀⠀⠀⠀⠀⠀⠀⠀
                ⠀⠀⠀⢶⣗⠧⡀⢳⠀⠀⠀⠀⢸⣀⣸⠀⠀⠀⢀⡜⠀⣸⢤⣶⠀⠀⠀⠀⠀⠀
                ⠀⠀⠀⠈⠻⣿⣦⣈⣧⡀⠀⠀⢸⣿⣿⠀⠀⢀⣼⡀⣨⣿⡿⠁⠀⠀⠀⠀⠀⠀
                ⠀⠀⠀⠀⠀⠈⠻⠿⠿⠓⠄⠤⠘⠉⠙⠤⢀⠾⠿⣿⠟⠋⠀⠀
                        ░░░░░░░░║░░░
                        ░░░╔╗╦╗╔╣░░░
                        ░░░╠╝║║║║░░░
                        ░░░╚╝╝╚╚╩░░░
                        ░░░░░░░░░░░░
                ."""
    return result, warning


def getProcessCode(
    spss_file: BytesIO, xlsx_file: BytesIO, xlsx_file_LC: BytesIO, checkinclude=False, condition=None
):
    file_xlsx = get_temp_file(xlsx_file)
    varsList = pd.read_excel(
        file_xlsx,
        usecols="A,B,D,E",
        skiprows=3,
        names=["vars", "varsTypes", "Scales", "descendOrder"],
    ).dropna(subset=["vars"])
    colVarsList = pd.melt(
        pd.read_excel(file_xlsx, nrows=2), var_name="colVars", value_name="colVarsNames"
    ).drop(0)
    result = ""
    warning = ""
    colvars = colVarsList.iloc[:, 0]
    temp_file_name = get_temp_file(spss_file)
    data, study_metadata = pyreadstat.read_sav(
        temp_file_name, apply_value_formats=False
    )
    for i in range(len(colvars)):
        var = colvars[i + 1]
        if re.search("^[PFSV].*[1-90].*A", var):
            colvars[i + 1] = "$COL_" + re.search(".*A", var).group()[:-1]
        else:
            colvars[i + 1] = "COL_" + var

    for i in range(len(varsList)):
        if varsList.iloc[i][1] != "A":
            if varsList.iloc[i][3] != "D" and varsList.iloc[i][3] != "P":
                result += writeQuestion(
                    varsList.iloc[i][0],
                    varsList.iloc[i][1],
                    colvars,
                    includeall=checkinclude,
                )
            elif varsList.iloc[i][3] == "P":
                result += writeQuestion(
                    varsList.iloc[i][0],
                    varsList.iloc[i][1],
                    colvars,
                    includeall=checkinclude,
                    custom_order=varsList.iloc[i][2],
                )
            else:
                result += writeQuestion(
                    varsList.iloc[i][0],
                    varsList.iloc[i][1],
                    colvars,
                    descendingorder=True,
                    includeall=checkinclude,
                )
            if (
                not pd.isnull(varsList.iloc[i][2])
                and not pd.isnull(varsList.iloc[i][3])
                and varsList.iloc[i][3] not in ["I", "D"]
            ):
                varlabeloriginal = ""
                for var, label in study_metadata.column_names_to_labels.items():
                    if var == varsList.iloc[i][0]:
                        varlabeloriginal = label
                for tipo in str(varsList.iloc[i][3]).split():
                    if tipo == "T2B":
                        result += (
                            "\nDELETE VARIABLES "
                            + varsList.iloc[i][2]
                            + tipo
                            + ".\nEXECUTE."
                        )
                        result += (
                            "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            + varsList.iloc[i][2]
                            + '\n/OPTIONS FIX="'
                            + tipo
                            + '" FIXTYPE=SUFFIX ACTION=RUN.\nEXECUTE.'
                        )
                        result += (
                            "\nVALUE LABELS "
                            + varsList.iloc[i][2]
                            + tipo
                            + ' 1 "'
                            + tipo
                            + '".'
                        )
                        result += (
                            "\nRECODE "
                            + varsList.iloc[i][2]
                            + tipo
                            + " (5=1) (4=1) (3=SYSMIS) (2=SYSMIS) (1=SYSMIS).\nEXECUTE.\n"
                        )
                        result += (
                            "\nVARIABLE LABELS "
                            + varsList.iloc[i][0]
                            + ' "'
                            + varlabeloriginal
                            + " - "
                            + tipo
                            + '".\nEXECUTE.\n'
                        )
                        result += writeQuestion(
                            varsList.iloc[i][0],
                            varsList.iloc[i][1],
                            colvars,
                            includeall=checkinclude,
                            varanidada=varsList.iloc[i][2] + tipo,
                        )
                    elif tipo == "TB":
                        result += (
                            "\nDELETE VARIABLES "
                            + varsList.iloc[i][2]
                            + tipo
                            + ".\nEXECUTE."
                        )
                        result += (
                            "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            + varsList.iloc[i][2]
                            + '\n/OPTIONS FIX="'
                            + tipo
                            + '" FIXTYPE=SUFFIX ACTION=RUN.\nEXECUTE.'
                        )
                        result += (
                            "\nVALUE LABELS "
                            + varsList.iloc[i][2]
                            + tipo
                            + ' 1 "'
                            + tipo
                            + '".'
                        )
                        result += (
                            "\nRECODE "
                            + varsList.iloc[i][2]
                            + tipo
                            + " (5=1) (4=SYSMIS) (3=SYSMIS) (2=SYSMIS) (1=SYSMIS).\nEXECUTE.\n"
                        )
                        result += (
                            "\nVARIABLE LABELS "
                            + varsList.iloc[i][0]
                            + ' "'
                            + varlabeloriginal
                            + " - "
                            + tipo
                            + '".'
                        )
                        result += writeQuestion(
                            varsList.iloc[i][0],
                            varsList.iloc[i][1],
                            colvars,
                            includeall=checkinclude,
                            varanidada=varsList.iloc[i][2] + tipo,
                        )
                    elif tipo == "MB":
                        result += (
                            "\nDELETE VARIABLES "
                            + varsList.iloc[i][2]
                            + tipo
                            + ".\nEXECUTE."
                        )
                        result += (
                            "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            + varsList.iloc[i][2]
                            + '\n/OPTIONS FIX="'
                            + tipo
                            + '" FIXTYPE=SUFFIX ACTION=RUN.\nEXECUTE.'
                        )
                        result += (
                            "\nVALUE LABELS "
                            + varsList.iloc[i][2]
                            + tipo
                            + ' 1 "'
                            + tipo
                            + '".'
                        )
                        result += (
                            "\nRECODE "
                            + varsList.iloc[i][2]
                            + tipo
                            + " (5=SYSMIS) (4=SYSMIS) (3=1) (2=SYSMIS) (1=SYSMIS).\nEXECUTE.\n"
                        )
                        result += (
                            "\nVARIABLE LABELS "
                            + varsList.iloc[i][0]
                            + ' "'
                            + varlabeloriginal
                            + " - "
                            + tipo
                            + '".'
                        )
                        result += writeQuestion(
                            varsList.iloc[i][0],
                            varsList.iloc[i][1],
                            colvars,
                            includeall=checkinclude,
                            varanidada=varsList.iloc[i][2] + tipo,
                        )
                    elif tipo == "B2B":
                        result += (
                            "\nDELETE VARIABLES "
                            + varsList.iloc[i][2]
                            + tipo
                            + ".\nEXECUTE."
                        )
                        result += (
                            "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            + varsList.iloc[i][2]
                            + '\n/OPTIONS FIX="'
                            + tipo
                            + '" FIXTYPE=SUFFIX ACTION=RUN.\nEXECUTE.'
                        )
                        result += (
                            "\nVALUE LABELS "
                            + varsList.iloc[i][2]
                            + tipo
                            + ' 1 "'
                            + tipo
                            + '".'
                        )
                        result += (
                            "\nRECODE "
                            + varsList.iloc[i][2]
                            + tipo
                            + " (5=SYSMIS) (4=SYSMIS) (3=SYSMIS) (2=1) (1=1).\nEXECUTE.\n"
                        )
                        result += (
                            "\nVARIABLE LABELS "
                            + varsList.iloc[i][0]
                            + ' "'
                            + varlabeloriginal
                            + " - "
                            + tipo
                            + '".'
                        )
                        result += writeQuestion(
                            varsList.iloc[i][0],
                            varsList.iloc[i][1],
                            colvars,
                            includeall=checkinclude,
                            varanidada=varsList.iloc[i][2] + tipo,
                        )
                    elif tipo == "B3B":
                        result += (
                            "\nDELETE VARIABLES "
                            + varsList.iloc[i][2]
                            + tipo
                            + ".\nEXECUTE."
                        )
                        result += (
                            "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            + varsList.iloc[i][2]
                            + '\n/OPTIONS FIX="'
                            + tipo
                            + '" FIXTYPE=SUFFIX ACTION=RUN.\nEXECUTE.'
                        )
                        result += (
                            "\nVALUE LABELS "
                            + varsList.iloc[i][2]
                            + tipo
                            + ' 1 "'
                            + tipo
                            + '".'
                        )
                        result += (
                            "\nRECODE "
                            + varsList.iloc[i][2]
                            + tipo
                            + " (5=SYSMIS) (4=SYSMIS) (3=1) (2=1) (1=1).\nEXECUTE.\n"
                        )
                        result += (
                            "\nVARIABLE LABELS "
                            + varsList.iloc[i][0]
                            + ' "'
                            + varlabeloriginal
                            + " - "
                            + tipo
                            + '".'
                        )
                        result += writeQuestion(
                            varsList.iloc[i][0],
                            varsList.iloc[i][1],
                            colvars,
                            includeall=checkinclude,
                            varanidada=varsList.iloc[i][2] + tipo,
                        )
                result += (
                    "\nVARIABLE LABELS "
                    + varsList.iloc[i][0]
                    + ' "'
                    + varlabeloriginal
                    + '".'
                )
        elif varsList.iloc[i][1] == "A":
            result_abierta, result_warning = getProcessAbiertas(
                spss_file,
                xlsx_file,
                xlsx_file_LC,
                checkinclude,
                varsList.iloc[i][0],
                condition=condition,
            )
            result += result_abierta
            warning += result_warning
    return result, warning

def getWarning(
    spss_file: BytesIO, xlsx_file: BytesIO, xlsx_file_LC: BytesIO, checkinclude=False, condition=None
):
    file_xlsx = get_temp_file(xlsx_file)
    varsList = pd.read_excel(
        file_xlsx,
        usecols="A,B,D,E",
        skiprows=3,
        names=["vars", "varsTypes", "Scales", "descendOrder"],
    ).dropna(subset=["vars"])
    warning = ""

    for i in range(len(varsList)):
        if varsList.iloc[i][1] == "A":
            _, result_warning = getProcessAbiertas(
                spss_file,
                xlsx_file,
                xlsx_file_LC,
                checkinclude,
                varsList.iloc[i][0],
                condition=condition,
            )
            warning += result_warning
    return warning


def getPreProcessAbiertas(spss_file: BytesIO, xlsx_file: BytesIO, xlsx_file_LC: BytesIO):
    result = ""
    temp_file_name = get_temp_file(spss_file)
    data, study_metadata = pyreadstat.read_sav(
        temp_file_name, apply_value_formats=False
    )
    file_xlsx = get_temp_file(xlsx_file)
    file_xlsx_LC = get_temp_file(xlsx_file_LC)
    varsList = pd.read_excel(
        file_xlsx, usecols="A,C", skiprows=3, names=["vars", "sheetNames"]
    ).dropna()
    for i in range(len(varsList)):
        lcTable = pd.read_excel(
            file_xlsx_LC,
            sheet_name=varsList.iloc[i][1],
            usecols="A,B",
            skiprows=1,
            names=["vars", "sheetNames"],
        )
        varAbierta = varsList.iloc[i][0]

        prefix = re.search("^[PFSV].*[1-90].*A", varAbierta).group()
        multis = []
        variab = ""

        labelvar = ""
        for var, label in study_metadata.column_names_to_labels.items():
            if re.search("^[PFSV].*[1-90].*A", var):
                if re.search(".*A", var).group() == prefix:
                    multis.append(var)
                    if labelvar == "":
                        labelvar = label
        delmultis = []
        for multi in multis:
            try:
                if (
                    study_metadata.column_names_to_labels[multi]
                    .lower()
                    .startswith("otro")
                ):
                    delmultis.append(multi)
            except Exception:
                continue
        for multi2 in delmultis:
            multis.remove(multi2)
        for i in range(len(multis)):
            variab += multis[i] + " "
        result += getAbiertasPreCode(variab, lcTable)
        if len(multis) > 1:
            result += writeAgrupMulti(prefix, multis, labelvar)
    return result


def getAbiertasPreCode(var, lcTable):
    abiertascode = ""
    principal = ""
    subcodes = []
    options = []
    for i in range(len(lcTable)):
        if str(lcTable.iloc[i][0]).strip() == "NETO":
            continue
        if lcTable.isnull().iloc[i][1]:
            subcodes.append(lcTable.iloc[i][0])
        else:
            if subcodes:
                abiertascode += "\nRECODE " + var
                for cod in subcodes:
                    abiertascode += " (" + str(cod) + "=" + str(principal) + ")"
                abiertascode += "."
                subcodes = []
            principal = lcTable.iloc[i][0]
            options.append((principal, lcTable.iloc[i][1]))
    abiertascode += "\nVALUE LABELS " + var
    for num, option in options:
        abiertascode += "\n" + str(num).strip() + ' "' + str(option).strip() + '"'
    abiertascode += ".\nEXECUTE.\n"
    return abiertascode


def getProcessAbiertas(
    spss_file: BytesIO,
    xlsx_file: BytesIO,
    xlsx_file_LC: BytesIO,
    checkinclude=False,
    namevar="",
    condition=None,
):
    result = ""
    warning = ""
    temp_file_name = get_temp_file(spss_file)
    data2, study_metadata = pyreadstat.read_sav(
        temp_file_name, apply_value_formats=False
    )

    if condition is None:
        data = data2
    else:
        data = data2[condition]

    file_xlsx = get_temp_file(xlsx_file)
    file_xlsx_LC = get_temp_file(xlsx_file_LC)
    varsList = pd.read_excel(
        file_xlsx,
        usecols="A,C,D,E",
        skiprows=3,
        names=["vars", "sheetNames", "varanidada", "typesegment"],
    ).dropna(subset=["sheetNames"])
    colVarsList = pd.melt(
        pd.read_excel(file_xlsx, nrows=2), var_name="colVars", value_name="colVarsNames"
    ).drop(0)
    colvars = colVarsList.iloc[:, 0]
    for i in range(len(colvars)):
        var = colvars[i + 1]
        if re.search("^[PFSV].*[1-90].*A", var):
            colvars[i + 1] = "$COL_" + re.search(".*A", var).group()[:-1]
        else:
            colvars[i + 1] = "COL_" + var
    for i in range(len(varsList)):
        if namevar == "" or namevar == varsList.iloc[i][0]:
            lcTable = pd.read_excel(
                file_xlsx_LC,
                sheet_name=varsList.iloc[i][1],
                usecols="A,B",
                skiprows=1,
                names=["vars", "sheetNames"],
            ).dropna()
            varAbierta = varsList.iloc[i][0]
            varlabeloriginal = ""

            prefix = re.search("^[PFSV].*[1-90].*A", varAbierta).group()
            multis = []
            for var, label in study_metadata.column_names_to_labels.items():
                if var == varAbierta:
                    varlabeloriginal = label
                if re.search("^[PFSV].*[1-90].*A", var):
                    if re.search(".*A", var).group() == prefix:
                        multis.append(var)
            listNetos = []
            parNeto = []
            if lcTable.iloc[0][0] != "NETO":
                parNeto = ["First", []]
            lista_codigos = []
            lista_final_codigos = []
            for j in range(len(lcTable)):
                if lcTable.iloc[j][0] == "NETO":
                    if parNeto != []:
                        listNetos.append(parNeto)
                    parNeto = [lcTable.iloc[j][1], []]
                elif lcTable.iloc[j][0] == 95:
                    if parNeto != []:
                        listNetos.append(parNeto)
                    parNeto = ["End", [95]]
                else:
                    parNeto[1].append(lcTable.iloc[j][0])
                lista_final_codigos.append(lcTable.iloc[j][0])

            lcTable2 = pd.read_excel(
                file_xlsx_LC,
                sheet_name=varsList.iloc[i][1],
                usecols="A,B",
                skiprows=1,
                names=["vars", "sheetNames"],
            )
            for j in range(len(lcTable2)):
                lista_codigos.append(lcTable2.iloc[j][0])

            if parNeto != []:
                listNetos.append(parNeto)
            listatotal = []

            for var in multis:
                listatotal += data[var].dropna().tolist()
            count = Counter(listatotal)
            listafinalorde = []
            num = 1
            for net in listNetos:
                if net[0] != "First" and net[0] != "End":
                    if any(count[ele] > 0 for ele in net[1]):
                        listafinalorde.append(990 + num)
                        result += "\nADD VALUE LABEL "
                        for multivar in multis:
                            result += multivar + " "
                        result += str(990 + num) + ' "NETO ' + net[0] + '".\nEXECUTE.\n'
                    num += 1
                if net[0] != "End":
                    for in1 in count.most_common():
                        if in1[0] in net[1]:
                            listafinalorde.append(int(in1[0]))
                else:
                    for end in net[1]:
                        if count[end] > 0:
                            listafinalorde.append(end)
            result += writeAbiertasQuestion(
                varAbierta, colvars, listafinalorde, includeall=checkinclude
            )

            flag_preginclude = True
            flag_preginclude2 = True

            for cod in count:
                if cod not in lista_final_codigos and flag_preginclude2:
                    warning += "#Code without label " + prefix + "# "
                    flag_preginclude2 = False
                if cod not in lista_codigos:
                    if flag_preginclude:
                        warning += "\nCode Missing " + prefix + " : "
                        flag_preginclude = False
                    warning += str(cod) + " | "

            listatotaluniq = list(set(listatotal))
            for net in listNetos:
                if (
                    net[0] != "First"
                    and net[0] != "End"
                    and any(count[ele] > 0 for ele in net[1])
                ):
                    nombreneto = net[0].strip().replace(" ", "_") + "_" + varAbierta
                    result += "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                    for col in multis:
                        result += col + " "
                    result += '\n/OPTIONS FIX="NETO_" FIXTYPE=PREFIX ACTION=RUN.\n'
                    newmultis = ("NETO_" + variablee for variablee in multis)
                    result += "RECODE "
                    for newvar in newmultis:
                        result += newvar + " "
                    result += "\n(SYSMIS=0)"
                    for num in listatotaluniq:
                        if num in net[1]:
                            result += " (" + str(int(num)) + "=1)"
                        else:
                            result += " (" + str(int(num)) + "=0)"
                    result += ".\nEXECUTE.\n\nCOMPUTE NETO_" + nombreneto + "="
                    newmultis = ("NETO_" + variablee for variablee in multis)
                    for col in newmultis:
                        result += col + "+"
                    result = result[:-1] + ".\nRECODE NETO_" + nombreneto
                    for inde in range(len(multis)):
                        result += " (" + str(inde + 1) + "=1)"
                    result += ".\nEXECUTE.\nDELETE VARIABLES"
                    newmultis = ("NETO_" + variablee for variablee in multis)
                    for newvar in newmultis:
                        result += " " + newvar
                    result += (
                        ".\nEXECUTE.\n\nformats NETO_"
                        + nombreneto
                        + "(f8.0).\nVALUE LABELS NETO_"
                        + nombreneto
                        + ' 1 "NETO '
                        + net[0].strip()
                        + '".\nEXECUTE.\n'
                    )
                    result+=f"TEMPORARY.\nFILTER BY {multis[0]}.\n"
                    result += writeQuestion(
                        "NETO_" + nombreneto, "T", colvars, includeall=checkinclude
                    )

            if varsList.iloc[i][2] and varsList.iloc[i][3]:
                prefix = re.search("^[PFSV].*[1-90].*A", varAbierta).group()
                for tipo in str(varsList.iloc[i][3]).split():
                    if tipo == "T2B":
                        result += (
                            "\nDELETE VARIABLES "
                            + varsList.iloc[i][2]
                            + tipo
                            + ".\nEXECUTE."
                        )
                        result += (
                            "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            + varsList.iloc[i][2]
                            + '\n/OPTIONS FIX="'
                            + tipo
                            + '" FIXTYPE=SUFFIX ACTION=RUN.\nEXECUTE.'
                        )
                        result += (
                            "\nVALUE LABELS "
                            + varsList.iloc[i][2]
                            + tipo
                            + ' 1 "'
                            + tipo
                            + '".'
                        )
                        result += (
                            "\nRECODE "
                            + varsList.iloc[i][2]
                            + tipo
                            + " (5=1) (4=1) (3=SYSMIS) (2=SYSMIS) (1=SYSMIS).\nEXECUTE.\n"
                        )
                        result += writeAgrupMulti(
                            prefix, multis, varlabeloriginal + " - " + tipo
                        )
                        condition1 = data[varsList.iloc[i][2]] == 5
                        condition2 = data[varsList.iloc[i][2]] == 4
                        filtro = condition1 | condition2
                        result += writeAbiertasQuestion(
                            varAbierta,
                            colvars,
                            getListOrderConditions(multis, data, listNetos, filtro),
                            includeall=checkinclude,
                            varanidada=varsList.iloc[i][2] + tipo,
                        )
                        listatotal = []
                        for var in multis:
                            listatotal += data[filtro][var].tolist()
                        count2 = Counter(listatotal)
                        listatotaluniq = list(set(listatotal))
                        for net in listNetos:
                            if (
                                net[0] != "First"
                                and net[0] != "End"
                                and any(count2[ele] > 0 for ele in net[1])
                            ):
                                nombreneto = (
                                    net[0].strip().replace(" ", "_") + "_" + varAbierta
                                )
                                result+=f"TEMPORARY.\nSELECT IF (nvalid({multis[0]})=1 and {varsList.iloc[i][2] + tipo}=1).\n"
                                result += writeQuestion(
                                    "NETO_" + nombreneto,
                                    "T",
                                    colvars,
                                    includeall=checkinclude,
                                    varanidada=varsList.iloc[i][2] + tipo,
                                )
                    elif tipo == "TB":
                        result += (
                            "\nDELETE VARIABLES "
                            + varsList.iloc[i][2]
                            + tipo
                            + ".\nEXECUTE."
                        )
                        result += (
                            "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            + varsList.iloc[i][2]
                            + '\n/OPTIONS FIX="'
                            + tipo
                            + '" FIXTYPE=SUFFIX ACTION=RUN.\nEXECUTE.'
                        )
                        result += (
                            "\nVALUE LABELS "
                            + varsList.iloc[i][2]
                            + tipo
                            + ' 1 "'
                            + tipo
                            + '".'
                        )
                        result += (
                            "\nRECODE "
                            + varsList.iloc[i][2]
                            + tipo
                            + " (5=1) (4=SYSMIS) (3=SYSMIS) (2=SYSMIS) (1=SYSMIS).\nEXECUTE.\n"
                        )
                        result += writeAgrupMulti(
                            prefix, multis, varlabeloriginal + " - " + tipo
                        )
                        filtro = data[varsList.iloc[i][2]] == 5
                        result += writeAbiertasQuestion(
                            varAbierta,
                            colvars,
                            getListOrderConditions(multis, data, listNetos, filtro),
                            includeall=checkinclude,
                            varanidada=varsList.iloc[i][2] + tipo,
                        )
                        listatotal = []
                        for var in multis:
                            listatotal += data[filtro][var].tolist()
                        count2 = Counter(listatotal)
                        listatotaluniq = list(set(listatotal))
                        for net in listNetos:
                            if (
                                net[0] != "First"
                                and net[0] != "End"
                                and any(count2[ele] > 0 for ele in net[1])
                            ):
                                nombreneto = (
                                    net[0].strip().replace(" ", "_") + "_" + varAbierta
                                )
                                result+=f"TEMPORARY.\nSELECT IF (nvalid({multis[0]})=1 and {varsList.iloc[i][2] + tipo}=1).\n"
                                result += writeQuestion(
                                    "NETO_" + nombreneto,
                                    "T",
                                    colvars,
                                    includeall=checkinclude,
                                    varanidada=varsList.iloc[i][2] + tipo,
                                )
                    elif tipo == "MB":
                        result += (
                            "\nDELETE VARIABLES "
                            + varsList.iloc[i][2]
                            + tipo
                            + ".\nEXECUTE."
                        )
                        result += (
                            "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            + varsList.iloc[i][2]
                            + '\n/OPTIONS FIX="'
                            + tipo
                            + '" FIXTYPE=SUFFIX ACTION=RUN.\nEXECUTE.'
                        )
                        result += (
                            "\nVALUE LABELS "
                            + varsList.iloc[i][2]
                            + tipo
                            + ' 1 "'
                            + tipo
                            + '".'
                        )
                        result += (
                            "\nRECODE "
                            + varsList.iloc[i][2]
                            + tipo
                            + " (5=SYSMIS) (4=SYSMIS) (3=1) (2=SYSMIS) (1=SYSMIS).\nEXECUTE.\n"
                        )
                        result += writeAgrupMulti(
                            prefix, multis, varlabeloriginal + " - " + tipo
                        )
                        filtro = data[varsList.iloc[i][2]] == 3
                        result += writeAbiertasQuestion(
                            varAbierta,
                            colvars,
                            getListOrderConditions(multis, data, listNetos, filtro),
                            includeall=checkinclude,
                            varanidada=varsList.iloc[i][2] + tipo,
                        )
                        listatotal = []
                        for var in multis:
                            listatotal += data[filtro][var].tolist()
                        count2 = Counter(listatotal)
                        listatotaluniq = list(set(listatotal))
                        for net in listNetos:
                            if (
                                net[0] != "First"
                                and net[0] != "End"
                                and any(count2[ele] > 0 for ele in net[1])
                            ):
                                nombreneto = (
                                    net[0].strip().replace(" ", "_") + "_" + varAbierta
                                )
                                result+=f"TEMPORARY.\nSELECT IF (nvalid({multis[0]})=1 and {varsList.iloc[i][2] + tipo}=1).\n"
                                result += writeQuestion(
                                    "NETO_" + nombreneto,
                                    "T",
                                    colvars,
                                    includeall=checkinclude,
                                    varanidada=varsList.iloc[i][2] + tipo,
                                )
                    elif tipo == "B2B":
                        result += (
                            "\nDELETE VARIABLES "
                            + varsList.iloc[i][2]
                            + tipo
                            + ".\nEXECUTE."
                        )
                        result += (
                            "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            + varsList.iloc[i][2]
                            + '\n/OPTIONS FIX="'
                            + tipo
                            + '" FIXTYPE=SUFFIX ACTION=RUN.\nEXECUTE.'
                        )
                        result += (
                            "\nVALUE LABELS "
                            + varsList.iloc[i][2]
                            + tipo
                            + ' 1 "'
                            + tipo
                            + '".'
                        )
                        result += (
                            "\nRECODE "
                            + varsList.iloc[i][2]
                            + tipo
                            + " (5=SYSMIS) (4=SYSMIS) (3=SYSMIS) (2=1) (1=1).\nEXECUTE.\n"
                        )
                        result += writeAgrupMulti(
                            prefix, multis, varlabeloriginal + " - " + tipo
                        )
                        condition1 = data[varsList.iloc[i][2]] == 1
                        condition2 = data[varsList.iloc[i][2]] == 2
                        filtro = condition1 | condition2
                        result += writeAbiertasQuestion(
                            varAbierta,
                            colvars,
                            getListOrderConditions(multis, data, listNetos, filtro),
                            includeall=checkinclude,
                            varanidada=varsList.iloc[i][2] + tipo,
                        )
                        listatotal = []
                        for var in multis:
                            listatotal += data[filtro][var].tolist()
                        count2 = Counter(listatotal)
                        listatotaluniq = list(set(listatotal))
                        for net in listNetos:
                            if (
                                net[0] != "First"
                                and net[0] != "End"
                                and any(count2[ele] > 0 for ele in net[1])
                            ):
                                nombreneto = (
                                    net[0].strip().replace(" ", "_") + "_" + varAbierta
                                )
                                result+=f"TEMPORARY.\nSELECT IF (nvalid({multis[0]})=1 and {varsList.iloc[i][2] + tipo}=1).\n"
                                result += writeQuestion(
                                    "NETO_" + nombreneto,
                                    "T",
                                    colvars,
                                    includeall=checkinclude,
                                    varanidada=varsList.iloc[i][2] + tipo,
                                )
                    elif tipo == "B3B":
                        result += (
                            "\nDELETE VARIABLES "
                            + varsList.iloc[i][2]
                            + tipo
                            + ".\nEXECUTE."
                        )
                        result += (
                            "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            + varsList.iloc[i][2]
                            + '\n/OPTIONS FIX="'
                            + tipo
                            + '" FIXTYPE=SUFFIX ACTION=RUN.\nEXECUTE.'
                        )
                        result += (
                            "\nVALUE LABELS "
                            + varsList.iloc[i][2]
                            + tipo
                            + ' 1 "'
                            + tipo
                            + '".'
                        )
                        result += (
                            "\nRECODE "
                            + varsList.iloc[i][2]
                            + tipo
                            + " (5=SYSMIS) (4=SYSMIS) (3=1) (2=1) (1=1).\nEXECUTE.\n"
                        )
                        result += writeAgrupMulti(
                            prefix, multis, varlabeloriginal + " - " + tipo
                        )
                        condition1 = data[varsList.iloc[i][2]] == 1
                        condition2 = data[varsList.iloc[i][2]] == 2
                        condition3 = data[varsList.iloc[i][2]] == 3
                        filtro = condition1 | condition2 | condition3
                        result += writeAbiertasQuestion(
                            varAbierta,
                            colvars,
                            getListOrderConditions(multis, data, listNetos, filtro),
                            includeall=checkinclude,
                            varanidada=varsList.iloc[i][2] + tipo,
                        )
                        listatotal = []
                        for var in multis:
                            listatotal += data[filtro][var].tolist()
                        count2 = Counter(listatotal)
                        listatotaluniq = list(set(listatotal))
                        for net in listNetos:
                            if (
                                net[0] != "First"
                                and net[0] != "End"
                                and any(count2[ele] > 0 for ele in net[1])
                            ):
                                nombreneto = (
                                    net[0].strip().replace(" ", "_") + "_" + varAbierta
                                )
                                result+=f"TEMPORARY.\nSELECT IF (nvalid({multis[0]})=1 and {varsList.iloc[i][2] + tipo}=1).\n"
                                result += writeQuestion(
                                    "NETO_" + nombreneto,
                                    "T",
                                    colvars,
                                    includeall=checkinclude,
                                    varanidada=varsList.iloc[i][2] + tipo,
                                )
            result += writeAgrupMulti(prefix, multis, varlabeloriginal)
    return result, warning


def getListOrderConditions(multis, data, listNetos, condition):
    listatotal = []
    for var in multis:
        listatotal += data[condition][var].tolist()
    count = Counter(listatotal)
    listafinalorde = []
    num = 1
    for net in listNetos:
        if net[0] != "First" and net[0] != "End":
            if any(count[ele] > 0 for ele in net[1]):
                listafinalorde.append(990 + num)
            num += 1
        if net[0] != "End":
            for i in count.most_common():
                if i[0] in net[1]:
                    listafinalorde.append(int(i[0]))
        else:
            for end in net[1]:
                if count[end] > 0:
                    listafinalorde.append(end)
    return listafinalorde


def getPenaltysCode2(spss_file: BytesIO, xlsx_file: BytesIO, rutaarchivo=""):
    result = getPenaltysCode(xlsx_file)
    if getPenaltysCode(xlsx_file) != "":
        file_xlsx = get_temp_file(xlsx_file)
        nombrehoja = "Penaltys"
        try:
            sufijo = (
                pd.read_excel(file_xlsx, usecols="P", skiprows=3, names=["name"])
                .dropna()
                .iloc[0, 0]
            )
            sufijo = " " + str(sufijo)
        except Exception:
            sufijo = ""
        result += (
            "\nOUTPUT EXPORT\n  /CONTENTS  EXPORT=VISIBLE  LAYERS=VISIBLE  MODELVIEWS=PRINTSETTING\n  /XLSX  DOCUMENTFILE='"
            + rutaarchivo
            + "'\n     OPERATION=CREATESHEET  SHEET='"
            + nombrehoja
            + sufijo
            + "'\n     LOCATION=LASTCOLUMN  NOTESCAPTIONS=NO.\n"
            + "OUTPUT CLOSE NAME=*.\nEXECUTE.\n"
        )
        temp_file_name = get_temp_file(spss_file)
        data, study_metadata = pyreadstat.read_sav(
            temp_file_name, apply_value_formats=False
        )
        result += (
            "\n*___TOTAL____________________________________________________________________________\n ____________________________________________________________________________________\n ______"
            + nombrehoja
            + sufijo
            + "______________________________________________________________________________.\n"
        )
        varsList = (
            pd.read_excel(file_xlsx, usecols="M", skiprows=3, names=["varsSegment"])
            .dropna()["varsSegment"]
            .tolist()
        )
        varsList_segment = (
            pd.read_excel(file_xlsx, usecols="N", skiprows=3, names=["vars_Segment"])
            .dropna()["vars_Segment"]
            .tolist()
        )

        penaltys_code = getPenaltysCode(xlsx_file)
        if not varsList_segment:
            for var in varsList:
                refdict = study_metadata.variable_value_labels[var]
                refs_unique = data[var].dropna().unique()
                refs_unique.sort()
                for refindex in refs_unique:
                    name_var = limpiar_texto(refdict[refindex])
                    name_dataset = name_var
                    name_sheet = (
                        nombrehoja
                        + " "
                        + unidecode(refdict[refindex]).replace(".", "")
                        + sufijo
                    )
                    if len(name_sheet) > 30:
                        name_sheet = (
                            nombrehoja
                            + " "
                            + unidecode(refdict[refindex]).replace(".", "")[:10]
                            + sufijo
                        )
                    result += "DATASET ACTIVATE REF_" + name_dataset + ".\n"
                    result += penaltys_code
                    result += (
                        "\nOUTPUT EXPORT\n  /CONTENTS  EXPORT=VISIBLE  LAYERS=VISIBLE  MODELVIEWS=PRINTSETTING\n  /XLSX  DOCUMENTFILE='"
                        + rutaarchivo
                        + "'\n     OPERATION=CREATESHEET  SHEET='"
                        + name_sheet
                        + "'\n     LOCATION=LASTCOLUMN  NOTESCAPTIONS=NO.\n"
                        + "OUTPUT CLOSE NAME=*.\n"
                    )
                    result += "DATASET CLOSE REF_" + name_dataset + ".\n"
                result += (
                    "\n*____________________________________________________________________________________\n ______"
                    + var
                    + "______________________________________________________________________________\n ______"
                    + nombrehoja
                    + sufijo
                    + "______________________________________________________________________________.\nEXECUTE.\n"
                )
        else:
            for var_segment in varsList_segment:
                refdict_segment = study_metadata.variable_value_labels[var_segment]
                refs_segment_unique = data[var_segment].dropna().unique()
                refs_segment_unique.sort()
                for refindex_segment in refs_segment_unique:
                    name_var_segment = limpiar_texto(refdict_segment[refindex_segment])
                    for var in varsList:
                        refdict = study_metadata.variable_value_labels[var]
                        refs_unique = data[var].dropna().unique()
                        refs_unique.sort()
                        for refindex in refs_unique:
                            name_var = limpiar_texto(refdict[refindex])
                            name_dataset = name_var + "_" + name_var_segment
                            name_sheet = (
                                nombrehoja
                                + " "
                                + unidecode(refdict[refindex]).replace(".", "")
                                + " "
                                + unidecode(refdict_segment[refindex_segment]).replace(".", "")
                                + sufijo
                            )
                            if len(name_sheet) > 30:
                                name_sheet = (
                                    nombrehoja
                                    + " "
                                    + unidecode(refdict[refindex]).replace(".", "")[:10]
                                    + " "
                                    + unidecode(refdict_segment[refindex_segment]).replace(".", "")[:10]
                                    + sufijo
                                )
                            result += "DATASET ACTIVATE REF_" + name_dataset + ".\n"
                            result += penaltys_code
                            result += (
                                "\nOUTPUT EXPORT\n  /CONTENTS  EXPORT=VISIBLE  LAYERS=VISIBLE  MODELVIEWS=PRINTSETTING\n  /XLSX  DOCUMENTFILE='"
                                + rutaarchivo
                                + "'\n     OPERATION=CREATESHEET  SHEET='"
                                + name_sheet
                                + "'\n     LOCATION=LASTCOLUMN  NOTESCAPTIONS=NO.\n"
                                + "OUTPUT CLOSE NAME=*.\n"
                            )
                            result += "DATASET CLOSE REF_" + name_dataset + ".\n"
                        result += (
                            "\n*____________________________________________________________________________________"
                            + "\n ______"
                            + var
                            + "______________________________________________________________________________"
                            + "\n ______"
                            + nombrehoja
                            + sufijo
                            + "_"
                            + name_var_segment
                            + "______________________________________________________________________________.\nEXECUTE.\n"
                        )

        result += """*
                    ⠸⣷⣦⠤⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣀⣠⣤⠀⠀⠀
                    ⠀⠙⣿⡄⠈⠑⢄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣀⠔⠊⠉⣿⡿⠁⠀⠀⠀
                    ⠀⠀⠈⠣⡀⠀⠀⠑⢄⠀⠀⠀⠀⠀⠀⠀⠀⠀⡠⠊⠁⠀⠀⣰⠟⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠈⠢⣄⠀⡈⠒⠊⠉⠁⠀⠈⠉⠑⠚⠀⠀⣀⠔⢊⣠⠤⠒⠊⡽⠀⠀
                    ⠀⠀⠀⠀⠀⠀⠀⡽⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠩⡔⠊⠁⠀⠀⠀⠀⠀⠀⠇
                    ⠀⠀⠀⠀⠀⠀⠀⡇⢠⡤⢄⠀⠀⠀⠀⠀⡠⢤⣄⠀⡇⠀⠀⠀⠀⠀⠀⠀⢰⠀
                    ⠀⠀⠀⠀⠀⠀⢀⠇⠹⠿⠟⠀⠀⠤⠀⠀⠻⠿⠟⠀⣇⠀⠀⡀⠠⠄⠒⠊⠁⠀
                    ⠀⠀⠀⠀⠀⠀⢸⣿⣿⡆⠀⠰⠤⠖⠦⠴⠀⢀⣶⣿⣿⠀⠙⢄⠀⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠀⠀⠀⢻⣿⠃⠀⠀⠀⠀⠀⠀⠀⠈⠿⡿⠛⢄⠀⠀⠱⣄⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠀⠀⠀⢸⠈⠓⠦⠀⣀⣀⣀⠀⡠⠴⠊⠹⡞⣁⠤⠒⠉⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠀⠀⣠⠃⠀⠀⠀⠀⡌⠉⠉⡤⠀⠀⠀⠀⢻⠿⠆⠀⠀⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠀⠰⠁⡀⠀⠀⠀⠀⢸⠀⢰⠃⠀⠀⠀⢠⠀⢣⠀⠀⠀⠀⠀⠀⠀⠀
                    ⠀⠀⠀⢶⣗⠧⡀⢳⠀⠀⠀⠀⢸⣀⣸⠀⠀⠀⢀⡜⠀⣸⢤⣶⠀⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠈⠻⣿⣦⣈⣧⡀⠀⠀⢸⣿⣿⠀⠀⢀⣼⡀⣨⣿⡿⠁⠀⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠀⠈⠻⠿⠿⠓⠄⠤⠘⠉⠙⠤⢀⠾⠿⣿⠟⠋⠀⠀
                          ░░░░░░░░║░░░
                          ░░░╔╗╦╗╔╣░░░
                          ░░░╠╝║║║║░░░
                          ░░░╚╝╝╚╚╩░░░
                          ░░░░░░░░░░░░
                    ."""
    return result


def getPenaltysCode(xlsx_file: BytesIO):
    try:
        file_xlsx = get_temp_file(xlsx_file)
        varsList = pd.read_excel(
            file_xlsx, usecols="A,B", skiprows=3, names=["vars", "varsTypes"]
        ).dropna()
        penaltyList = pd.read_excel(
            file_xlsx, usecols="K", skiprows=3, names=["penaltyVars"]
        ).dropna()
        ref = penaltyList.iloc[0][0]
        penaltyList = penaltyList.drop(0)
        penaltyList = penaltyList.iloc[:, 0]
        penaltyCode = ""
        for i in range(len(varsList)):
            typevar = varsList.iloc[i][1]
            if typevar == "J":
                var1 = varsList.iloc[i][0]
                var2 = ""
                try:
                    for penal in penaltyList:
                        if (
                            re.search("_.*", var1).group()
                            == re.search("_.*", penal).group()
                        ):
                            var2 = penal
                            break
                except Exception:
                    for penal in penaltyList:
                        var2 = penal
                penaltyCode += (
                    "\nCTABLES"
                    + "\n  /VLABELS VARIABLES="
                    + var1
                    + " "
                    + ref
                    + " DISPLAY=LABEL  /VLABELS VARIABLES="
                    + var2
                    + " DISPLAY=NONE"
                    + "\n  /PCOMPUTE &cat3 = EXPR([4] + [5])"
                    + '\n  /PPROPERTIES &cat3 LABEL = "TOP TWO" FORMAT=COUNT F40.0 HIDESOURCECATS=YES'
                    + "\n  /PCOMPUTE &cat2 = EXPR([3])"
                    + '\n  /PPROPERTIES &cat2 LABEL = "JUST RIGHT" FORMAT=COUNT F40.0 HIDESOURCECATS=YES'
                    + "\n  /PCOMPUTE &cat1 = EXPR([1] + [2])"
                    + '\n  /PPROPERTIES &cat1 LABEL = "BOTTOM TWO" FORMAT=COUNT F40.0 HIDESOURCECATS=YES'
                    + "\n  /TABLE "
                    + var1
                    + " [C][COUNT F40.0] + "
                    + var1
                    + " [C] > "
                    + var2
                    + " [C][COUNT F40.0] BY "
                    + ref
                    + " [C]"
                    + "\n  /SLABELS VISIBLE=NO"
                    + "\n  /CATEGORIES VARIABLES="
                    + var1
                    + " [1, 2, 3, 4, 5, &cat3, &cat2, &cat1, OTHERNM] EMPTY=INCLUDE TOTAL=YES POSITION=AFTER"
                    + "\n  /CATEGORIES VARIABLES="
                    + var2
                    + " ORDER=A KEY=VALUE EMPTY=INCLUDE TOTAL=YES POSITION=AFTER"
                    + "\n  /CATEGORIES VARIABLES="
                    + ref
                    + " ORDER=A KEY=VALUE EMPTY=EXCLUDE"
                    + "\n  /CRITERIA CILEVEL=95."
                )
        return penaltyCode
    except Exception:
        return ""


def getCruces2(
    spss_file: BytesIO, xlsx_file: BytesIO, checkinclude=False, rutaarchivo=""
):
    result = getCruces(spss_file, xlsx_file, checkinclude)
    if getCruces(spss_file, xlsx_file, checkinclude) != "":
        file_xlsx = get_temp_file(xlsx_file)
        nombrehoja = "Cruces"
        try:
            sufijo = (
                pd.read_excel(file_xlsx, usecols="P", skiprows=3, names=["name"])
                .dropna()
                .iloc[0, 0]
            )
            sufijo = " " + str(sufijo)
        except Exception:
            sufijo = ""
        result += "\nOUTPUT MODIFY\n  /SELECT ALL EXCEPT (TABLES)\n  /DELETEOBJECT DELETE = YES."
        result += (
            "\nOUTPUT EXPORT\n  /CONTENTS  EXPORT=VISIBLE  LAYERS=VISIBLE  MODELVIEWS=PRINTSETTING\n  /XLSX  DOCUMENTFILE='"
            + rutaarchivo
            + "'\n     OPERATION=CREATESHEET  SHEET='"
            + nombrehoja
            + sufijo
            + "'\n     LOCATION=LASTCOLUMN  NOTESCAPTIONS=NO.\n"
            + "OUTPUT CLOSE NAME=*.\nEXECUTE.\n"
        )
        temp_file_name = get_temp_file(spss_file)
        data, study_metadata = pyreadstat.read_sav(
            temp_file_name, apply_value_formats=False
        )
        result += (
            "\n*___TOTAL____________________________________________________________________________\n ____________________________________________________________________________________\n ______"
            + nombrehoja
            + sufijo
            + "______________________________________________________________________________.\n"
        )
        varsList = (
            pd.read_excel(file_xlsx, usecols="M", skiprows=3, names=["varsSegment"])
            .dropna()["varsSegment"]
            .tolist()
        )
        varsList_segment = (
            pd.read_excel(file_xlsx, usecols="N", skiprows=3, names=["vars_Segment"])
            .dropna()["vars_Segment"]
            .tolist()
        )

        if not varsList_segment:
            for var in varsList:
                refdict = study_metadata.variable_value_labels[var]
                refs_unique = data[var].dropna().unique()
                refs_unique.sort()
                for refindex in refs_unique:
                    name_var = limpiar_texto(refdict[refindex])
                    name_dataset = name_var
                    name_sheet = (
                        nombrehoja
                        + " "
                        + unidecode(refdict[refindex]).replace(".", "")
                        + sufijo
                    )
                    if len(name_sheet) > 30:
                        name_sheet = (
                            nombrehoja
                            + " "
                            + unidecode(refdict[refindex]).replace(".", "")[:10]
                            + sufijo
                        )
                    result += "DATASET ACTIVATE REF_" + name_dataset + ".\n"
                    condition = data[var] == refindex
                    result += getCruces(
                        spss_file, xlsx_file, checkinclude, condition=condition
                    )
                    result += "\nOUTPUT MODIFY\n  /SELECT ALL EXCEPT (TABLES)\n  /DELETEOBJECT DELETE = YES."
                    result += (
                        "\nOUTPUT EXPORT\n  /CONTENTS  EXPORT=VISIBLE  LAYERS=VISIBLE  MODELVIEWS=PRINTSETTING\n  /XLSX  DOCUMENTFILE='"
                        + rutaarchivo
                        + "'\n     OPERATION=CREATESHEET  SHEET='"
                        + name_sheet
                        + "'\n     LOCATION=LASTCOLUMN  NOTESCAPTIONS=NO.\n"
                        + "OUTPUT CLOSE NAME=*.\n"
                    )
                    result += "DATASET CLOSE REF_" + name_dataset + ".\n"
                result += (
                    "\n*____________________________________________________________________________________\n ______"
                    + var
                    + "______________________________________________________________________________\n ______"
                    + nombrehoja
                    + sufijo
                    + "______________________________________________________________________________.\nEXECUTE.\n"
                )
        else:
            for var_segment in varsList_segment:
                refdict_segment = study_metadata.variable_value_labels[var_segment]
                refs_segment_unique = data[var_segment].dropna().unique()
                refs_segment_unique.sort()
                for refindex_segment in refs_segment_unique:
                    name_var_segment = limpiar_texto(refdict_segment[refindex_segment])
                    for var in varsList:
                        refdict = study_metadata.variable_value_labels[var]
                        refs_unique = data[var].dropna().unique()
                        refs_unique.sort()
                        for refindex in refs_unique:
                            name_var = limpiar_texto(refdict[refindex])
                            name_dataset = name_var + "_" + name_var_segment
                            name_sheet = (
                                nombrehoja
                                + " "
                                + unidecode(refdict[refindex]).replace(".", "")
                                + " "
                                + unidecode(refdict_segment[refindex_segment]).replace(".", "")
                                + sufijo
                            )
                            if len(name_sheet) > 30:
                                name_sheet = (
                                    nombrehoja
                                    + " "
                                    + unidecode(refdict[refindex]).replace(".", "")[:10]
                                    + " "
                                    + unidecode(refdict_segment[refindex_segment]).replace(".", "")[:10]
                                    + sufijo
                                )
                            result += "DATASET ACTIVATE REF_" + name_dataset + ".\n"
                            condition1 = data[var] == refindex
                            condition2 = data[var_segment] == refindex_segment
                            condition = condition1 & condition2
                            result += getCruces(
                                spss_file, xlsx_file, checkinclude, condition=condition
                            )
                            result += "\nOUTPUT MODIFY\n  /SELECT ALL EXCEPT (TABLES)\n  /DELETEOBJECT DELETE = YES."
                            result += (
                                "\nOUTPUT EXPORT\n  /CONTENTS  EXPORT=VISIBLE  LAYERS=VISIBLE  MODELVIEWS=PRINTSETTING\n  /XLSX  DOCUMENTFILE='"
                                + rutaarchivo
                                + "'\n     OPERATION=CREATESHEET  SHEET='"
                                + name_sheet
                                + "'\n     LOCATION=LASTCOLUMN  NOTESCAPTIONS=NO.\n"
                                + "OUTPUT CLOSE NAME=*.\n"
                            )
                            result += "DATASET CLOSE REF_" + name_dataset + ".\n"
                        result += (
                            "\n*____________________________________________________________________________________"
                            + "\n ______"
                            + var
                            + "______________________________________________________________________________"
                            + "\n ______"
                            + nombrehoja
                            + sufijo
                            + "_"
                            + name_var_segment
                            + "______________________________________________________________________________.\nEXECUTE.\n"
                        )

        result += """*
                    ⠸⣷⣦⠤⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣀⣠⣤⠀⠀⠀
                    ⠀⠙⣿⡄⠈⠑⢄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣀⠔⠊⠉⣿⡿⠁⠀⠀⠀
                    ⠀⠀⠈⠣⡀⠀⠀⠑⢄⠀⠀⠀⠀⠀⠀⠀⠀⠀⡠⠊⠁⠀⠀⣰⠟⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠈⠢⣄⠀⡈⠒⠊⠉⠁⠀⠈⠉⠑⠚⠀⠀⣀⠔⢊⣠⠤⠒⠊⡽⠀⠀
                    ⠀⠀⠀⠀⠀⠀⠀⡽⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠩⡔⠊⠁⠀⠀⠀⠀⠀⠀⠇
                    ⠀⠀⠀⠀⠀⠀⠀⡇⢠⡤⢄⠀⠀⠀⠀⠀⡠⢤⣄⠀⡇⠀⠀⠀⠀⠀⠀⠀⢰⠀
                    ⠀⠀⠀⠀⠀⠀⢀⠇⠹⠿⠟⠀⠀⠤⠀⠀⠻⠿⠟⠀⣇⠀⠀⡀⠠⠄⠒⠊⠁⠀
                    ⠀⠀⠀⠀⠀⠀⢸⣿⣿⡆⠀⠰⠤⠖⠦⠴⠀⢀⣶⣿⣿⠀⠙⢄⠀⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠀⠀⠀⢻⣿⠃⠀⠀⠀⠀⠀⠀⠀⠈⠿⡿⠛⢄⠀⠀⠱⣄⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠀⠀⠀⢸⠈⠓⠦⠀⣀⣀⣀⠀⡠⠴⠊⠹⡞⣁⠤⠒⠉⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠀⠀⣠⠃⠀⠀⠀⠀⡌⠉⠉⡤⠀⠀⠀⠀⢻⠿⠆⠀⠀⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠀⠰⠁⡀⠀⠀⠀⠀⢸⠀⢰⠃⠀⠀⠀⢠⠀⢣⠀⠀⠀⠀⠀⠀⠀⠀
                    ⠀⠀⠀⢶⣗⠧⡀⢳⠀⠀⠀⠀⢸⣀⣸⠀⠀⠀⢀⡜⠀⣸⢤⣶⠀⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠈⠻⣿⣦⣈⣧⡀⠀⠀⢸⣿⣿⠀⠀⢀⣼⡀⣨⣿⡿⠁⠀⠀⠀⠀⠀⠀
                    ⠀⠀⠀⠀⠀⠈⠻⠿⠿⠓⠄⠤⠘⠉⠙⠤⢀⠾⠿⣿⠟⠋⠀⠀
                          ░░░░░░░░║░░░
                          ░░░╔╗╦╗╔╣░░░
                          ░░░╠╝║║║║░░░
                          ░░░╚╝╝╚╚╩░░░
                          ░░░░░░░░░░░░
                    ."""
    return result


def getCruces(
    spss_file: BytesIO, xlsx_file: BytesIO, checkinclude=False, condition=None
):
    try:
        file_xlsx = get_temp_file(xlsx_file)
        varsList = pd.read_excel(
            file_xlsx,
            usecols="G,H,I,J",
            skiprows=3,
            names=["vars", "varsTypes", "crossVars", "sheetname"],
        ).dropna(subset=["vars"])
        crosscode = ""
        for i in range(len(varsList)):
            for crossvar in varsList.iloc[i][2].split():
                if re.search("^[PFSV].*[1-90].*A", crossvar):
                    crossvar = "$" + re.search(".*A", crossvar).group()[:-1]
                if varsList.iloc[i][1] != "A":
                    crosscode += writeQuestion(
                        varsList.iloc[i][0],
                        varsList.iloc[i][1],
                        [crossvar],
                        includeall=checkinclude,
                    )
                else:
                    lcTable = pd.read_excel(
                        file_xlsx,
                        sheet_name=varsList.iloc[i][3],
                        usecols="A,B",
                        skiprows=1,
                        names=["vars", "sheetNames"],
                    ).dropna()
                    temp_file_name = get_temp_file(spss_file)
                    data2, study_metadata = pyreadstat.read_sav(
                        temp_file_name, apply_value_formats=False
                    )
                    colvars = [crossvar]
                    if condition is None:
                        data = data2
                    else:
                        data = data2[condition]
                    varAbierta = varsList.iloc[i][0]
                    prefix = re.search("^[PFSV].*[1-90].*A", varAbierta).group()
                    multis = []
                    for var, label in study_metadata.column_names_to_labels.items():
                        if re.search("^[PFSV].*[1-90].*A", var):
                            if re.search(".*A", var).group() == prefix:
                                multis.append(var)
                    listNetos = []
                    parNeto = []
                    if lcTable.iloc[0][0] != "NETO":
                        parNeto = ["First", []]
                    for j in range(len(lcTable)):
                        if lcTable.iloc[j][0] == "NETO":
                            if parNeto != []:
                                listNetos.append(parNeto)
                            parNeto = [lcTable.iloc[j][1], []]
                        elif lcTable.iloc[j][0] == 95:
                            if parNeto != []:
                                listNetos.append(parNeto)
                            parNeto = ["End", [95]]
                        else:
                            parNeto[1].append(lcTable.iloc[j][0])
                    if parNeto != []:
                        listNetos.append(parNeto)
                    listatotal = []

                    for var in multis:
                        listatotal += data[var].dropna().tolist()
                    count = Counter(listatotal)
                    listafinalorde = []
                    num = 1
                    for net in listNetos:
                        if net[0] != "First" and net[0] != "End":
                            if any(count[ele] > 0 for ele in net[1]):
                                listafinalorde.append(990 + num)
                                crosscode += "\nADD VALUE LABEL "
                                for multivar in multis:
                                    crosscode += multivar + " "
                                crosscode += (
                                    str(990 + num)
                                    + ' "NETO '
                                    + net[0]
                                    + '".\nEXECUTE.\n'
                                )
                                num += 1
                        if net[0] != "End":
                            for in1 in count.most_common():
                                if in1[0] in net[1]:
                                    listafinalorde.append(int(in1[0]))
                        else:
                            for end in net[1]:
                                if count[end] > 0:
                                    listafinalorde.append(end)
                    crosscode += writeAbiertasQuestion(
                        varAbierta, colvars, listafinalorde, includeall=checkinclude
                    )

                    listatotaluniq = list(set(listatotal))
                    for net in listNetos:
                        if (
                            net[0] != "First"
                            and net[0] != "End"
                            and any(count[ele] > 0 for ele in net[1])
                        ):
                            crosscode += ".\nEXECUTE."
                            nombreneto = (
                                net[0].strip().replace(" ", "_") + "_" + varAbierta
                            )
                            crosscode += "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
                            for col in multis:
                                crosscode += col + " "
                            crosscode += (
                                '\n/OPTIONS FIX="NETO_" FIXTYPE=PREFIX ACTION=RUN.\n'
                            )
                            newmultis = ("NETO_" + variablee for variablee in multis)
                            crosscode += "RECODE "
                            for newvar in newmultis:
                                crosscode += newvar + " "
                            crosscode += "\n(SYSMIS=0)"
                            for num in listatotaluniq:
                                if num in net[1]:
                                    crosscode += " (" + str(int(num)) + "=1)"
                                else:
                                    crosscode += " (" + str(int(num)) + "=0)"
                            crosscode += (
                                ".\nEXECUTE.\n\nCOMPUTE NETO_" + nombreneto + "="
                            )
                            newmultis = ("NETO_" + variablee for variablee in multis)
                            for col in newmultis:
                                crosscode += col + "+"
                            crosscode = crosscode[:-1] + ".\nRECODE NETO_" + nombreneto
                            for inde in range(len(multis)):
                                crosscode += " (" + str(inde + 1) + "=1)"
                            crosscode += ".\nEXECUTE.\nDELETE VARIABLES"
                            newmultis = ("NETO_" + variablee for variablee in multis)
                            for newvar in newmultis:
                                crosscode += " " + newvar
                            crosscode += (
                                ".\nEXECUTE.\n\nformats NETO_"
                                + nombreneto
                                + "(f8.0).\nVALUE LABELS NETO_"
                                + nombreneto
                                + ' 1 "NETO '
                                + net[0].strip()
                                + '".\nEXECUTE.\n'
                            )
                            crosscode += writeQuestion(
                                "NETO_" + nombreneto,
                                "T",
                                colvars,
                                includeall=checkinclude,
                            )
        return crosscode
    except Exception:
        return ""


def getVarsSav(spss_file: BytesIO):
    temp_file_name = get_temp_file(spss_file)
    data, study_metadata = pyreadstat.read_sav(
        temp_file_name, apply_value_formats=False
    )
    return study_metadata.column_names


def writeAgrupMulti(prefix, listVars, label):
    try:
        if len(listVars) > 1:
            txt = (
                "\nMRSETS\n  /MCGROUP NAME=$"
                + prefix[:-1]
                + " LABEL='"
                + str(label)
                + "'\n    VARIABLES="
            )
            for var in listVars:
                txt += var + " "
            txt += ".\n"
            return txt
        return ""
    except Exception:
        return ""


def getCloneCodeVars(spss_file: BytesIO, xlsx_file: BytesIO):
    file_xlsx = get_temp_file(xlsx_file)
    colVars = pd.melt(
        pd.read_excel(file_xlsx, nrows=2), var_name="colVars", value_name="colVarsNames"
    ).drop(0)
    temp_file_name = get_temp_file(spss_file)
    data, study_metadata = pyreadstat.read_sav(
        temp_file_name, apply_value_formats=False
    )
    columnVars = colVars.iloc[:, 0]

    columnsclone = "\nDELETE VARIABLES"
    for col in columnVars:
        if not re.search("^[PFSV].*[1-90].*A", col):
            columnsclone += " COL_" + col
    columnsclone += ".\nEXECUTE.\n"
    columnsclone += "SPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
    for col in columnVars:
        if not re.search("^[PFSV].*[1-90].*A", col):
            columnsclone += col + " "
    columnsclone += '\n/OPTIONS FIX="COL_" FIXTYPE=PREFIX ACTION=RUN.\n'
    for row in range(len(colVars)):
        col = colVars.iloc[row][0]
        if re.search("^[PFSV].*[1-90].*A", col):
            prefix = re.search(".*A", col).group()
            serie = False
            multis = []
            for var, label in study_metadata.column_names_to_labels.items():
                if re.search(".A", var):
                    if re.search(".*A", var).group() == prefix:
                        serie = True
                        multis.append(var)
                if serie:
                    if (
                        not re.search(".A", var)
                        or re.search(".*A", var).group() != prefix
                    ):
                        columnsclone += writeAgrupMulti(
                            "COL_" + prefix, multis, colVars.iloc[row][1]
                        )
                        break
    for row in range(len(colVars)):
        if not re.search("^[PFSV].*[1-90].*A", colVars.iloc[row][0]):
            columnsclone += (
                "\nVARIABLE LABELS COL_"
                + colVars.iloc[row][0]
                + " '"
                + colVars.iloc[row][1]
                + "'."
            )
    return columnsclone


def getInverseCodeVars(spss_file: BytesIO, inverseVars):
    temp_file_name = get_temp_file(spss_file)
    data, study_metadata = pyreadstat.read_sav(
        temp_file_name, apply_value_formats=False
    )
    dictValues = study_metadata.variable_value_labels
    inverserecodes = ""
    inverserecodes = "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
    for var in inverseVars:
        inverserecodes += var + " "
    inverserecodes += '\n/OPTIONS FIX="BACKUP_INVERSE_" FIXTYPE=PREFIX ACTION=RUN.\n'
    for var in inverseVars:
        if not re.search("^\([0-9]\).*", dictValues[var][1]):
            inverserecodes += "\nRECODE " + var + " (5=1) (4=2) (2=4) (1=5)."
            inverserecodes += "\nVALUE LABELS " + var
            for i in range(1, 6):
                inverserecodes += (
                    "\n" + str(i) + ' "(' + str(i) + ") " + dictValues[var][6 - i] + '"'
                )
            inverserecodes += "."
    inverserecodes += "\nEXECUTE."
    return inverserecodes


def checkInverseCodeVars(spss_file: BytesIO, inverseVars):
    temp_file_name = get_temp_file(spss_file)
    data, study_metadata = pyreadstat.read_sav(
        temp_file_name, apply_value_formats=False
    )
    dictValues = study_metadata.variable_value_labels
    inverserecodes = ""
    inverserecodes = "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
    for var in inverseVars:
        inverserecodes += var + " "
    inverserecodes += '\n/OPTIONS FIX="BACKUP_INVERSE_" FIXTYPE=PREFIX ACTION=RUN.\n'
    for var in inverseVars:
        if not re.search("^\([0-9]\).*", dictValues[var][1]):
            return True
    return False


def getScaleCodeVars(spss_file: BytesIO, scaleVars):
    scalerecodes = ""
    try:
        temp_file_name = get_temp_file(spss_file)
        data, study_metadata = pyreadstat.read_sav(
            temp_file_name, apply_value_formats=False
        )
        dictValues = study_metadata.variable_value_labels
        scalerecodes = "\nSPSS_TUTORIALS_CLONE_VARIABLES VARIABLES="
        for i in range(len(scaleVars)):
            scalerecodes += scaleVars.iloc[i][0] + " "
        scalerecodes += '\n/OPTIONS FIX="BACKUP_SCALE_" FIXTYPE=PREFIX ACTION=RUN.\n'
        scalecodeback = scalerecodes

        for i in range(len(scaleVars)):
            scalecode1 = ""
            try:
                for num in range(len(scaleVars.iloc[i][1].split())):
                    float(scaleVars.iloc[i][1].split()[num])
                for num in range(len(scaleVars.iloc[i][1].split())):
                    (
                        "\n"
                        + scaleVars.iloc[i][1].split()[num]
                        + ' "('
                        + scaleVars.iloc[i][1].split()[num]
                        + ") "
                        + dictValues[scaleVars.iloc[i][0]][num + 1]
                        + '"'
                    )
                scalecode1 += "\nRECODE " + scaleVars.iloc[i][0]
                for num in range(len(scaleVars.iloc[i][1].split())):
                    scalecode1 += (
                        " ("
                        + str(num + 1)
                        + "="
                        + scaleVars.iloc[i][1].split()[num]
                        + ")"
                    )
                scalecode1 += "."
            except Exception:
                scalecode1 = ""
            scalerecodes += scalecode1
        if scalerecodes == scalecodeback:
            return ""
        scalerecodes += "\nEXECUTE."

        for i in range(len(scaleVars)):
            try:
                scalecode2 = ""
                for num in range(len(scaleVars.iloc[i][1].split())):
                    float(scaleVars.iloc[i][1].split()[num])
                scalecode2 += "\nVALUE LABELS " + scaleVars.iloc[i][0]
                for num in range(len(scaleVars.iloc[i][1].split())):
                    scalecode2 += (
                        "\n"
                        + scaleVars.iloc[i][1].split()[num]
                        + ' "('
                        + scaleVars.iloc[i][1].split()[num]
                        + ") "
                        + dictValues[scaleVars.iloc[i][0]][num + 1]
                        + '"'
                    )
                scalecode2 += "."
            except Exception:
                scalecode2 = ""
            scalerecodes += scalecode2
        return scalerecodes
    except Exception:
        return ""


def getGroupCreateMultisCode(spss_file: BytesIO):
    agrupresult = ""
    temp_file_name = get_temp_file(spss_file)
    data, study_metadata = pyreadstat.read_sav(
        temp_file_name, apply_value_formats=False
    )
    serie = False
    prefix = ""
    multis = []
    label2 = ""
    for var, label in study_metadata.column_names_to_labels.items():
        if re.search("^[PFSV].*[1-90].*A", var):
            if not serie:
                multis = []
                prefix = re.search(".*A", var).group()
                multis.append(var)
                serie = True
            else:
                if re.search(".*A", var).group() == prefix:
                    multis.append(var)
                else:
                    agrupresult += writeAgrupMulti(prefix, multis, label2)
                    multis = []
                    prefix = re.search(".*A", var).group()
                    multis.append(var)
        elif serie:
            agrupresult += writeAgrupMulti(prefix, multis, label2)
            multis = []
            serie = False
        label2 = label
    return agrupresult


def getSegmentCode(spss_file: BytesIO, xlsx_file: BytesIO):
    try:
        temp_file_name = get_temp_file(spss_file)
        data, study_metadata = pyreadstat.read_sav(
            temp_file_name, apply_value_formats=False
        )
        file_xlsx = get_temp_file(xlsx_file)
        varsList = (
            pd.read_excel(file_xlsx, usecols="M", skiprows=3, names=["varsSegment"])
            .dropna()["varsSegment"]
            .tolist()
        )
        varsList_segment = (
            pd.read_excel(file_xlsx, usecols="N", skiprows=3, names=["vars_Segment"])
            .dropna()["vars_Segment"]
            .tolist()
        )

        filterdatabase = ""
        namedatasetspss = "ConjuntoDatos1"
        if not varsList_segment:
            for var in varsList:
                refdict = study_metadata.variable_value_labels[var]
                refs_unique = data[var].dropna().unique()
                refs_unique.sort()
                for refindex in refs_unique:
                    filterdatabase += "DATASET ACTIVATE " + namedatasetspss + ".\n"
                    filterdatabase += (
                        "DATASET COPY REF_"
                        + limpiar_texto(refdict[refindex])
                        + ".\nDATASET ACTIVATE REF_"
                        + limpiar_texto(refdict[refindex])
                        + ".\nFILTER OFF.\nUSE ALL.\n"
                    )
                    filterdatabase += (
                        "SELECT IF ("
                        + var
                        + " = "
                        + str(int(refindex))
                        + ").\nEXECUTE.\n\n"
                    )
        else:
            for var_segment in varsList_segment:
                refdict_segment = study_metadata.variable_value_labels[var_segment]
                refs_segment_unique = data[var_segment].dropna().unique()
                refs_segment_unique.sort()
                for refindex_segment in refs_segment_unique:
                    name_var_segment = limpiar_texto(refdict_segment[refindex_segment])
                    for var in varsList:
                        refdict = study_metadata.variable_value_labels[var]
                        refs_unique = data[var].dropna().unique()
                        refs_unique.sort()
                        for refindex in refs_unique:
                            name_var = limpiar_texto(refdict[refindex])
                            name_dataset = name_var + "_" + name_var_segment
                            filterdatabase += (
                                "DATASET ACTIVATE " + namedatasetspss + ".\n"
                            )
                            filterdatabase += (
                                "DATASET COPY REF_"
                                + name_dataset
                                + ".\nDATASET ACTIVATE REF_"
                                + name_dataset
                                + ".\nFILTER OFF.\nUSE ALL.\n"
                            )
                            filterdatabase += (
                                "SELECT IF ("
                                + var
                                + " = "
                                + str(int(refindex))
                                + " AND "
                                + var_segment
                                + " = "
                                + str(int(refindex_segment))
                                + ").\nEXECUTE.\n\n"
                            )
        return filterdatabase
    except Exception:
        return "Error with Variables to segment"


def writeQuestion(
    varName,
    qtype,
    colVars,
    descendingorder=False,
    includeall=False,
    varanidada="",
    custom_order="",
):
    txt = ""
    if qtype == "M":
        varName = "$" + re.search(".*A", varName).group()[:-1]
    txt += "\nCTABLES\n  /VLABELS VARIABLES=" + varName + " TOTAL "
    for colvar in colVars:
        txt += colvar + " "
    txt += "DISPLAY=LABEL"
    if qtype in ["E", "J"]:
        txt += (
            "\n  /PCOMPUTE &cat1 = EXPR([4]+[5])"
            + "\n  /PPROPERTIES &cat1 LABEL = \"NETO TOP TWO BOX\" FORMAT=COUNT '1' F40.0 HIDESOURCECATS=NO"
            + "\n  /PCOMPUTE &cat2 = EXPR([2]+[1])\n  /PPROPERTIES &cat2 LABEL = \"NETO BOTTOM TWO BOX\" FORMAT=COUNT '1' F40.0 HIDESOURCECATS=NO"
        )

    txt += "\n  /TABLE " + varName + " [C][COUNT '1' F40.0, TOTALS["
    if qtype in ["E", "N"]:
        txt += "MEAN 'Promedio:' F40.2, STDDEV 'Desviación estándar:' F40.2, SEMEAN 'Error estándar:' F40.2,\n"
    if qtype in ["M"]:
        txt += "COUNT 'Total' F40.0, RESPONSES 'Total Respuestas' F40.0, COLPCT.RESPONSES.COUNT '%' F40.0]] BY "
    else:
        txt += "COUNT 'Total' F40.0, TOTALN 'Total Respuestas' F40.0, COLPCT.COUNT '%' F40.0]] BY "
    if varanidada != "":
        txt += varanidada + " > ("
    txt += "TOTAL[C]"
    for colvar in colVars:
        txt += " + " + colvar + " [C]"
    if varanidada != "":
        txt += ")"
    txt += "\n  /SLABELS POSITION=ROW\n  /CATEGORIES VARIABLES=" + varName
    if custom_order != "":
        txt += " " + custom_order + " "
    elif qtype in ["E", "J"]:
        txt += " [&cat1, 5, 4, 3, 2, 1, &cat2] "
    elif qtype in ["M"]:
        txt += " ORDER=D KEY=COUNT "
    elif descendingorder:
        txt += " ORDER=D KEY=VALUE "
    else:
        txt += " ORDER=A KEY=VALUE "

    if qtype in ["E", "J"]:
        txt += "EMPTY=INCLUDE TOTAL=YES POSITION=AFTER"
    else:
        txt += "EMPTY=EXCLUDE TOTAL=YES POSITION=AFTER"

    txt += "\n  /CATEGORIES VARIABLES=TOTAL "
    if varanidada != "":
        txt += varanidada + " "
    for colvar in colVars:
        txt += colvar + " "
    if includeall:
        txt += "ORDER=A EMPTY=INCLUDE"
    else:
        txt += "ORDER=A EMPTY=EXCLUDE"

    txt += (
        "\n  /CRITERIA CILEVEL=95\n  /COMPARETEST TYPE=PROP ALPHA=0.05 ADJUST=BONFERRONI ORIGIN=COLUMN INCLUDEMRSETS=YES"
        + "\n  CATEGORIES=SUBTOTALS MERGE=YES STYLE=SIMPLE SHOWSIG=NO.\n"
    )
    return txt


def writeAbiertasQuestion(varName, colVars, orderlist, includeall=False, varanidada=""):
    txt = ""
    varName = "$" + re.search(".*A", varName).group()[:-1]
    txt += "\nCTABLES\n  /VLABELS VARIABLES=" + varName + " TOTAL "
    for colvar in colVars:
        txt += colvar + " "
    txt += "DISPLAY=LABEL\n  /TABLE " + varName + " [C][COUNT '1' F40.0, TOTALS["
    txt += "COUNT 'Total' F40.0, RESPONSES 'Total Respuestas' F40.0, COLPCT.RESPONSES.COUNT '%' F40.0]] BY "
    if varanidada != "":
        txt += varanidada + " > ("
    txt += "TOTAL[C]"
    for colvar in colVars:
        txt += " + " + colvar + " [C]"
    if varanidada != "":
        txt += ")"
    txt += "\n  /SLABELS POSITION=ROW\n  /CATEGORIES VARIABLES=" + varName
    txt += " [" + ", ".join(str(x) for x in orderlist) + "] "
    txt += "EMPTY=INCLUDE TOTAL=YES POSITION=AFTER"
    txt += "\n  /CATEGORIES VARIABLES=TOTAL "
    if varanidada != "":
        txt += varanidada + " "
    for colvar in colVars:
        txt += colvar + " "
    if includeall:
        txt += "ORDER=A EMPTY=INCLUDE"
    else:
        txt += "ORDER=A EMPTY=EXCLUDE"
    txt += (
        "\n  /CRITERIA CILEVEL=95\n  /COMPARETEST TYPE=PROP ALPHA=0.05 ADJUST=BONFERRONI ORIGIN=COLUMN INCLUDEMRSETS=YES"
        + "\n  CATEGORIES=SUBTOTALS MERGE=YES STYLE=SIMPLE SHOWSIG=NO.\n"
    )
    return txt


def limpiar_texto(texto):
    texto = unidecode(texto)
    texto = re.sub(r'[^a-zA-Z0-9\s]', '', texto)
    texto = re.sub(r'\s+', '_', texto)
    return texto

def entero_a_romano(num):
    valores = [
        (1000, "M"), (900, "CM"), (500, "D"), (400, "CD"),
        (100, "C"), (90, "XC"), (50, "L"), (40, "XL"),
        (10, "X"), (9, "lX"), (5, "V"), (4, "lV"), (1, "l")
    ]
    resultado = ""
    for valor, simbolo in valores:
        while num >= valor:
            resultado += simbolo
            num -= valor
    return resultado

def getVarsForPlantilla(spss_file: BytesIO):
    temp_file_name = get_temp_file(spss_file)
    data, study_metadata = pyreadstat.read_sav(
        temp_file_name, apply_value_formats=False
    )
    dict_values = study_metadata.variable_value_labels
    list_vars = study_metadata.column_names
    n_total = study_metadata.number_rows
    var_type_base = study_metadata.original_variable_types  # F-- Float / A-- String
    dict_labels=study_metadata.column_names_to_labels
    # Create a new Workbook
    wb_new = Workbook()
    # Remove the default sheet created with the new workbook
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)

    ws_plantilla = wb_new.create_sheet(title="Estadisticas Plantilla")

    redFill = PatternFill(start_color="C80000", end_color="C80000", fill_type="solid")

    yellowFill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    greenFillTitle = PatternFill(
        start_color="70AD47", end_color="70AD47", fill_type="solid"
    )

    medium_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    ws_plantilla.cell(row=1, column=1).value = "Variable"
    ws_plantilla.cell(row=1, column=2).value = "Tipo de Variable"
    ws_plantilla.cell(row=1, column=3).value = "NValids"
    ws_plantilla.cell(row=1, column=4).value = "Unique Distinct Values"
    ws_plantilla.cell(row=1, column=5).value = "Total Options Values"
    ws_plantilla.cell(row=1, column=6).value = "Option 5"
    ws_plantilla.cell(row=1,column=7).value="Nombre Var"
    ws_plantilla.cell(row=1,column=8).value="Opt. Labels"

    for col in range(1,9):
        ws_plantilla.cell(row=1,column=col).fill=greenFillTitle
        ws_plantilla.cell(row=1,column=col).font = Font(color = "FFFFFF")
        ws_plantilla.cell(row=1,column=col).border = medium_border
        column_letter = get_column_letter(col)
        if col==1:
            width_col=22
        elif col in [7]:
            width_col=20
        elif col in [8]:
            width_col=25
        else:
            width_col=7
        ws_plantilla.column_dimensions[column_letter].width = width_col

    ws_plantilla.auto_filter.ref = ws_plantilla.dimensions

    textPlantilla = "Variable\tTipo de Variable\tNValids\tUnique Values"
    multis = []
    row_num = 2

    vars_to_ignore = [
        "s1",
        "s2",
        "Response_ID",
        "IMAGEN",
        "MARCA",
        "PRECIO",
        "TelA",
        "N.encuesta",
        "tipo_super",
        "ABIERTAS",
        "ETIQUETAS",
        "tipo.super",
        "Tel.",
        "N.enc.",
        "tipo.",
        "acabado.",
    ]
    for var in list_vars:
        if var_type_base[var].startswith("A") or any(
            var.lower().startswith(ignore.lower()) for ignore in vars_to_ignore
        ):
            continue
        if re.search("^[FPS].*A.*[1-90]", var):
            group_multi = re.search(".*A", var).group()
            if group_multi in multis:
                continue
            multis.append(group_multi)

            textPlantilla += "\n" + var + "\t"
            ws_plantilla.cell(row=row_num, column=1).value = var
            try:
                dict_values[var]
                textPlantilla += "M\t"
                ws_plantilla.cell(row=row_num, column=2).value = "M"
            except Exception:
                textPlantilla += "A\t"
                ws_plantilla.cell(row=row_num, column=2).value = "A"
            index_list = []
            count_unique = 0
            count_total = 0
            for var2 in list_vars:
                if re.search(group_multi, var2):
                    count_total += 1
                    if len(data[var2].dropna().index.tolist()) != 0:
                        count_unique += 1
                    for ind in data[var2].dropna().index.tolist():
                        index_list.append(ind)
            textPlantilla += (
                str(len(list(set(index_list))))
                + "\t"
                + str(count_unique)
                + "|"
                + str(count_total)
            )
            ws_plantilla.cell(row=row_num, column=3).value = str(
                len(list(set(index_list)))
            )
            if ws_plantilla.cell(row=row_num, column=3).value != str(n_total):
                ws_plantilla.cell(row=row_num, column=3).fill = yellowFill
            if ws_plantilla.cell(row=row_num, column=3).value == "0":
                ws_plantilla.cell(row=row_num, column=3).fill = redFill
            ws_plantilla.cell(row=row_num, column=4).value = (
                str(count_unique) + "|" + str(count_total)
            )
            if ws_plantilla.cell(row=row_num, column=4).value == "1|1":
                ws_plantilla.cell(row=row_num, column=4).fill = yellowFill
        else:
            textPlantilla += "\n" + var + "\t"
            ws_plantilla.cell(row=row_num, column=1).value = var
            try:
                if len(dict_values[var]) == 5:
                    if "just" in dict_values[var][3] or "Just" in dict_values[var][3]:
                        textPlantilla += "J\t"
                        ws_plantilla.cell(row=row_num, column=2).value = "J"
                    else:
                        textPlantilla += "E\t"
                        ws_plantilla.cell(row=row_num, column=2).value = "E"
                        if "5" not in dict_values[var][5]:
                            ws_plantilla.cell(row=row_num, column=6).value = "not 5"
                            ws_plantilla.cell(row=row_num, column=6).fill = yellowFill
                elif dict_values[var][1] == "":
                    textPlantilla += "N\t"
                    ws_plantilla.cell(row=row_num, column=2).value = "N"
                else:
                    textPlantilla += "U\t"
                    ws_plantilla.cell(row=row_num, column=2).value = "U"
            except Exception:
                textPlantilla += "U\t"
                ws_plantilla.cell(row=row_num, column=2).value = "U"
            textPlantilla += str(len(data[var].dropna())) + "\t"
            ws_plantilla.cell(row=row_num, column=3).value = str(
                len(data[var].dropna())
            )
            if ws_plantilla.cell(row=row_num, column=3).value != str(n_total):
                ws_plantilla.cell(row=row_num, column=3).fill = yellowFill
            if ws_plantilla.cell(row=row_num, column=3).value == "0":
                ws_plantilla.cell(row=row_num, column=3).fill = redFill

            ws_plantilla.cell(row=row_num, column=4).value = str(
                len(list(set(data[var].dropna())))
            )
            if ws_plantilla.cell(row=row_num, column=4).value == "1":
                ws_plantilla.cell(row=row_num, column=4).fill = yellowFill

            try:
                textPlantilla += (
                    str(len(list(set(data[var].dropna()))))
                    + "\t/"
                    + str(len(dict_values[var]))
                )
                ws_plantilla.cell(row=row_num, column=5).value = "/" + str(
                    len(dict_values[var])
                )
            except Exception:
                textPlantilla += str(len(list(set(data[var].dropna())))) + "\t/--"
                ws_plantilla.cell(row=row_num, column=5).value = "/--"
        try:
            labelval1=dict_values[var]
        except:
            labelval1="None"
        label_base1=dict_labels[var]
        ws_plantilla.cell(row=row_num,column=7).value=str(label_base1)
        ws_plantilla.cell(row=row_num,column=8).value=str(labelval1)
        ws_plantilla.cell(row=row_num,column=9).value=" "
        row_num += 1
    return textPlantilla, write_temp_excel(wb_new)


def get_comparison_tables(spss_file1: BytesIO, spss_file2: BytesIO):
    # Create a new Workbook
    wb_new = Workbook()
    # Remove the default sheet created with the new workbook
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)
    for caso in range(2):
        ws_plantilla = wb_new.create_sheet(
            title="Estadisticas Plantilla " + str(caso + 1)
        )
        if caso == 0:
            archivo1 = spss_file1
            archivo2 = spss_file2
        else:
            archivo1 = spss_file2
            archivo2 = spss_file1
        temp_file_name = get_temp_file(archivo1)
        data, study_metadata = pyreadstat.read_sav(
            temp_file_name, apply_value_formats=False
        )
        temp_file_name2 = get_temp_file(archivo2)
        data2, study_metadata2 = pyreadstat.read_sav(
            temp_file_name2, apply_value_formats=False
        )

        dict_values = study_metadata.variable_value_labels
        list_vars = study_metadata.column_names
        dict_values2 = study_metadata2.variable_value_labels
        list_vars2 = study_metadata2.column_names
        n_total = study_metadata.number_rows
        var_type_base = study_metadata.original_variable_types  # F-- Float / A-- String

        dict_labels = study_metadata.column_names_to_labels
        dict_labels2 = study_metadata2.column_names_to_labels

        redFill = PatternFill(
            start_color="C80000", end_color="C80000", fill_type="solid"
        )

        yellowFill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        blueFill = PatternFill(
            start_color="C5D9F1", end_color="C5D9F1", fill_type="solid"
        )

        grayFill = PatternFill(
            start_color="DBDBDB", end_color="DBDBDB", fill_type="solid"
        )

        greenFillTitle = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )

        medium_border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )
        # green_thin_border = Border(
        #     left=Side(border_style="thin", color="F7F9F1"),
        #     right=Side(border_style="thin", color="F7F9F1"),
        #     top=Side(border_style="thin", color="F7F9F1"),
        #     bottom=Side(border_style="thin", color="F7F9F1"),
        # )
        # greenBackground = PatternFill(
        #     start_color="EBF1DE", end_color="EBF1DE", fill_type="solid"
        # )
        # blueBackground = PatternFill(
        #     start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
        # )

        ws_plantilla.cell(row=1, column=1).value = "Variable"
        ws_plantilla.cell(row=1, column=2).value = "Tipo de Variable"
        ws_plantilla.cell(row=1, column=3).value = "NValids"
        ws_plantilla.cell(row=1, column=4).value = "Unique Distinct Values"
        ws_plantilla.cell(row=1, column=5).value = "Total Options Values"
        ws_plantilla.cell(row=1, column=6).value = "Option 5"
        ws_plantilla.cell(row=1, column=7).value = "Var in other Base"
        ws_plantilla.cell(row=1, column=8).value = "Same Value labels"
        ws_plantilla.cell(row=1, column=9).value = "Labels Base 1"
        ws_plantilla.cell(row=1, column=10).value = "Labels Base 2"
        ws_plantilla.cell(row=1, column=11).value = "T. Rate Options"
        ws_plantilla.cell(row=1, column=12).value = "Same label"
        ws_plantilla.cell(row=1, column=13).value = "Label 1"
        ws_plantilla.cell(row=1, column=14).value = "Label 2"
        ws_plantilla.cell(row=1, column=15).value = "Similar Rate"
        ws_plantilla.cell(row=1, column=16).value = archivo1.name
        ws_plantilla.cell(row=1, column=17).value = archivo2.name

        # for row in range(2,len(list_vars)+1):
        #     for col in range(1,16):
        #         if col in [7,8,9,12,13,17]:
        #             fillcol=blueBackground
        #             bordercol=blue_thin_border
        #         else:
        #             fillcol=greenBackground
        #             bordercol=green_thin_border
        #         ws_plantilla.cell(row=row,column=col).fill=fillcol
        #         ws_plantilla.cell(row=row,column=col).border=bordercol

        for col in range(1, 18):
            if col in [7, 8, 9, 12, 13, 17]:
                fillcol = blueFill
            else:
                fillcol = greenFillTitle
            ws_plantilla.cell(row=1, column=col).fill = fillcol
            ws_plantilla.cell(row=1, column=col).font = Font(color="FFFFFF")
            ws_plantilla.cell(row=1, column=col).border = medium_border
            column_letter = get_column_letter(col)
            if col in [1]:
                width_col = 22
            elif col in [11]:
                width_col = 9
            elif col in [16, 17]:
                width_col = 50
            else:
                width_col = 7
            ws_plantilla.column_dimensions[column_letter].width = width_col

        ws_plantilla.auto_filter.ref = ws_plantilla.dimensions

        multis = []
        row_num = 2

        vars_to_ignore = [
            "s1",
            "s2",
            "Response_ID",
            "IMAGEN",
            "MARCA",
            "PRECIO",
            "TelA",
            "N.encuesta",
            "tipo_super",
            "ABIERTAS",
            "ETIQUETAS",
            "tipo.super",
            "Tel.",
            "N.enc.",
            "tipo.",
            "acabado.",
        ]
        for var in list_vars:
            if var_type_base[var].startswith("A") or any(
                var.lower().startswith(ignore.lower()) for ignore in vars_to_ignore
            ):
                continue
            if re.search("^[FPS].*A.*[1-90]", var):
                group_multi = re.search(".*A", var).group()
                if group_multi in multis:
                    continue
                multis.append(group_multi)

                ws_plantilla.cell(row=row_num, column=1).value = var
                try:
                    dict_values[var]
                    ws_plantilla.cell(row=row_num, column=2).value = "M"
                except Exception:
                    ws_plantilla.cell(row=row_num, column=2).value = "A"
                index_list = []
                count_unique = 0
                count_total = 0
                for var2 in list_vars:
                    if re.search(group_multi, var2):
                        count_total += 1
                        if len(data[var2].dropna().index.tolist()) != 0:
                            count_unique += 1
                        for ind in data[var2].dropna().index.tolist():
                            index_list.append(ind)
                count_total2 = 0
                for var3 in list_vars2:
                    if re.search(group_multi, var3):
                        count_total2 += 1
                ws_plantilla.cell(row=row_num, column=3).value = str(
                    len(list(set(index_list)))
                )
                if ws_plantilla.cell(row=row_num, column=3).value != str(n_total):
                    ws_plantilla.cell(row=row_num, column=3).fill = yellowFill
                if ws_plantilla.cell(row=row_num, column=3).value == "0":
                    ws_plantilla.cell(row=row_num, column=3).fill = redFill
                ws_plantilla.cell(row=row_num, column=4).value = (
                    str(count_unique) + "|" + str(count_total)
                )
                if ws_plantilla.cell(row=row_num, column=4).value == "1|1":
                    ws_plantilla.cell(row=row_num, column=4).fill = yellowFill
                if var in list_vars2:
                    ws_plantilla.cell(row=row_num, column=7).value = var
                    try:
                        labelval1 = dict_values[var]
                    except Exception:
                        labelval1 = ""
                    try:
                        labelval2 = dict_values2[var]
                    except Exception:
                        labelval2 = ""
                    label_base1 = dict_labels[var]
                    label_base2 = dict_labels2[var]
                    if (labelval1 == labelval2) and (count_total == count_total2):
                        ws_plantilla.cell(row=row_num, column=8).value = "Yes"
                    elif (labelval1 == labelval2) and (count_total != count_total2):
                        ws_plantilla.cell(row=row_num, column=8).value = "Same Labels"
                        ws_plantilla.cell(row=row_num, column=8).fill = yellowFill
                        ws_plantilla.cell(
                            row=row_num, column=9
                        ).value = "Distint number of variables in multiple"
                        ws_plantilla.cell(row=row_num, column=10).value = (
                            str(count_total2) + " | " + str(count_total)
                        )
                    else:
                        ws_plantilla.cell(row=row_num, column=8).value = "No"
                        ws_plantilla.cell(row=row_num, column=8).fill = redFill
                        ws_plantilla.cell(row=row_num, column=9).value = str(labelval2)
                        ws_plantilla.cell(row=row_num, column=9).fill = grayFill
                        ws_plantilla.cell(row=row_num, column=7).fill = grayFill
                        ws_plantilla.cell(row=row_num, column=10).value = str(labelval1)
                        ws_plantilla.cell(row=row_num, column=11).value = (
                            str(len(labelval2))
                            + "|"
                            + str(len(labelval1))
                            + "--"
                            + "{0:.0%}".format(
                                SequenceMatcher(
                                    None, str(labelval2).lower(), str(labelval1).lower()
                                ).ratio()
                            )
                        )
                    if label_base1 == label_base2:
                        ws_plantilla.cell(row=row_num, column=12).value = "Yes"
                    else:
                        ws_plantilla.cell(row=row_num, column=12).value = "No"
                        ws_plantilla.cell(row=row_num, column=12).fill = blueFill
                        ws_plantilla.cell(row=row_num, column=13).value = str(
                            label_base2
                        )
                        ws_plantilla.cell(row=row_num, column=13).fill = grayFill
                        ws_plantilla.cell(row=row_num, column=7).fill = grayFill
                        ws_plantilla.cell(row=row_num, column=14).value = str(
                            label_base1
                        )
                        ws_plantilla.cell(
                            row=row_num, column=15
                        ).value = "{0:.0%}".format(
                            SequenceMatcher(
                                None, str(label_base2).lower(), str(label_base1).lower()
                            ).ratio()
                        )
                else:
                    ws_plantilla.cell(row=row_num, column=7).value = "No"
                    ws_plantilla.cell(row=row_num, column=7).fill = blueFill
            else:
                ws_plantilla.cell(row=row_num, column=1).value = var
                try:
                    if len(dict_values[var]) == 5:
                        if (
                            "just" in dict_values[var][3]
                            or "Just" in dict_values[var][3]
                        ):
                            ws_plantilla.cell(row=row_num, column=2).value = "J"
                        else:
                            ws_plantilla.cell(row=row_num, column=2).value = "E"
                            if "5" not in dict_values[var][5]:
                                ws_plantilla.cell(row=row_num, column=6).value = "not 5"
                                ws_plantilla.cell(
                                    row=row_num, column=6
                                ).fill = yellowFill
                    elif dict_values[var][1] == "":
                        ws_plantilla.cell(row=row_num, column=2).value = "N"
                    else:
                        ws_plantilla.cell(row=row_num, column=2).value = "U"
                except Exception:
                    ws_plantilla.cell(row=row_num, column=2).value = "U"
                ws_plantilla.cell(row=row_num, column=3).value = str(
                    len(data[var].dropna())
                )
                if ws_plantilla.cell(row=row_num, column=3).value != str(n_total):
                    ws_plantilla.cell(row=row_num, column=3).fill = yellowFill
                if ws_plantilla.cell(row=row_num, column=3).value == "0":
                    ws_plantilla.cell(row=row_num, column=3).fill = redFill

                ws_plantilla.cell(row=row_num, column=4).value = str(
                    len(list(set(data[var].dropna())))
                )
                if ws_plantilla.cell(row=row_num, column=4).value == "1":
                    ws_plantilla.cell(row=row_num, column=4).fill = yellowFill
                try:
                    ws_plantilla.cell(row=row_num, column=5).value = "/" + str(
                        len(dict_values[var])
                    )
                except Exception:
                    ws_plantilla.cell(row=row_num, column=5).value = "/--"
                if var in list_vars2:
                    ws_plantilla.cell(row=row_num, column=7).value = var
                    try:
                        labelval1 = dict_values[var]
                    except Exception:
                        labelval1 = ""
                    try:
                        labelval2 = dict_values2[var]
                    except Exception:
                        labelval2 = ""

                    label_base1 = dict_labels[var]
                    label_base2 = dict_labels2[var]

                    if labelval1 == labelval2:
                        ws_plantilla.cell(row=row_num, column=8).value = "Yes"
                    else:
                        ws_plantilla.cell(row=row_num, column=8).value = "No"
                        ws_plantilla.cell(row=row_num, column=8).fill = redFill
                        ws_plantilla.cell(row=row_num, column=9).value = str(labelval2)
                        ws_plantilla.cell(row=row_num, column=9).fill = grayFill
                        ws_plantilla.cell(row=row_num, column=7).fill = grayFill
                        ws_plantilla.cell(row=row_num, column=10).value = str(labelval1)
                        ws_plantilla.cell(row=row_num, column=11).value = (
                            str(len(labelval2))
                            + "|"
                            + str(len(labelval1))
                            + "--"
                            + "{0:.0%}".format(
                                SequenceMatcher(
                                    None, str(labelval2).lower(), str(labelval1).lower()
                                ).ratio()
                            )
                        )
                    if label_base1 == label_base2:
                        ws_plantilla.cell(row=row_num, column=12).value = "Yes"
                    else:
                        ws_plantilla.cell(row=row_num, column=12).value = "No"
                        ws_plantilla.cell(row=row_num, column=12).fill = blueFill
                        ws_plantilla.cell(row=row_num, column=13).value = str(
                            label_base2
                        )
                        ws_plantilla.cell(row=row_num, column=13).fill = grayFill
                        ws_plantilla.cell(row=row_num, column=7).fill = grayFill
                        ws_plantilla.cell(row=row_num, column=14).value = str(
                            label_base1
                        )
                        ws_plantilla.cell(
                            row=row_num, column=15
                        ).value = "{0:.0%}".format(
                            SequenceMatcher(
                                None, str(label_base2).lower(), str(label_base1).lower()
                            ).ratio()
                        )
                else:
                    ws_plantilla.cell(row=row_num, column=7).value = "No"
                    ws_plantilla.cell(row=row_num, column=7).fill = blueFill
            row_num += 1

    return write_temp_excel(wb_new)


def get_lc_comparison(lc_xlsx_file1: BytesIO, lc_xlsx_file2: BytesIO):
    file_xlsx1 = get_temp_file(lc_xlsx_file1, ".xlsx")
    file_xlsx2 = get_temp_file(lc_xlsx_file2, ".xlsx")

    wb_lc_1_first = load_workbook(file_xlsx1)
    wb_lc_2_first = load_workbook(file_xlsx2)

    wb_new = Workbook()
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)

    for caso in range(2):
        if caso == 0:
            wb_lc1 = wb_lc_1_first
            wb_lc2 = wb_lc_2_first
            name_sheet=lc_xlsx_file1.name
            name_sheet2=lc_xlsx_file2.name
        else:
            wb_lc1 = wb_lc_2_first
            wb_lc2 = wb_lc_1_first
            name_sheet=lc_xlsx_file2.name
            name_sheet2=lc_xlsx_file1.name

        ws_lc_comparison = wb_new.create_sheet(
            title="LC comparison -"+name_sheet[:10]+ str(caso + 1)
        )

        ws_lc_comparison.cell(row=1, column=1).value = "Sheet Name"
        ws_lc_comparison.cell(row=1, column=2).value = "Codes duplicated"
        ws_lc_comparison.cell(row=1, column=3).value = "Verbatims duplicated"
        ws_lc_comparison.cell(row=1, column=4).value = "In LC 2"
        ws_lc_comparison.cell(row=1, column=5).value = "LC equals"
        ws_lc_comparison.cell(row=1, column=6).value = "Codes Missing in LC2"
        ws_lc_comparison.cell(row=1, column=7).value = "Distinct labels in LC2"
        ws_lc_comparison.cell(row=1, column=8).value = f"LC1={name_sheet}"
        ws_lc_comparison.cell(row=1, column=9).value = f"LC2={name_sheet2}"
        ws_lc_comparison.cell(row=1, column=10).value = " "

        blueFill = PatternFill(
            start_color="C5D9F1", end_color="C5D9F1", fill_type="solid"
        )
        yellowFill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        medium_border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )

        for col in range(1, 10):
            ws_lc_comparison.cell(row=1, column=col).fill = blueFill
            ws_lc_comparison.cell(row=1, column=col).border = medium_border

        row_num=2
        for sheet in wb_lc1:
            ws_lc_comparison.cell(row=row_num, column=1).value = sheet.title

            for col_dup in range(2):
                valores = []
                for fila in sheet.iter_rows(min_col=col_dup+1, max_col=col_dup+1, min_row=2, values_only=True):
                    if fila[0] is not None and fila[0]!='NETO':  # Ignorar celdas vacías
                        valores.append(fila[0])

                # Buscar duplicados
                duplicados = set([valor for valor in valores if valores.count(valor) > 1])

                if duplicados:
                    ws_lc_comparison.cell(row=row_num, column=col_dup+2).fill=yellowFill
                    text_dups=""
                    for dup in duplicados:
                        list_rows=[]
                        list_values=[]
                        for rowi in range(1,sheet.max_row+1):
                            if sheet.cell(row=rowi, column=col_dup+1).value==dup:
                                list_rows.append(rowi)
                                list_values.append(sheet.cell(row=rowi, column=2-col_dup).value)
                        text_dups+=f"{dup} \t |-- values: {list_values}  --|-- rows: {list_rows}\n"
                    ws_lc_comparison.cell(row=row_num, column=col_dup+2).value=text_dups[:-2]
                else:
                    ws_lc_comparison.cell(row=row_num, column=col_dup+2).value="No"

            if sheet.title in wb_lc2.sheetnames:
                ws_lc_comparison.cell(row=row_num, column=4).value = "Yes"
                ws1=sheet
                ws2=wb_lc2[sheet.title]

                dict_code_and_values1 = defaultdict(list)
                dict_code_and_values2 = defaultdict(list)
                max_filas = max(ws1.max_row, ws2.max_row)

                for fila in range(1, max_filas + 1):
                    valor_ws1_col1 = ws1.cell(row=fila, column=1).value
                    valor_ws1_col2 = ws1.cell(row=fila, column=2).value

                    valor_ws2_col1 = ws2.cell(row=fila, column=1).value
                    valor_ws2_col2 = ws2.cell(row=fila, column=2).value

                    if valor_ws1_col1 != valor_ws2_col1 or valor_ws1_col2 != valor_ws2_col2:
                        ws_lc_comparison.cell(row=row_num, column=5).value = "No"

                    if valor_ws1_col1 is not None:
                        dict_code_and_values1[valor_ws1_col1].append(valor_ws1_col2 if valor_ws1_col2 is not None else "")

                    if valor_ws2_col1 is not None:
                        dict_code_and_values2[valor_ws2_col1].append(valor_ws2_col2 if valor_ws2_col2 is not None else "")

                if ws_lc_comparison.cell(row=row_num, column=5).value != "No":
                    ws_lc_comparison.cell(row=row_num, column=5).value = "Yes"

                codes_missings=[]
                distints_labels=[]
                dict_distint_lab1=[]
                dict_distint_lab2=[]
                count_sames=0
                for key_lc1 in dict_code_and_values1.keys():
                    if not key_lc1 in dict_code_and_values2.keys():
                        codes_missings.append(key_lc1)
                    elif dict_code_and_values1[key_lc1]!=dict_code_and_values2[key_lc1]:
                        distints_labels.append(str(key_lc1)+"-"+str(dict_code_and_values1[key_lc1])+"|"+str(dict_code_and_values2[key_lc1])+"|"+"{0:.0%}".format(
                            SequenceMatcher(
                                None, str(dict_code_and_values1[key_lc1]).lower(), str(dict_code_and_values2[key_lc1]).lower()
                            ).ratio()
                        ))
                        dict_distint_lab1.append(str(dict_code_and_values1[key_lc1]))
                        dict_distint_lab2.append(str(dict_code_and_values2[key_lc1]))
                    elif dict_code_and_values1[key_lc1]==dict_code_and_values2[key_lc1]:
                        count_sames+=1
                if codes_missings:
                    ws_lc_comparison.cell(row=row_num, column=6).value = str(codes_missings)
                else:
                    ws_lc_comparison.cell(row=row_num, column=6).value = "         LC2 Contain all codes"
                if distints_labels:
                    text_distint_labels=""
                    subcodes1=dict_distint_lab1.count("['']")
                    subcodes2=dict_distint_lab2.count("['']")
                    for val in distints_labels:
                        if re.search("\[''\]",val):
                            continue
                        text_distint_labels+=val+"\n"
                    ws_lc_comparison.cell(row=row_num, column=7).value = text_distint_labels+ f"S1={subcodes1},S2={subcodes2},Sames={count_sames}"
                else:
                    ws_lc_comparison.cell(row=row_num, column=7).value = "         All Labels are equal"
                ws_lc_comparison.cell(row=row_num, column=8).value=" "
            else:
                ws_lc_comparison.cell(row=row_num, column=4).value = "No"
            row_num+=1

        for row in ws_lc_comparison.iter_rows():
            max_lines = 1  # Valor mínimo de líneas por celda
            for cell in row:
                if cell.value and isinstance(cell.value, str):  # Si el contenido es texto
                    max_lines = max(max_lines, cell.value.count("\n") + 1)
                    cell.alignment = Alignment(wrap_text=True)

            # Ajustar la altura de la fila (cada línea adicional suma un tamaño base)
            ws_lc_comparison.row_dimensions[row[0].row].height = max_lines * 15  # Ajustar según necesidad

        columnas_ajustar = ["B", "C", "G"]
        for col in columnas_ajustar:
            max_length = 0
            for cell in ws_lc_comparison[col]:
                if cell.value and isinstance(cell.value, str):
                    lineas = cell.value.split("\n")  # Dividir en líneas
                    max_line_length = max(len(linea) for linea in lineas)  # Buscar la línea más larga
                    max_length = max(max_length, max_line_length)
            max_length=min(max_length,100)
            ws_lc_comparison.column_dimensions[col].width = max_length*0.9

        ws_lc_comparison.column_dimensions["H"].width = 30
        ws_lc_comparison.column_dimensions["I"].width = 30
    return write_temp_excel(wb_new)


def merge_with_border_range(ws, cell_range: str, border: Border):
    """
    Combina celdas según un rango estilo Excel (ej: 'A2:C2') y aplica el borde a todas las celdas del rango.

    :param ws: Hoja de cálculo
    :param cell_range: Rango tipo 'A2:C2'
    :param border: Objeto Border a aplicar
    """
    # Combinar celdas
    ws.merge_cells(cell_range)

    # Obtener coordenadas numéricas del rango
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    # Aplicar borde a cada celda del rango
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border

def apply_outer_border_range(ws, cell_range: str, border_style="medium"):
    """
    Aplica un borde exterior al rango dado, tipo 'B2:E10'
    """
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    side = Side(style=border_style)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            current = cell.border

            cell.border = Border(
                left=side if col == min_col else current.left,
                right=side if col == max_col else current.right,
                top=side if row == min_row else current.top,
                bottom=side if row == max_row else current.bottom,
            )

def apply_medium_bottom_border(cell):
    """
    Applies a medium bottom border to the given cell,
    preserving the existing top, left, and right borders.
    """
    # Define medium bottom border
    medium = Side(style="medium")

    # Get current borders of the cell
    current_border = cell.border

    # Create a new border, keeping other sides unchanged
    new_border = Border(
        left=current_border.left,
        right=current_border.right,
        top=current_border.top,
        bottom=medium
    )

    # Assign the new border to the cell
    cell.border = new_border

def apply_medium_top_border(cell):
    """
    Applies a medium top border to the given cell,
    preserving the existing bottom, left, and right borders.
    """
    # Define medium top border
    medium = Side(style="medium")

    # Get current borders of the cell
    current_border = cell.border

    # Create a new border, keeping other sides unchanged
    new_border = Border(
        left=current_border.left,
        right=current_border.right,
        bottom=current_border.bottom,
        top=medium
    )

    # Assign the new border to the cell
    cell.border = new_border

def merge_and_style_with_wrap(ws, cell_range: str, border):
    """
    Combines cells in the given range, applies border, sets bold text,
    and enables wrap text in the visible merged cell.
    """
    # Paso 1: merge + border
    merge_with_border_range(ws, cell_range, border)

    # Paso 2: aplicar estilo a la celda visible (superior izquierda)
    min_col, min_row, _, _ = range_boundaries(cell_range)
    cell = ws.cell(row=min_row, column=min_col)

    # Aplicar negrita y ajuste de texto
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

def merge_identical_cells_with_border(ws, col_letter: str, start_row: int):
    """
    Merges consecutive identical cells in a column starting at a given row and applies a medium border to each merged group.

    :param ws: Worksheet
    :param col_letter: Column letter (e.g., "A")
    :param start_row: Starting row for merging
    :return: List of ending rows of each merged block
    """
    medium_side = Side(style="medium")
    full_border = Border(top=medium_side, bottom=medium_side, left=medium_side, right=medium_side)

    current_value = None
    merge_start = start_row
    end_rows = []

    for row in range(start_row, ws.max_row + 2):
        cell_value = ws[f"{col_letter}{row}"].value if row <= ws.max_row else None

        if cell_value != current_value:
            if current_value is not None:
                if row - merge_start > 1:
                    cell_range = f"{col_letter}{merge_start}:{col_letter}{row - 1}"
                else:
                    cell_range = f"{col_letter}{merge_start}"
                merge_and_style_with_wrap(ws, cell_range, full_border)
                end_rows.append(row - 1)
            current_value = cell_value
            merge_start = row

    return end_rows

def ensure_file_and_workbook(input_data):
        if isinstance(input_data, Workbook):
            # Convertir a BytesIO y guardar temporalmente
            bio = BytesIO()
            input_data.save(bio)
            bio.seek(0)
            path = get_temp_file(bio, ".xlsx")
            wb = input_data
        elif isinstance(input_data, BytesIO):
            input_data.seek(0)
            path = get_temp_file(input_data, ".xlsx")
            wb = load_workbook(path)
        else:
            raise TypeError("Se esperaba un Workbook o un BytesIO.")
        return path, wb

def get_kpis_tables(xlsx_tablas, xlsx_kpis_list):

    # file_xlsx_tablas = get_temp_file(xlsx_tablas, ".xlsx")
    # file_xlsx_kpis_list = get_temp_file(xlsx_kpis_list, ".xlsx")

    # wb_tablas = load_workbook(file_xlsx_tablas)
    # wb_kpis_list = load_workbook(file_xlsx_kpis_list)

    file_xlsx_tablas, wb_tablas = ensure_file_and_workbook(xlsx_tablas)
    file_xlsx_kpis_list, wb_kpis_list = ensure_file_and_workbook(xlsx_kpis_list)

    # Create a new Workbook
    wb_new = Workbook()

    # Remove the default sheet created with the new workbook
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)

    sheets_dfs = pd.read_excel(file_xlsx_tablas, sheet_name=None)


    kpis_df_questions = pd.read_excel(
        file_xlsx_kpis_list,
        usecols="A,B,C,D",
        names=["group_kpi","names_kpis", "number_question_kpi", "number_question2_kpi"],
    ).dropna(subset=["names_kpis"])

    # Limpiar espacios al inicio y final de celdas tipo texto
    kpis_df_questions = kpis_df_questions.applymap(
        lambda x: x.strip() if isinstance(x, str) else x
    )

    visits_df_names = pd.read_excel(
        file_xlsx_kpis_list,
        usecols="E,F",
        names=["visit", "name_visit"],
    ).dropna(subset=["visit"])

    # Limpiar espacios al inicio y final de celdas tipo texto
    visits_df_names = visits_df_names.applymap(
        lambda x: x.strip() if isinstance(x, str) else x
    )

    # Para la columna 2
    kpis_df_questions[kpis_df_questions.columns[2]] = kpis_df_questions[kpis_df_questions.columns[2]].apply(
        lambda x: f"{x.upper()}." if isinstance(x, str) and len(x.strip().split()) == 1 and not x.endswith('.') else x
    )

    # Para la columna 3
    kpis_df_questions[kpis_df_questions.columns[3]] = kpis_df_questions[kpis_df_questions.columns[3]].apply(
        lambda x: f"{x.upper()}." if isinstance(x, str) and len(x.strip().split()) == 1 and not x.endswith('.') else x
    )


    grayFill = PatternFill(
        start_color="BFBFBF", end_color="BFBFBF", fill_type="solid"
    )
    greenFill = PatternFill(
    start_color="D8E4BC", end_color="D8E4BC", fill_type="solid"
    )
    redFill = PatternFill(
        start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
    )
    centrado = Alignment(horizontal='center', vertical='center')

    # Definir estilos de borde
    medium = Side(style="medium")     # Borde grueso
    thin = Side(style="thin")
    complete_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    simple_border = Border(left=thin, right=thin, top=thin, bottom=thin)


    word_multis="MULTI"
    first_row=6

    for sheet_name, data in sheets_dfs.items():
        if data.empty:
            continue

        if sheet_name.lower().startswith("grillas") and not "mom" in sheet_name.lower():
            part2 = sheet_name.split(" ", 1)
            namepart2=" "+part2[1] if len(part2) > 1 else ""
            kpi_sheet_name= "KPI'S"+ namepart2

            ws_kpis = wb_new.create_sheet(title=kpi_sheet_name)
            ws_tables=wb_tablas[sheet_name]
            dict_multis={}

            # ws_kpis.merge_cells(f"A{first_row+2}:C{first_row+2}")
            merge_with_border_range(ws_kpis,f"A{first_row+2}:C{first_row+2}",complete_border)
            ws_kpis[f"A{first_row+2}"]="Base"
            ws_kpis[f"A{first_row+2}"].border=complete_border

            refs_list = [
                cell.value
                for cell in ws_tables[2]
                if cell.value not in (None, "", "TOTAL")
            ]

            refs_col_indexes = [
                cell.column
                for cell in ws_tables[2]
                if cell.value not in (None, "", "TOTAL")
            ]

            n = len(refs_list)

            columnas = {}
            offset = 4  # Primer columna para las refs

            visits_list= visits_df_names.iloc[:, 0].tolist()
            visits_name_list= visits_df_names.iloc[:, 1].tolist()

            for i, var in enumerate(visits_list):
                inicio = offset + i * (n + 1)
                columnas[var] = list(range(inicio, inicio + n))

            df_refs_x_visits = pd.DataFrame(columnas, index=refs_list)
            """      V1  V2  V3  V4
                Q2B   4   7  10  13
                R6Z   5   8  11  14
                df.loc["fila", "Columna"]"""

            for visit in visits_list:
                indexs=df_refs_x_visits[visit]

                fila_inicio = None
                # Paso 1: Buscar la fila donde la segunda palabra de columna A sea "V#"
                for row in ws_tables.iter_rows(min_col=1, max_col=1):
                    cell = row[0]
                    if isinstance(cell.value, str):
                        palabras = cell.value.strip().split()
                        if len(palabras) >= 2 and palabras[1] == visit:
                            fila_inicio = cell.row
                            break
                fila_total = None
                if fila_inicio:
                    # Paso 2: Buscar desde fila_inicio hacia abajo la primera ocurrencia de "Total" en columna B
                    fila_total = None
                    for row in ws_tables.iter_rows(min_row=fila_inicio, min_col=2, max_col=2):
                        cell = row[0]
                        if isinstance(cell.value, str) and cell.value.strip().lower() == "total":
                            fila_total = cell.row
                            break
                else:
                    df_refs_x_visits.drop(columns=[visit], inplace=True)

                if fila_total:
                    first_col=get_column_letter(min(indexs))
                    last_col=get_column_letter(max(indexs))
                    merge_cells_visits= f"{first_col}{first_row}:{last_col}{first_row}"
                    merge_with_border_range(ws_kpis,merge_cells_visits,complete_border)
                    ws_kpis[f"{first_col}{first_row}"]=visits_name_list[visits_list.index(visit)]
                    ws_kpis[f"{first_col}{first_row}"].fill=grayFill
                    for i, var in enumerate(indexs):
                        ws_kpis[f"{get_column_letter(var)}{first_row+1}"]=refs_list[i]
                        if fila_total:
                            ws_kpis[f"{get_column_letter(var)}{first_row+2}"]=ws_tables[f"{get_column_letter(refs_col_indexes[i])}{fila_total}"].value

            actual_row=first_row+3

            for _, row in kpis_df_questions.iterrows():
                kpi = row["names_kpis"]
                question = row["number_question_kpi"]
                question2 = row["number_question2_kpi"]
                group_kpi = row["group_kpi"]

                # Si está vacío, usar el último válido
                if pd.isna(group_kpi) or group_kpi == "":
                    group_kpi = last_valid_group
                else:
                    last_valid_group = group_kpi

                if kpi != word_multis:
                    jr_question=False
                    unique_question=False
                    promedio_question=False
                    # Paso 1: Buscar la fila donde empiece por "P#. V# "
                    for row in ws_tables.iter_rows(min_col=1, max_col=1):
                        cell = row[0]
                        if isinstance(cell.value, str):
                            palabras = cell.value.strip().split()
                            if len(palabras) >= 1 and (palabras[0] == question or palabras[0] == question2):
                                fila_inicio = cell.row
                                visit_num=palabras[1]
                                indexs=df_refs_x_visits[visit_num]
                                cell_val_option3 = ws_tables.cell(row=fila_inicio + 3, column=2).value
                                cell_val_option_NETO = ws_tables.cell(row=fila_inicio, column=2).value
                                #Si no es pregunta de escala
                                if isinstance(cell_val_option3, str) and "neto" not in cell_val_option_NETO.lower():
                                    unique_question=True
                                    fila_final = None
                                    fila_promedio= None
                                    for fila in range(fila_inicio, ws_tables.max_row):
                                        valor = ws_tables.cell(row=fila, column=2).value
                                        #Si encuentra el promedio antes que el total y no es de escala
                                        if isinstance(valor, str) and valor.strip().lower() == "promedio:":
                                            fila_promedio = fila
                                            promedio_question=True
                                            break
                                        elif isinstance(valor, str) and valor.strip().lower() == "total":
                                            fila_final = fila - 1
                                            break

                                    if promedio_question:
                                        for i, var in enumerate(indexs):
                                            ws_kpis[f"{get_column_letter(var)}{actual_row}"]=ws_tables.cell(row=fila_promedio , column=refs_col_indexes[i]).value
                                    else:
                                        for row_offset, fila in enumerate(range(fila_inicio, fila_final + 1)):
                                            for i, var in enumerate(indexs):
                                                ws_kpis[f"{get_column_letter(var)}{actual_row + row_offset}"] = ws_tables.cell(row=fila_inicio + row_offset, column=refs_col_indexes[i]).value
                                elif isinstance(cell_val_option3, str) and "just" in cell_val_option3.lower():
                                    jr_question=True
                                    for i, var in enumerate(indexs):
                                        ws_kpis[f"{get_column_letter(var)}{actual_row}"]=ws_tables.cell(row=fila_inicio , column=refs_col_indexes[i]).value
                                        ws_kpis[f"{get_column_letter(var)}{actual_row+1}"]=ws_tables.cell(row=fila_inicio + 3, column=refs_col_indexes[i]).value
                                        ws_kpis[f"{get_column_letter(var)}{actual_row+2}"]=ws_tables.cell(row=fila_inicio + 6, column=refs_col_indexes[i]).value
                                else:
                                    for i, var in enumerate(indexs):
                                        ws_kpis[f"{get_column_letter(var)}{actual_row}"]=ws_tables.cell(row=fila_inicio, column=refs_col_indexes[i]).value
                                        ws_kpis[f"{get_column_letter(var)}{actual_row+1}"]=ws_tables.cell(row=fila_inicio + 1, column=refs_col_indexes[i]).value

                    if jr_question:
                        merge_with_border_range(ws_kpis,f"B{actual_row}:B{actual_row+2}",simple_border)
                        ws_kpis[f"B{actual_row}"]=kpi
                        ws_kpis[f"C{actual_row}"]=ws_tables.cell(row=fila_inicio + 1, column=2).value
                        ws_kpis[f"C{actual_row+1}"]="JR"
                        ws_kpis[f"C{actual_row+2}"]=ws_tables.cell(row=fila_inicio + 5, column=2).value
                        ws_kpis[f"C{actual_row}"].border =simple_border
                        ws_kpis[f"C{actual_row+1}"].border =simple_border
                        ws_kpis[f"C{actual_row+2}"].border =simple_border
                        ws_kpis[f"A{actual_row}"]=group_kpi
                        ws_kpis[f"A{actual_row+1}"]=group_kpi
                        ws_kpis[f"A{actual_row+2}"]=group_kpi
                        actual_row+=3
                    elif promedio_question:
                        ws_kpis[f"B{actual_row}"]=kpi
                        ws_kpis[f"C{actual_row}"]="Promedio"
                        ws_kpis[f"C{actual_row}"].border =simple_border
                        ws_kpis[f"A{actual_row}"]=group_kpi
                        actual_row+=1
                    elif unique_question:
                        merge_with_border_range(ws_kpis,f"B{actual_row}:B{actual_row+fila_final-fila_inicio}",simple_border)
                        ws_kpis[f"B{actual_row}"]=kpi
                        for i in range(fila_final-fila_inicio+1):
                            ws_kpis[f"C{actual_row+i}"]=ws_tables.cell(row=fila_inicio+i, column=2).value
                            ws_kpis[f"C{actual_row+i}"].border =simple_border
                            ws_kpis[f"A{actual_row+i}"]=group_kpi
                        actual_row+=fila_final-fila_inicio+1
                    else:
                        merge_with_border_range(ws_kpis,f"B{actual_row}:B{actual_row+1}",simple_border)
                        ws_kpis[f"B{actual_row}"]=kpi
                        ws_kpis[f"C{actual_row}"]="T2B"
                        ws_kpis[f"C{actual_row+1}"]="TB"
                        ws_kpis[f"C{actual_row}"].border =simple_border
                        ws_kpis[f"C{actual_row+1}"].border =simple_border
                        ws_kpis[f"A{actual_row}"]=group_kpi
                        ws_kpis[f"A{actual_row+1}"]=group_kpi
                        actual_row+=2
                else:
                    if pd.notna(question2) and pd.notna(question) and str(question2).strip() != "" and len(question2.strip().split()) >= 2:
                        for row2 in ws_tables.iter_rows(min_col=1, max_col=1):
                            cell2 = row2[0]
                            if isinstance(cell2.value, str):
                                palabras2 = cell2.value.strip().split()
                                if len(palabras2) >= 1 and (palabras2[0] == question):
                                    atribute_row = " ".join(palabras2[2:]) if len(palabras2) > 2 else ""
                                    if atribute_row==question2:
                                        fila_inicio = cell2.row
                                        visit_num=palabras2[1]
                                        indexs=df_refs_x_visits[visit_num]
                                        for i, var in enumerate(indexs):
                                                ws_kpis[f"{get_column_letter(var)}{actual_row}"]=ws_tables.cell(row=fila_inicio, column=refs_col_indexes[i]).value
                                                ws_kpis[f"{get_column_letter(var)}{actual_row+1}"]=ws_tables.cell(row=fila_inicio + 1, column=refs_col_indexes[i]).value
                        merge_with_border_range(ws_kpis,f"B{actual_row}:B{actual_row+1}",simple_border)
                        ws_kpis[f"B{actual_row}"]=question2
                        ws_kpis[f"C{actual_row}"]="T2B"
                        ws_kpis[f"C{actual_row+1}"]="TB"
                        ws_kpis[f"C{actual_row}"].border =simple_border
                        ws_kpis[f"C{actual_row+1}"].border =simple_border
                        ws_kpis[f"A{actual_row}"]=group_kpi
                        ws_kpis[f"A{actual_row+1}"]=group_kpi
                        actual_row+=2
                        if question not in dict_multis:
                            dict_multis[question]=[]
                        dict_multis[question].append(question2)
                    else:
                        list_atributes= dict_multis[question] if question in dict_multis else []
                        for row in ws_tables.iter_rows(min_col=1, max_col=1):
                            cell = row[0]
                            if isinstance(cell.value, str):
                                palabras = cell.value.strip().split()
                                if len(palabras) >= 1 and (palabras[0] == question or palabras[0] == question2):
                                    atribute_question = " ".join(palabras[2:]) if len(palabras) > 2 else ""
                                    if not atribute_question in list_atributes:
                                        list_atributes.append(atribute_question)
                                        for row2 in ws_tables.iter_rows(min_col=1, max_col=1):
                                            cell2 = row2[0]
                                            if isinstance(cell2.value, str):
                                                palabras2 = cell2.value.strip().split()
                                                if len(palabras2) >= 1 and (palabras2[0] == question or palabras2[0] == question2):
                                                    atribute_row = " ".join(palabras2[2:]) if len(palabras2) > 2 else ""
                                                    if atribute_row==atribute_question:
                                                        fila_inicio = cell2.row
                                                        visit_num=palabras2[1]
                                                        indexs=df_refs_x_visits[visit_num]
                                                        for i, var in enumerate(indexs):
                                                                ws_kpis[f"{get_column_letter(var)}{actual_row}"]=ws_tables.cell(row=fila_inicio, column=refs_col_indexes[i]).value
                                                                ws_kpis[f"{get_column_letter(var)}{actual_row+1}"]=ws_tables.cell(row=fila_inicio + 1, column=refs_col_indexes[i]).value

                                        merge_with_border_range(ws_kpis,f"B{actual_row}:B{actual_row+1}",simple_border)
                                        ws_kpis[f"B{actual_row}"]=atribute_question
                                        ws_kpis[f"C{actual_row}"]="T2B"
                                        ws_kpis[f"C{actual_row+1}"]="TB"
                                        ws_kpis[f"C{actual_row}"].border =simple_border
                                        ws_kpis[f"C{actual_row+1}"].border =simple_border
                                        ws_kpis[f"A{actual_row}"]=group_kpi
                                        ws_kpis[f"A{actual_row+1}"]=group_kpi
                                        actual_row+=2

            for ref_name, fila_columnas in df_refs_x_visits.iterrows():
                columnas = fila_columnas.tolist()  # Ej: [4, 7, 10, 13]

                pares = list(combinations(columnas, 2))

                # Iterar sobre las filas de la hoja Excel a partir de first_row + 3
                for row in ws_kpis.iter_rows(min_row=first_row + 3, values_only=False):
                    fila_excel = row[0].row  # Número de fila actual

                    for col1, col2 in pares:
                        x1val = ws_kpis.cell(row=fila_excel, column=col1).value
                        x2val = ws_kpis.cell(row=fila_excel, column=col2).value
                        n1val = ws_kpis.cell(row=first_row+2, column=col1).value
                        n2val = ws_kpis.cell(row=first_row+2, column=col2).value


                        # Saltar si alguno es None o no es string
                        if x1val in (None, "") or x2val in (None, ""):
                            continue

                        # Intentar extraer el primer número entero antes del primer espacio
                        try:
                            px1 = int(str(x1val).strip().split()[0])
                            px2 = int(str(x2val).strip().split()[0])
                            n1 = int(str(n1val).strip().split()[0])
                            n2 = int(str(n2val).strip().split()[0])
                            x1 = round((px1 / 100) * n1)
                            x2 = round((px2 / 100) * n2)
                        except (IndexError, ValueError):
                            continue  # Saltar si no se puede convertir

                        # Aquí haces tu comparación
                        if calculate_differences(x1, x2, n1, n2):
                            if (x1 / n1) > (x2 / n2):
                                cell_val = str(ws_kpis.cell(row=fila_excel, column=col1).value).strip()
                                last_word = cell_val.replace("\n", " ").split()[-1]

                                if "l" in last_word:
                                    ws_kpis.cell(row=fila_excel, column=col1).value+="-"+entero_a_romano(columnas.index(col2)+1)
                                else:
                                    ws_kpis.cell(row=fila_excel, column=col1).value+=" "+entero_a_romano(columnas.index(col2)+1)
                            else:
                                cell_val = str(ws_kpis.cell(row=fila_excel, column=col2).value).strip()
                                last_word = cell_val.replace("\n", " ").split()[-1]

                                if "l" in last_word:
                                    ws_kpis.cell(row=fila_excel, column=col2).value+="-"+entero_a_romano(columnas.index(col1)+1)
                                else:
                                    ws_kpis.cell(row=fila_excel, column=col2).value+=" "+entero_a_romano(columnas.index(col1)+1)
                    for col in columnas:
                        apply_red_and_blue_color_to_letter(ws_kpis.cell(row=fila_excel, column=col))

            bottom_thick_border_list = merge_identical_cells_with_border(ws_kpis,"A",first_row+3)
            left_thick_border_list = df_refs_x_visits.iloc[0].tolist()
            right_thick_border_list = df_refs_x_visits.iloc[-1].tolist()


            columns_refs_data = df_refs_x_visits.values.flatten().tolist()
            for row in ws_kpis.iter_rows():
                for cell in row:
                    cell.alignment = centrado

                    col_idx = cell.column  # número de columna
                    fila = cell.row
                    if col_idx in [3] and fila in bottom_thick_border_list:
                        apply_medium_bottom_border(cell)
                    if col_idx in [2] and fila-1 in bottom_thick_border_list:
                        apply_medium_top_border(cell)
                    if fila >= first_row+1 and col_idx in columns_refs_data:
                        cell.alignment = centrado

                        if cell.value in (None, ""):
                            cell.value="-"

                        # Obtener valores en la columna 3 (columna C = índice 3)
                        val_actual = ws_kpis.cell(row=fila, column=3).value
                        val_arriba = ws_kpis.cell(row=fila - 1, column=3).value if fila > 1 else None
                        val_abajo = ws_kpis.cell(row=fila + 1, column=3).value

                        # Verifica condiciones especiales
                        if val_actual == "T2B" or val_abajo == "JR":
                            cell.fill=greenFill
                        elif val_arriba == "JR":
                            cell.fill=redFill

                        left_border = medium if col_idx in left_thick_border_list else thin
                        right_border = medium if col_idx in right_thick_border_list else thin
                        bottom_border = medium if fila in bottom_thick_border_list or fila == first_row + 2 else thin

                        top_border=thin

                        custom_border = Border(
                            left=left_border,
                            top=top_border,
                            right=right_border,
                            bottom=bottom_border
                        )

                        cell.border=custom_border

            if ws_kpis.max_row-1 in bottom_thick_border_list:
                apply_medium_top_border(ws_kpis.cell(row=ws_kpis.max_row, column=2))

            """
            Set specific pixel-based column widths:
            - B: 235 px
            - C: 122 px
            - D to last column with data:
                - 80 px if in columns_refs_data
                - 20 px if not
            """

            # Pixel-to-width approximations for openpyxl (1 width unit ≈ 7.001 pixels for Calibri 11)
            px_to_width = lambda px: round(px / 7.001, 2)

            # Set fixed widths for B and C
            ws_kpis.column_dimensions["B"].width = px_to_width(235)
            ws_kpis.column_dimensions["C"].width = px_to_width(122)

            # Get last column with data
            last_col_idx = max(cell.column for row in ws_kpis.iter_rows() for cell in row if cell.value)

            for col_idx in range(4, last_col_idx + 1):
                col_letter = get_column_letter(col_idx)
                if col_idx in columns_refs_data:
                    ws_kpis.column_dimensions[col_letter].width = px_to_width(80)
                else:
                    ws_kpis.column_dimensions[col_letter].width = px_to_width(20)

    return write_temp_excel(wb_new),wb_new

def correction_word_anios(texto):
    # Definir patrones y reemplazos con mayúsculas y minúsculas
    patrones = [
        (r'\banos\b', 'años'),
        (r'\bAnos\b', 'Años'),
        (r'\bANOS\b', 'AÑOS'),
        (r'\bano\b', 'año'),
        (r'\bAno\b', 'Año'),
        (r'\bANO\b', 'AÑO'),
    ]
    for patron, reemplazo in patrones:
        texto = re.sub(patron, reemplazo, texto)
    return texto
def corregir_nombre_hojas_con_anios(wb):
    for sheet in wb.worksheets:
        nombre_original = sheet.title
        nuevo_nombre = correction_word_anios(nombre_original)
        if nuevo_nombre != nombre_original:
            if nuevo_nombre not in wb.sheetnames:
                sheet.title = nuevo_nombre

def copy_sheets(source_wb, target_wb, prefix=""):
    for ws in source_wb.worksheets:
        new_title = f"{prefix}{ws.title}"
        new_ws = target_wb.create_sheet(title=new_title)

        # Copiar celdas
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue  # Ignorar celdas fusionadas intermedias

                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)

                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copiar celdas combinadas
        for merged_range in ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))

        # Copiar anchos de columnas
        for col_letter, col_dim in ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width

        # Copiar alturas de filas
        for row_idx, row_dim in ws.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = row_dim.height

def update_penalty_sheets(wb_postprocess_kpis, xlsx_kpis_list):
    wb_new_penaltys= Workbook()
    # Remove the default sheet created with the new workbook
    default_sheet = wb_new_penaltys.active
    wb_new_penaltys.remove(default_sheet)


    file_xlsx_kpis_list, wb_kpis_list = ensure_file_and_workbook(xlsx_kpis_list)
    # Cargar el archivo
    kpis_df_questions = pd.read_excel(
        file_xlsx_kpis_list,
        usecols="C,D",
        names=["number_question_kpi", "number_question2_kpi"]
    )

    # Limpiar espacios al inicio y final de celdas tipo texto
    kpis_df_questions = kpis_df_questions.applymap(
        lambda x: x.strip() if isinstance(x, str) else x
    )

    # Eliminar filas con valores nulos en ambas columnas
    kpis_df_questions.dropna(subset=["number_question_kpi", "number_question2_kpi"], inplace=True)

    # Filtrar para que ambas columnas tengan exactamente una palabra (sin espacios)
    kpis_df_questions = kpis_df_questions[
        kpis_df_questions["number_question_kpi"].astype(str).str.strip().str.count(r"\s") == 0
    ]
    kpis_df_questions = kpis_df_questions[
        kpis_df_questions["number_question2_kpi"].astype(str).str.strip().str.count(r"\s") == 0
    ]
    # Para la columna 2
    kpis_df_questions[kpis_df_questions.columns[0]] = kpis_df_questions[kpis_df_questions.columns[0]].apply(
        lambda x: f"{x.upper()}." if isinstance(x, str) and len(x.strip().split()) == 1 and not x.endswith('.') else x
    )

    # Para la columna 3
    kpis_df_questions[kpis_df_questions.columns[1]] = kpis_df_questions[kpis_df_questions.columns[1]].apply(
        lambda x: f"{x.upper()}." if isinstance(x, str) and len(x.strip().split()) == 1 and not x.endswith('.') else x
    )

    question_pairs = kpis_df_questions[["number_question_kpi", "number_question2_kpi"]].values.tolist()

    # Limpiar espacios al inicio y final de celdas tipo texto
    kpis_df_questions = kpis_df_questions.applymap(
        lambda x: x.strip() if isinstance(x, str) else x
    )
    file_xlsx_kpis_list, _ = ensure_file_and_workbook(xlsx_kpis_list)
    visits_df_names = pd.read_excel(
        file_xlsx_kpis_list,
        usecols="E,F",
        names=["visit", "name_visit"],
    ).dropna(subset=["visit"])

    # Limpiar espacios al inicio y final de celdas tipo texto
    visits_df_names = visits_df_names.applymap(
        lambda x: x.strip() if isinstance(x, str) else x
    )
    grayFill = PatternFill(
        start_color="BFBFBF", end_color="BFBFBF", fill_type="solid"
    )
    redFill = PatternFill(
        start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
    )
    # Definir estilos de borde
    medium = Side(style="medium")     # Borde grueso
    thin = Side(style="thin")
    complete_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    simple_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)

    first_row=6
    for sheet_name in wb_postprocess_kpis.sheetnames:
        if sheet_name.lower().startswith("penaltys"):
            ws_kpis = wb_new_penaltys.create_sheet(title=sheet_name)
            ws_tables=wb_postprocess_kpis[sheet_name]

            merge_with_border_range(ws_kpis,f"A{first_row+2}:B{first_row+2}",complete_border)
            ws_kpis[f"A{first_row+2}"]="Base"
            ws_kpis[f"A{first_row+2}"].border=complete_border

            refs_list = [
                cell.value
                for cell in ws_tables[1]
                if cell.value not in (None, "", "Grouped Variable")
            ]

            refs_col_indexes = [
                cell.column
                for cell in ws_tables[1]
                if cell.value not in (None, "", "Grouped Variable")
            ]

            n = len(refs_list)

            columnas = {}
            offset = 3  # Primer columna para las refs

            visits_list= visits_df_names.iloc[:, 0].tolist()
            visits_name_list= visits_df_names.iloc[:, 1].tolist()

            for i, var in enumerate(visits_list):
                inicio = offset + i * (n + 1)
                columnas[var] = list(range(inicio, inicio + n))

            df_refs_x_visits = pd.DataFrame(columnas, index=refs_list)
            """      V1  V2  V3  V4
                Q2B   4   7  10  13
                R6Z   5   8  11  14
                df.loc["fila", "Columna"]"""

            for visit in visits_list:
                indexs=df_refs_x_visits[visit]

                fila_inicio = None
                # Paso 1: Buscar la fila donde la segunda palabra de columna A sea "V#"
                for row in ws_tables.iter_rows(min_col=1, max_col=1):
                    cell = row[0]
                    if isinstance(cell.value, str):
                        palabras = cell.value.strip().split()
                        if len(palabras) >= 2 and palabras[1] == visit:
                            fila_inicio = cell.row
                            break
                fila_total = None
                if fila_inicio:
                    # Paso 2: Buscar desde fila_inicio hacia abajo la primera ocurrencia de "Total" en columna B
                    fila_total = None
                    for row in ws_tables.iter_rows(min_row=fila_inicio, min_col=2, max_col=2):
                        cell = row[0]
                        if isinstance(cell.value, str) and cell.value.strip().lower() == "total":
                            fila_total = cell.row
                            break
                else:
                    df_refs_x_visits.drop(columns=[visit], inplace=True)

                if fila_total:
                    first_col=get_column_letter(min(indexs))
                    last_col=get_column_letter(max(indexs))
                    merge_cells_visits= f"{first_col}{first_row}:{last_col}{first_row}"
                    merge_with_border_range(ws_kpis,merge_cells_visits,complete_border)
                    ws_kpis[f"{first_col}{first_row}"]=visits_name_list[visits_list.index(visit)]
                    ws_kpis[f"{first_col}{first_row}"].fill=grayFill
                    for i, var in enumerate(indexs):
                        ws_kpis[f"{get_column_letter(var)}{first_row+1}"]=refs_list[i]
                        if fila_total:
                            ws_kpis[f"{get_column_letter(var)}{first_row+2}"]=ws_tables[f"{get_column_letter(refs_col_indexes[i])}{fila_total}"].value

            actual_row=first_row+3
            list_questions=[]
            bottom_thick_border_list=[]
            for row in ws_tables.iter_rows():
                label_question = row[0]
                init_row=label_question.row
                if label_question.value is None:
                    continue
                label_question_split = label_question.value.strip().split()
                question =label_question_split[0]
                if question not in list_questions:
                    questions=[]
                    if any(question in pair for pair in question_pairs):
                        questions =[pair for pair in question_pairs if question in pair][0]
                        for q in questions:
                            list_questions.append(q)
                    else:
                        questions=[question]
                        list_questions.append(question)

                    for row2 in ws_tables.iter_rows(min_col=1, max_col=1):
                        cell = row2[0]
                        if isinstance(cell.value, str):
                            palabras = cell.value.strip().split()
                            if len(palabras) >= 1 and palabras[0] in questions:
                                fila_inicio = cell.row
                                visit_num=palabras[1]
                                indexs=df_refs_x_visits[visit_num]
                                for i, var in enumerate(indexs):
                                    ws_kpis[f"{get_column_letter(var)}{actual_row}"]=ws_tables.cell(row=fila_inicio + 6, column=refs_col_indexes[i]).value
                                    ws_kpis[f"{get_column_letter(var)}{actual_row+1}"]=ws_tables.cell(row=fila_inicio + 7, column=refs_col_indexes[i]).value
                                    for offset in [0, 1]:  # Para actual_row y actual_row + 1
                                        cell = ws_kpis[f"{get_column_letter(var)}{actual_row + offset}"]
                                        if isinstance(cell.value, (int, float, str)) and str(cell.value).strip() != "":
                                            try:
                                                val = round(float(cell.value), 2)
                                                cell.value = val
                                                cell.number_format = "0.00"
                                                if val <= -3:
                                                    cell.fill = redFill
                                            except ValueError:
                                                pass

                    if isinstance(label_question.value, str):
                        words = label_question.value.strip().split()
                        if len(words) > 1:
                            label_question_str = " ".join([words[0]] + words[2:])
                        else:
                            label_question_str = label_question.value.strip()
                    merge_with_border_range(ws_kpis,f"A{actual_row}:A{actual_row+1}",simple_border)
                    ws_kpis[f"A{actual_row}"]=label_question_str
                    ws_kpis[f"B{actual_row}"]=ws_tables.cell(row=init_row + 6, column=2).value
                    ws_kpis[f"B{actual_row+1}"]=ws_tables.cell(row=init_row + 7, column=2).value
                    ws_kpis[f"B{actual_row}"].border =simple_border
                    ws_kpis[f"B{actual_row+1}"].border =simple_border
                    bottom_thick_border_list.append(actual_row+1)
                    actual_row+=2

            left_thick_border_list = df_refs_x_visits.iloc[0].tolist()
            right_thick_border_list = df_refs_x_visits.iloc[-1].tolist()


            columns_refs_data = df_refs_x_visits.values.flatten().tolist()
            for row in ws_kpis.iter_rows():
                for cell in row:
                    cell.alignment = centrado

                    col_idx = cell.column  # número de columna
                    fila = cell.row
                    if col_idx in [2] and fila in bottom_thick_border_list:
                        apply_medium_bottom_border(cell)
                    if col_idx in [1] and fila-1 in bottom_thick_border_list:
                        apply_medium_top_border(cell)
                    if fila >= first_row+1 and col_idx in columns_refs_data:
                        cell.alignment = centrado

                        if cell.value in (None, ""):
                            cell.value="-"

                        left_border = medium if col_idx in left_thick_border_list else thin
                        right_border = medium if col_idx in right_thick_border_list else thin
                        bottom_border = medium if fila in bottom_thick_border_list or fila == first_row + 2 else thin

                        top_border=thin

                        custom_border = Border(
                            left=left_border,
                            top=top_border,
                            right=right_border,
                            bottom=bottom_border
                        )

                        cell.border=custom_border

            if ws_kpis.max_row-1 in bottom_thick_border_list:
                apply_medium_top_border(ws_kpis.cell(row=ws_kpis.max_row, column=1))
            # Pixel-to-width approximations for openpyxl (1 width unit ≈ 7.001 pixels for Calibri 11)
            px_to_width = lambda px: round(px / 7.001, 2)

            # Set fixed widths for B and C
            ws_kpis.column_dimensions["A"].width = px_to_width(235)
            ws_kpis.column_dimensions["B"].width = px_to_width(150)
            # Get last column with data
            last_col_idx = max(cell.column for row in ws_kpis.iter_rows() for cell in row if cell.value)

            for col_idx in range(3, last_col_idx + 1):
                col_letter = get_column_letter(col_idx)
                if col_idx in columns_refs_data:
                    ws_kpis.column_dimensions[col_letter].width = px_to_width(80)
                else:
                    ws_kpis.column_dimensions[col_letter].width = px_to_width(20)
            for row in range(1, ws_kpis.max_row + 1):
                ws_kpis.row_dimensions[row].height = 15


    for sheet_name in wb_new_penaltys.sheetnames:
        if sheet_name in wb_postprocess_kpis.sheetnames:
            wb_postprocess_kpis.remove(wb_postprocess_kpis[sheet_name])
    copy_sheets(wb_new_penaltys,wb_postprocess_kpis)


def get_diferences_with_kpis(xlsx_pretables_file: BytesIO, xlsx_kpis_list: BytesIO):
    _, wb_tables=processing(xlsx_pretables_file)
    _, wb_kpis=get_kpis_tables(wb_tables,xlsx_kpis_list)

    nombre_indice="indice"
    wb_new = Workbook()
    ws_index = wb_new.active
    ws_index.title = nombre_indice

    # Copiar hojas de wb_kpis
    copy_sheets(wb_kpis, wb_new)

    # Copiar hojas de wb_tables
    copy_sheets(wb_tables, wb_new)


    corregir_nombre_hojas_con_anios(wb_new)
    for sheet_name in wb_new.sheetnames:
        if sheet_name.lower().startswith("grilla") and "mom" not in sheet_name.lower():
            # Obtener la hoja
            sheet = wb_new[sheet_name]

            # Obtener la segunda fila (índice 2, pero openpyxl empieza en 1)
            second_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]

            # Filtrar elementos no vacíos y que no sean 'TOTAL'
            refs_list = [cell for cell in second_row if cell and str(cell).strip().upper() != "TOTAL"]
            break


    # Estilos# Definir estilos de borde
    medium = Side(style="medium")     # Borde grueso
    thin = Side(style="thin")
    bold = Font(bold=True)
    enlace = Font(color="0000EE", underline="single")
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fondo_header = PatternFill("solid", fgColor="F79646")
    fondo_sections = PatternFill("solid", fgColor="FCD5B4")
    fondo_subsections = PatternFill("solid", fgColor="FDE9D9")
    complete_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    simple_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    style_border_tables="medium"

    # Título
    ws_index["B2"] = "ÍNDICE"
    merge_with_border_range(ws_index,"B2:H2",complete_border)
    ws_index["B2"].font = Font(bold=True, color="FFFFFF", size=14)
    ws_index["B2"].fill = fondo_header

    if "Filtros" in wb_new.sheetnames:
        cellf = ws_index.cell(row=4, column=2, value="Filtros")   # D con hipervínculo
        cellf.hyperlink = f"#'Filtros'!A1"
        cellf.font = enlace
        cellf.border=complete_border

    # Obtener listas desde el wb
    grillas_sheets = [s for s in wb_new.sheetnames if s.lower().startswith("grillas")]
    penaltys_sheets = [s for s in wb_new.sheetnames if s.lower().startswith("penaltys")]
    kpis_sheets = [s for s in wb_new.sheetnames if s.lower().startswith("kpi")]
    abiertas_sheets = [s for s in wb_new.sheetnames if s.lower().startswith("abiertas")]

    # refs_list desde J3 hacia abajo
    start_row_refs = 3
    start_col_refs = 10  # J
    for i, ref in enumerate(refs_list):
        ws_index.cell(row=start_row_refs + i, column=start_col_refs, value=ref)
    apply_outer_border_range(ws_index,f"J{start_row_refs}:L{start_row_refs+len(refs_list)-1}",style_border_tables)

    start_row_tables = start_row_refs+len(refs_list)+4

    if grillas_sheets:
        # Grillas
        ws_index[f"B{start_row_tables}"] = "Grillas"
        ws_index[f"B{start_row_tables}"].font = bold
        ws_index[f"B{start_row_tables}"].fill = fondo_sections
        for i, name in enumerate(grillas_sheets, start=start_row_tables):
            part2 = name.split(" ", 1)
            namepart2=part2[1] if len(part2) > 1 else ""
            safe_name = quote_sheetname(name)
            if namepart2=="":
                cell = ws_index.cell(row=i, column=3, value="Total")   # D con hipervínculo
                cell.hyperlink = f"#{safe_name}!A1"
                cell.font = enlace
                merge_with_border_range(ws_index,f"C{i}:D{i}",complete_border)
            else:
                ws_index.cell(row=i, column=3).fill = fondo_subsections  # C vacía con color
                cell = ws_index.cell(row=i, column=4, value=namepart2)   # D con hipervínculo
                cell.hyperlink = f"#{safe_name}!A1"
                cell.font = enlace
                cell.border=simple_border
        merge_with_border_range(ws_index,f"B{start_row_tables}:B{start_row_tables+len(grillas_sheets)-1}",complete_border)
        apply_outer_border_range(ws_index,f"C{start_row_tables}:D{start_row_tables+len(grillas_sheets)-1}",style_border_tables)

    if penaltys_sheets:
        # Penaltys
        ws_index[f"F{start_row_tables}"] = "Penaltys"
        ws_index[f"F{start_row_tables}"].font = bold
        ws_index[f"F{start_row_tables}"].fill = fondo_sections
        for i, name in enumerate(penaltys_sheets, start=start_row_tables):
            part2 = name.split(" ", 1)
            namepart2=part2[1] if len(part2) > 1 else ""
            safe_name = quote_sheetname(name)
            if namepart2=="":
                cell = ws_index.cell(row=i, column=7, value="Total")   # D con hipervínculo
                cell.hyperlink = f"#{safe_name}!A1"
                cell.font = enlace
                merge_with_border_range(ws_index,f"G{i}:H{i}",complete_border)
            else:
                ws_index.cell(row=i, column=7).fill = fondo_subsections  # C vacía con color
                cell = ws_index.cell(row=i, column=8, value=namepart2)   # D con hipervínculo
                cell.hyperlink = f"#{safe_name}!A1"
                cell.font = enlace
                cell.border=simple_border
        merge_with_border_range(ws_index,f"F{start_row_tables}:F{start_row_tables+len(penaltys_sheets)-1}",complete_border)
        apply_outer_border_range(ws_index,f"G{start_row_tables}:H{start_row_tables+len(penaltys_sheets)-1}",style_border_tables)

    if kpis_sheets:
        # KPIs
        ws_index[f"J{start_row_tables}"] = "KPI's"
        ws_index[f"J{start_row_tables}"].font = bold
        ws_index[f"J{start_row_tables}"].fill = fondo_sections
        for i, name in enumerate(kpis_sheets, start=start_row_tables):
            part2 = name.split(" ", 1)
            namepart2=part2[1] if len(part2) > 1 else ""
            safe_name = quote_sheetname(name)
            if namepart2=="":
                cell = ws_index.cell(row=i, column=11, value="Total")   # D con hipervínculo
                cell.hyperlink = f"#{safe_name}!A1"
                cell.font = enlace
            else:
                cell = ws_index.cell(row=i, column=11, value=namepart2)   # D con hipervínculo
                cell.hyperlink = f"#{safe_name}!A1"
                cell.font = enlace
                cell.border=simple_border
        merge_with_border_range(ws_index,f"J{start_row_tables}:J{start_row_tables+len(kpis_sheets)-1}",complete_border)
        apply_outer_border_range(ws_index,f"K{start_row_tables}:K{start_row_tables+len(kpis_sheets)-1}",style_border_tables)

    if abiertas_sheets:
        # Abiertas
        ws_index[f"M{start_row_tables}"] = "Abiertas"
        ws_index[f"M{start_row_tables}"].font = bold
        ws_index[f"M{start_row_tables}"].fill = fondo_sections
        for i, name in enumerate(abiertas_sheets, start=start_row_tables):
            part2 = name.split(" ", 1)
            namepart2=part2[1] if len(part2) > 1 else ""
            safe_name = quote_sheetname(name)
            if namepart2=="":
                cell = ws_index.cell(row=i, column=14, value="Total")   # D con hipervínculo
                cell.hyperlink = f"#{safe_name}!A1"
                cell.font = enlace
                merge_with_border_range(ws_index,f"N{i}:O{i}",complete_border)
            else:
                ws_index.cell(row=i, column=14).fill = fondo_subsections  # C vacía con color
                cell = ws_index.cell(row=i, column=15, value=namepart2)   # D con hipervínculo
                cell.hyperlink = f"#{safe_name}!A1"
                cell.font = enlace
                cell.border=simple_border
        merge_with_border_range(ws_index,f"M{start_row_tables}:M{start_row_tables+len(grillas_sheets)-1}",complete_border)
        apply_outer_border_range(ws_index,f"N{start_row_tables}:O{start_row_tables+len(grillas_sheets)-1}",style_border_tables)

    for row in ws_index.iter_rows():
        for cell in row:
            cell.alignment = centrado

    # Pixel-to-width approximations for openpyxl (1 width unit ≈ 7.001 pixels for Calibri 11)
    px_to_width = lambda px: round(px / 7.001, 2)
    # Get last column with data
    last_col_idx = max(cell.column for row in ws_index.iter_rows() for cell in row if cell.value)

    columns_section=[2,6,10,13]
    columns_subsection=[3,7,12,14]
    columns_data=[4,8,11,15]
    columns_separator=[5]
    for col_idx in range(1, last_col_idx + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx in columns_section:
            ws_index.column_dimensions[col_letter].width = px_to_width(76)
        elif col_idx in columns_subsection:
            ws_index.column_dimensions[col_letter].width = px_to_width(100)
        elif col_idx in columns_data:
            ws_index.column_dimensions[col_letter].width = px_to_width(150)
        elif col_idx in columns_separator:
            ws_index.column_dimensions[col_letter].width = px_to_width(23)

    ws_index.sheet_view.zoomScale = 70

    update_penalty_sheets(wb_new, xlsx_kpis_list)
    return write_temp_excel(wb_new)
