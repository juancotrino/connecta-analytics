import warnings
from io import BytesIO
import string
from itertools import product

import numpy as np
import pandas as pd
from statsmodels.stats.proportion import proportions_ztest

from openpyxl import Workbook, load_workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from app.modules.utils import get_temp_file, write_temp_excel

letters_list = list(string.ascii_uppercase)


def extract_digits(cell):
    cell = str(cell)
    match = pd.Series(cell).str.extract("(\d+)")[0][0]
    return int(match) if pd.notna(match) else None


def group_consecutive_indexes(index_list):
    if not index_list:
        return []

    sorted_indexes = sorted(set(index_list))
    groups = []
    group = [sorted_indexes[0]]

    for i in range(1, len(sorted_indexes)):
        if sorted_indexes[i] == sorted_indexes[i - 1] + 1:
            group.append(sorted_indexes[i])
        else:
            groups.append(group)
            group = [sorted_indexes[i]]

    groups.append(group)
    return groups


def calculate_percentages(
    inner_df: pd.DataFrame, total_df: pd.DataFrame, total_index: int
) -> pd.DataFrame:
    percentage_inner_df = inner_df.copy()
    for column in inner_df.columns:
        percentage_inner_df[column] = (
            inner_df[column] / total_df.loc[total_index, column]
        ) * 100
    return percentage_inner_df


def calculate_differences(
    x1: int, x2: int, n1: int, n2: int, sigma: float = 0.05
) -> bool:
    if n1 < 30 or n2 < 30 or x1 == 0 or x2 == 0 or n1 == 0 or n2 == 0:
        return False

    # Número de éxitos y tamaño de muestra en cada grupo
    counts = np.array([x1, x2])
    nobs = np.array([n1, n2])

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=RuntimeWarning)
        # Prueba z de dos proporciones
        _, p_value = proportions_ztest(counts, nobs)

    return p_value < sigma


def significant_differences(
    inner_df: pd.DataFrame,
    data: pd.DataFrame,
    total_index: int,
    letters_inner_dict: dict[str, str],
) -> pd.DataFrame:
    columns = inner_df.columns
    n_cols = len(columns)

    inner_differences_df = pd.DataFrame(
        "", index=inner_df.index, columns=inner_df.columns
    )

    for i in range(n_cols):
        for j in range(i + 1, n_cols):
            col1 = columns[i]
            col2 = columns[j]

            for index in inner_df.index:
                x1 = inner_df.at[index, col1]
                x2 = inner_df.at[index, col2]
                n1 = data.loc[total_index, col1]
                n2 = data.loc[total_index, col2]

                if calculate_differences(x1, x2, n1, n2):
                    if (x1 / n1) > (x2 / n2):
                        if inner_differences_df.at[index, col1]:
                            inner_differences_df.at[index, col1] += (
                                f",{letters_inner_dict[col2]}"
                            )
                        else:
                            inner_differences_df.at[index, col1] = letters_inner_dict[
                                col2
                            ]
                    else:
                        if inner_differences_df.at[index, col2]:
                            inner_differences_df.at[index, col2] += (
                                f",{letters_inner_dict[col1]}"
                            )
                        else:
                            inner_differences_df.at[index, col2] = letters_inner_dict[
                                col1
                            ]

    return inner_differences_df


def combine_values(num: int | float, string: str, decimals: int = 2):
    if pd.isna(num) and pd.isna(string):
        return np.nan
    elif pd.isna(num):
        return string.strip()
    elif pd.isna(string):
        if isinstance(num, float):
            return f"{num:.{decimals}f}".strip()
        else:
            return str(num).strip()
    else:
        if isinstance(num, float):
            num = f"{num:.{decimals}f}"
        return f"{num} {string}".strip()


# Función para combinar los dataframes
def combine_dataframes(df1, df2, decimals=2):
    combined_data = {
        col: [
            combine_values(num, string, decimals)
            for num, string in zip(df1[col], df2[col])
        ]
        for col in df1.columns
    }

    combined_df = pd.concat(
        [pd.Series(values, name=col) for col, values in combined_data.items()], axis=1
    )

    return combined_df


# Function to copy cell styles
def copy_styles(cell_source, cell_target):
    if cell_source.has_style:
        cell_target.font = Font(
            name=cell_source.font.name,
            size=cell_source.font.size,
            bold=cell_source.font.bold,
            italic=cell_source.font.italic,
            vertAlign=cell_source.font.vertAlign,
            underline=cell_source.font.underline,
            strike=cell_source.font.strike,
            color=cell_source.font.color,
        )

        cell_target.border = Border(
            left=Side(
                border_style=cell_source.border.left.style,
                color=cell_source.border.left.color,
            ),
            right=Side(
                border_style=cell_source.border.right.style,
                color=cell_source.border.right.color,
            ),
            top=Side(
                border_style=cell_source.border.top.style,
                color=cell_source.border.top.color,
            ),
            bottom=Side(
                border_style=cell_source.border.bottom.style,
                color=cell_source.border.bottom.color,
            ),
        )

        cell_target.fill = PatternFill(
            fill_type=cell_source.fill.fill_type,
            start_color=cell_source.fill.start_color,
            end_color=cell_source.fill.end_color,
        )

        cell_target.number_format = cell_source.number_format
        cell_target.protection = Protection(
            locked=cell_source.protection.locked, hidden=cell_source.protection.hidden
        )
        cell_target.alignment = Alignment(
            horizontal=cell_source.alignment.horizontal,
            vertical=cell_source.alignment.vertical,
            text_rotation=cell_source.alignment.text_rotation,
            wrap_text=cell_source.alignment.wrap_text,
            shrink_to_fit=cell_source.alignment.shrink_to_fit,
            indent=cell_source.alignment.indent,
        )


def delete_col_with_merged_ranges(sheet, idx):
    sheet.delete_cols(idx)

    for mcr in sheet.merged_cells:
        if idx < mcr.min_col:
            mcr.shift(col_shift=-1)
        elif idx < mcr.max_col:
            mcr.shrink(right=1)


# Function to apply red color to the letter in the cell
def apply_red_color_to_letter(cell):
    value = cell.value
    if isinstance(value, str) and any(char.isalpha() for char in value):
        # Find the first non-digit character
        index = 0
        while index < len(value) and value[index].isdigit():
            index += 1

        # Split the string based on the index
        num = value[:index]
        letter = value[index:].strip()  # Remove any leading/trailing spaces

        # Apply formatting

        red = InlineFont(color="00FF0000")
        rich_text_cell = CellRichText()
        rich_text_cell.append(f"{num} ")
        rich_text_cell.append(TextBlock(red, letter))
        cell.value = rich_text_cell


def write_statistical_siginificance_sheet(
    ws_existing,
    ws_new,
    first_all_nan_index,
    combined_differences_df,
    question_groups,
    category_indexes,
):
    # Copy styles and merged cells from existing sheet to new sheet
    for row in ws_existing.iter_rows():
        for cell in row:
            new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
            copy_styles(cell, new_cell)

    # Copy column widths
    for col in ws_existing.columns:
        column_letter = col[first_all_nan_index + 1].column_letter
        ws_new.column_dimensions[column_letter].width = ws_existing.column_dimensions[
            column_letter
        ].width

    # Write DataFrame to the new worksheet, starting after the existing data
    start_row = 1
    start_column = 1
    for r_idx, row in enumerate(
        dataframe_to_rows(combined_differences_df, index=False, header=True),
        start=start_row,
    ):
        for c_idx, value in enumerate(row, start=start_column):
            if not (isinstance(value, str) and "Unnamed" in value):
                new_cell = ws_new.cell(row=r_idx, column=c_idx, value=value)

    # Copy merged cell ranges
    for merged_cell in ws_existing.merged_cells.ranges:
        ws_new.merge_cells(str(merged_cell))

    # Iterate through merged cell ranges and unmerge those in column B
    for merged_range in list(ws_new.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(merged_range.coord)

        if min_col == 2 and max_col == 2:  # Check if the merged range is in column B
            ws_new.unmerge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col,
            )

    # Apply bold formatting to the first two columns
    for row in ws_new.iter_rows():
        for cell in row[:3]:  # First two columns
            cell.font = Font(bold=True)

    # Delete column B
    delete_col_with_merged_ranges(ws_new, 2)

    # Loop through each range and apply the formatting
    for question_group in question_groups:
        for row in question_group:
            for category_group in category_indexes:
                col_start, col_end = category_group
                for col in range(col_start, col_end + 1):
                    cell = ws_new.cell(row=row + 2, column=col)
                    apply_red_color_to_letter(cell)

    ws_new.cell(row=1, column=1).value = ""

    ws_new.column_dimensions["A"].width = 400 / 8.43
    ws_new.column_dimensions["B"].width = 150 / 8.43

    # Set fixed column width for all columns
    fixed_column_width = 80 / 8.43
    for col in range(3, ws_new.max_column + 1):
        column_letter = get_column_letter(col)
        ws_new.column_dimensions[column_letter].width = fixed_column_width

    # Set fixed row height for all rows
    fixed_row_height = 20 / 1.33
    for row in range(1, ws_new.max_row + 1):
        ws_new.row_dimensions[row].height = fixed_row_height

    # Iterate through the rows and columns starting from the second column
    for row in ws_new.iter_rows(min_col=2):
        for cell in row:
            if cell.alignment.wrap_text:
                # Preserve other alignment properties and set wrap_text to False
                alignment = cell.alignment
                new_alignment = Alignment(
                    horizontal=alignment.horizontal,
                    vertical=alignment.vertical,
                    text_rotation=alignment.text_rotation,
                    wrap_text=False,
                    shrink_to_fit=alignment.shrink_to_fit,
                    indent=alignment.indent,
                    justifyLastLine=alignment.justifyLastLine,
                    readingOrder=alignment.readingOrder,
                )
                cell.alignment = new_alignment


def write_penalty_sheet(result_df: pd.DataFrame, worksheet):
    # Split the DataFrame by unique questions
    unique_questions = result_df["question"].unique().tolist()
    dfs = {
        question: result_df[result_df["question"] == question].copy()
        for question in unique_questions
    }

    # Hide gridlines
    worksheet.sheet_view.showGridLines = False

    # Write each DataFrame to the Excel sheet with formatting
    start_row = 0
    for question, df in dfs.items():
        # Write the headers
        headers = ["Grouped Variable"] + df.columns[2:].to_list()
        for col_num, header in enumerate(headers, 2):
            cell = worksheet.cell(row=start_row + 1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(border_style="thin"))
            cell.alignment = Alignment(horizontal="center")

        # Write the data
        for row_num, row_data in enumerate(
            dataframe_to_rows(df, index=False, header=False), start=start_row + 2
        ):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                if col_num > 2:  # Apply number format to numeric columns
                    cell.number_format = "0.00"
                # Apply thin borders to data cells
                if col_num > 1:
                    cell.border = Border(
                        top=Side(border_style="thin"), bottom=Side(border_style="thin")
                    )

        # Write the question in the merged cell
        cell = worksheet.cell(row=start_row + 2, column=1)
        cell.value = question
        cell.alignment = Alignment(vertical="top", wrapText=True)
        worksheet.merge_cells(
            start_row=start_row + 2,
            start_column=1,
            end_row=start_row + 1 + len(df),
            end_column=1,
        )

        for col in range(1, len(df.columns) + 1):
            worksheet.cell(row=start_row + 1, column=col).border = Border(
                top=Side(border_style="thick")
            )
            worksheet.cell(row=start_row + 1 + len(df), column=col).border = Border(
                bottom=Side(border_style="thick")
            )

        worksheet.column_dimensions["A"].width = 400 / 8.43
        worksheet.column_dimensions["B"].width = 250 / 8.43

        # Update start_row for the next table, adding 2 rows of separation
        start_row += len(df) + 3


def delete_row_with_merged_ranges(sheet, idx):
    sheet.delete_rows(idx)
    for mcr in sheet.merged_cells:
        if idx < mcr.min_row:
            mcr.shift(row_shift=-1)
        elif idx < mcr.max_row:
            mcr.shrink(bottom=1)


def transform_headers(data: pd.DataFrame):
    c = 0
    new_columns = []
    for column in data.loc[0, :]:
        if pd.isnull(column):
            new_columns.append(f"Unnamed: {c}")
            c += 1
        else:
            new_columns.append(column)

    data.columns = new_columns
    data = data.drop(0).reset_index(drop=True)
    return data


def composite_columns(n: int, start: str = "AA"):
    letters = string.ascii_uppercase
    start_index = letters.index(start[0]) * 26 + letters.index(start[1]) + 1
    columns = ["".join(pair) for pair in product(letters, repeat=2)]
    return columns[start_index - 1 : start_index - 1 + n]


def get_totals_from_pretables(xlsx_file: BytesIO):
    temp_file_name_xlsx = get_temp_file(xlsx_file, ".xlsx")

    # Load the existing Excel file
    wb_existing = load_workbook(temp_file_name_xlsx)

    wb_new = Workbook()
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)
    sheets_dfs = pd.read_excel(temp_file_name_xlsx, sheet_name=None)

    ws_totals = wb_new.create_sheet(title="TOTALES")
    ws_option_5 = wb_new.create_sheet(title="OPCION 5")
    index_col_totals = 1
    index_col_opt_5 = 1

    # Iterate over all sheets
    for sheet_name, data in sheets_dfs.items():
        if data.empty:
            continue

        if not sheet_name.lower().startswith("penal"):
            ws_existing = wb_existing[sheet_name]
            ws_totals.cell(row=1, column=index_col_totals).value = sheet_name
            ws_option_5.cell(row=1, column=index_col_opt_5).value = sheet_name

            redFill = PatternFill(
                start_color="C80000", end_color="C80000", fill_type="solid"
            )

            yellowFill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

            blueFill = PatternFill(
                start_color="C5D9F1", end_color="C5D9F1", fill_type="solid"
            )

            max_col = ws_existing.max_column
            index_row_totals = 2
            index_ini = 0
            index_row_opt_5 = 2
            flag_opt_5 = False
            for row_actual in range(1, ws_existing.max_row + 1):
                val_a = ws_existing["A" + str(row_actual)].value
                val_b = ws_existing["B" + str(row_actual)].value
                val_c = ws_existing["C" + str(row_actual)].value
                if val_a:
                    if val_c == "1":
                        ws_totals.cell(
                            row=index_row_totals, column=index_col_totals
                        ).value = val_a
                elif val_c == "Total":
                    if index_ini == 0:
                        index_ini = row_actual
                    for i in range(4, max_col + 1):
                        ws_totals.cell(
                            row=index_row_totals, column=index_col_totals + i - 3
                        ).value = ws_existing.cell(row=row_actual, column=i).value
                        if ws_existing.cell(row=row_actual, column=i).value is np.nan:
                            ws_totals.cell(
                                row=index_row_totals, column=index_col_totals + i - 3
                            ).fill = redFill
                        elif (
                            not ws_existing.cell(row=row_actual, column=i).value
                            == ws_existing.cell(row=index_ini, column=i).value
                        ):
                            ws_totals.cell(
                                row=index_row_totals, column=index_col_totals + i - 3
                            ).fill = yellowFill
                    index_row_totals += 1
                if val_b == "NETO TOP TWO BOX":
                    ws_option_5.cell(
                        row=index_row_opt_5, column=index_col_opt_5
                    ).value = ws_existing.cell(row=row_actual, column=1).value
                    ws_option_5.cell(
                        row=index_row_opt_5, column=index_col_opt_5 + 1
                    ).value = ws_existing.cell(row=row_actual + 1, column=2).value
                    if (
                        "5"
                        not in ws_option_5.cell(
                            row=index_row_opt_5, column=index_col_opt_5 + 1
                        ).value
                    ):
                        ws_option_5.cell(
                            row=index_row_opt_5, column=index_col_opt_5 + 1
                        ).fill = yellowFill
                    index_row_opt_5 += 1
                    flag_opt_5 = True

            index_col_totals += max_col - 1
            if flag_opt_5:
                index_col_opt_5 += 3
    separators = []

    for col in range(1, ws_totals.max_column + 1):
        column_letter = get_column_letter(col)
        if (
            ws_totals.cell(row=1, column=col).value is None
            and ws_totals.cell(row=1, column=col + 1).value is None
        ):
            ws_totals.column_dimensions[column_letter].width = 3
        if (
            ws_totals.cell(row=1, column=col).value is None
            and ws_totals.cell(row=1, column=col + 1).value is not None
        ):
            ws_totals.column_dimensions[column_letter].width = 4
            separators.append(col)
        if ws_totals.cell(row=1, column=col).value is not None:
            ws_totals.column_dimensions[column_letter].width = 14

    for col in separators:
        for i in range(1, ws_totals.max_row + 1):
            ws_totals.cell(row=i, column=col).fill = blueFill

    separators = []

    for col in range(1, ws_option_5.max_column + 1):
        column_letter = get_column_letter(col)
        if (
            ws_option_5.cell(row=1, column=col).value is None
            and ws_option_5.cell(row=1, column=col + 1).value is None
        ):
            ws_option_5.column_dimensions[column_letter].width = 10
        if (
            ws_option_5.cell(row=1, column=col).value is None
            and ws_option_5.cell(row=1, column=col + 1).value is not None
        ):
            ws_option_5.column_dimensions[column_letter].width = 4
            separators.append(col)
        if ws_option_5.cell(row=1, column=col).value is not None:
            ws_option_5.column_dimensions[column_letter].width = 14

    for col in separators:
        for i in range(1, ws_option_5.max_row + 1):
            ws_option_5.cell(row=i, column=col).value = " "
            ws_option_5.cell(row=i, column=col).fill = blueFill

    return write_temp_excel(wb_new)


# @st.cache_data(show_spinner=False)
def processing(xlsx_file: BytesIO):
    temp_file_name_xlsx = get_temp_file(xlsx_file, ".xlsx")

    # Load the existing Excel file
    wb_existing = load_workbook(temp_file_name_xlsx)

    for sheet in wb_existing:
        if not sheet.title.lower().startswith("penal"):
            wstemp = wb_existing[sheet.title]
            maxcol = wstemp.max_column
            dictionary_netos = {}
            for rowi in range(1, wstemp.max_row + 1):
                valb = wstemp["B" + str(rowi)].value
                if (
                    valb
                    and valb.startswith("NETO")
                    and valb != "NETO TOP TWO BOX"
                    and valb != "NETO BOTTOM TWO BOX"
                ):
                    if valb not in dictionary_netos or dictionary_netos[valb] == 0:
                        dictionary_netos[valb] = 1
                        for rowf in range(rowi + 1, wstemp.max_row + 1):
                            if valb == wstemp["B" + str(rowf)].value:
                                for i in range(2, maxcol + 1):
                                    wstemp[get_column_letter(i) + str(rowi)] = wstemp[
                                        get_column_letter(i) + str(rowf)
                                    ].value
                                for merged_range in list(wstemp.merged_cells.ranges):
                                    if (
                                        merged_range.min_row >= rowf - 7
                                        and merged_range.min_row <= rowf + 5
                                    ):
                                        wstemp.merged_cells.ranges.remove(merged_range)
                                break
                    else:
                        dictionary_netos[valb] = dictionary_netos[valb] - 1
            for rowi in range(1, wstemp.max_row + 1):
                valb = wstemp["B" + str(rowi)].value
                if (
                    valb
                    and valb.startswith("NETO")
                    and valb != "NETO TOP TWO BOX"
                    and valb != "NETO BOTTOM TWO BOX"
                ):
                    for rowf in range(rowi + 1, wstemp.max_row + 1):
                        if valb == wstemp["B" + str(rowf)].value:
                            for j in range(11):
                                delete_row_with_merged_ranges(wstemp, rowf - 7)
                            break

            for rowi in range(wstemp.max_row + 1, 1, -1):
                if (
                    not wstemp["D" + str(rowi)].value
                    and not wstemp["C" + str(rowi)].value
                ):
                    delete_row_with_merged_ranges(wstemp, rowi)

    wb_existing.save(temp_file_name_xlsx)

    # Create a new Workbook
    wb_new = Workbook()

    # Remove the default sheet created with the new workbook
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)

    sheets_dfs = pd.read_excel(temp_file_name_xlsx, sheet_name=None)

    ws_totals = wb_new.create_sheet(title="TOTALES")
    index_totals = 1

    # Iterate over all sheets
    for sheet_name, data in sheets_dfs.items():
        if data.empty:
            continue

        if sheet_name.lower().startswith("penal"):
            first_row_with_data = data[~data.iloc[:, 3].isna()].index[0]
            data = data.iloc[first_row_with_data:]
            data.columns = [
                f"Unnamed: {i}" if pd.isna(col) else col
                for i, col in enumerate(data.iloc[0])
            ]
            data = data[1:]
            data = data.reset_index(drop=True)

            # Create a dataframe with the column names as the first row
            column_names_df = pd.DataFrame(
                [
                    [
                        column if not column.startswith("Unnamed") else np.nan
                        for column in data.columns
                    ]
                ],
                columns=list(range(len(data.columns))),
            )

            data.columns = range(len(data.columns))
            # Concatenate the new row with the original dataframe
            data = pd.concat([column_names_df, data]).reset_index(drop=True)

            # Rename the columns to integers
            data.columns = range(len(data.columns))

            questions = (
                data[data[0].str.startswith("P", na=False)][0]
                .dropna()
                .unique()
                .tolist()
            )

            question_idexes = np.array(
                [
                    data[data[0] == question].first_valid_index()
                    for question in questions
                ]
            )

            tables_first_indexes = (question_idexes - 2).tolist()

            tables_last_indexes = [
                data.iloc[start_idx + 2 :][
                    (data[1].iloc[start_idx + 2 :].isna())
                    & (data[2].iloc[start_idx + 2 :].isna())
                ].index[0]
                for start_idx in tables_first_indexes
            ]

            samples = data.loc[1, 3:].values.tolist()
            data.columns = ["question", "grouped_variable", "answer_option"] + samples

            tables_range_indexes = list(zip(tables_first_indexes, tables_last_indexes))

            results_dfs = []

            for question, (start, end) in zip(questions, tables_range_indexes):
                question_df = data.loc[start:end, :]

                # Finding the index of the first occurrence
                first_occurrence_index = question_df[
                    question_df["grouped_variable"].str.contains("Total", na=False)
                ].index[0]

                grouped_variables = (
                    question_df.loc[: first_occurrence_index - 1]
                    .dropna(subset="grouped_variable")["grouped_variable"]
                    .to_list()
                )

                results_calculations = (
                    grouped_variables
                    + [
                        f"MEAN {grouped_variable} VS. IC"
                        for grouped_variable in grouped_variables
                    ]
                    + [
                        f"PENALTY {grouped_variable}"
                        for grouped_variable in grouped_variables
                        if "just" not in grouped_variable.lower()
                    ]
                    + ["TOTAL"]
                )

                base_inner_df = pd.DataFrame(index=results_calculations)

                sub_df = question_df.loc[first_occurrence_index:]

                for sample in samples:
                    total = np.nan
                    for grouped_variable in grouped_variables:
                        grouped_variable_index_df = sub_df[
                            sub_df["grouped_variable"] == grouped_variable
                        ]
                        if grouped_variable_index_df.empty:
                            continue
                        grouped_variable_index = grouped_variable_index_df.index[0]
                        grouped_variable_df = sub_df.loc[
                            grouped_variable_index : grouped_variable_index + 4
                        ]

                        if grouped_variable_df[sample].sum() == 0:
                            continue

                        total = sub_df[:1][sample].values[0]

                        base_inner_df.loc[grouped_variable, sample] = (
                            grouped_variable_df[sample].sum() / total
                        )  # Percentage

                        base_inner_df.loc[f"MEAN {grouped_variable} VS. IC", sample] = (
                            grouped_variable_df[sample]
                            * np.array(list(range(0, 101, 25)))
                        ).sum() / grouped_variable_df[sample].sum()  # Mean

                    for grouped_variable in grouped_variables:
                        if (
                            "just" not in grouped_variable.lower()
                            and total is not np.nan
                        ):
                            percentage = base_inner_df.loc[grouped_variable, sample]
                            mean = base_inner_df.loc[
                                f"MEAN {grouped_variable} VS. IC", sample
                            ]
                            jr_mean = base_inner_df.loc[
                                f"MEAN {grouped_variables[1]} VS. IC", sample
                            ]

                            base_inner_df.loc[f"PENALTY {grouped_variable}", sample] = (
                                mean - jr_mean
                            ) * percentage  # Penalty

                    base_inner_df.loc["TOTAL", sample] = total  # Total

                base_inner_df = base_inner_df.reset_index(names="grouped_variable")
                base_inner_df.insert(0, "question", question)

                results_dfs.append(base_inner_df)

            result_df = pd.concat(results_dfs).reset_index(drop=True)

            worksheet = wb_new.create_sheet(title=sheet_name)
            write_penalty_sheet(result_df, worksheet)

        else:
            ws_existing = wb_existing[sheet_name]
            if "TOTAL" not in data.columns:
                data = transform_headers(data)
                delete_row_with_merged_ranges(ws_existing, 0)
            # data = pd.read_excel(temp_file_name_xlsx, sheet_name=sheet_name)
            data["Unnamed: 2"] = (
                pd.to_numeric(data["Unnamed: 2"], errors="coerce")
                .fillna(data["Unnamed: 2"])
                .fillna("")
            )
            data_differences = data.copy()

            float_types = data["Unnamed: 2"].apply(lambda x: isinstance(x, float))
            float_types = float_types[float_types]

            question_groups = group_consecutive_indexes(list(float_types.index))

            partial_df = data[
                data.index.isin([question_groups[0][0] - 1] + question_groups[0])
            ]
            initial_category_group = partial_df.loc[partial_df.index[0], "TOTAL"]  # (A)
            category_groups_columns = (
                partial_df.loc[partial_df.index[0], :].to_frame().reset_index()
            )

            initial_category_indexes = group_consecutive_indexes(
                list(
                    category_groups_columns[
                        category_groups_columns[1] == initial_category_group
                    ][1:].index
                )
            )

            category_indexes = [
                (
                    initial_category_indexes[i][-1],
                    initial_category_indexes[i + 1][0] - 1,
                )
                for i in range(len(initial_category_indexes) - 1)
            ] + [(initial_category_indexes[-1][0], len(category_groups_columns) - 1)]
            for cat in initial_category_indexes:
                if len(cat) > 1:
                    for cat1 in cat:
                        if cat1 != cat[-1]:
                            category_indexes += [(cat1, cat1)]

            total_differeces_df = pd.DataFrame(
                index=range(len(data_differences)), columns=data_differences.columns
            )

            for question_group in question_groups:
                df_total_search = data_differences.loc[
                    question_group[-1] : question_group[-1] + 6, :
                ]
                total_index = df_total_search[
                    df_total_search["Unnamed: 2"].str.contains("Total", na=False)
                ].index[0]
                data_differences.loc[total_index, "TOTAL"] = int(
                    data_differences.loc[total_index, "TOTAL"]
                )

                data_differences.update(
                    calculate_percentages(
                        data_differences[["TOTAL"]].loc[question_group, :].astype(int),
                        data_differences,
                        total_index,
                    )
                )

                for category_group in category_indexes:
                    columns_category_groups = category_groups_columns.loc[
                        category_group[0] : category_group[1]
                    ]["index"].to_list()

                    inner_df = (
                        data.loc[question_group, columns_category_groups]
                        .map(extract_digits)
                        .replace({None: np.nan})
                        .dropna(axis=1, how="all")
                    )

                    data_differences.update(inner_df)

                    data_differences.loc[total_index, inner_df.columns] = (
                        data_differences.loc[total_index, inner_df.columns]
                        .infer_objects(copy=False)
                        .fillna(0)
                        .astype(int)
                    )

                    data_differences.update(
                        calculate_percentages(inner_df, data_differences, total_index)
                    )

                    if len(inner_df.columns) > len(letters_list):
                        letters_inner_dict = {
                            column: letter
                            for column, letter in zip(
                                inner_df.columns,
                                composite_columns(len(inner_df.columns)),
                            )
                        }
                    else:
                        letters_inner_dict = {
                            column: letter
                            for column, letter in zip(
                                inner_df.columns, letters_list[: len(inner_df.columns)]
                            )
                        }

                    inner_differences_df = significant_differences(
                        inner_df, data_differences, total_index, letters_inner_dict
                    )

                    total_differeces_df.update(inner_differences_df)

            combined_differences_df = combine_dataframes(
                data_differences, total_differeces_df, 0
            )
            combined_differences_df["Unnamed: 2"] = np.where(
                combined_differences_df["Unnamed: 2"] == "1",
                combined_differences_df["Unnamed: 1"],
                combined_differences_df["Unnamed: 2"],
            )

            combined_differences_df["Unnamed: 2"] = combined_differences_df[
                "Unnamed: 2"
            ].replace("", np.nan)
            nan_df = combined_differences_df[combined_differences_df.isna().all(axis=1)]
            if not nan_df.empty:
                first_all_nan_index = combined_differences_df[
                    combined_differences_df.isna().all(axis=1)
                ].index[0]
            else:
                first_all_nan_index = 2

            ws_new = wb_new.create_sheet(title=sheet_name)

            write_statistical_siginificance_sheet(
                ws_existing,
                ws_new,
                first_all_nan_index,
                combined_differences_df,
                question_groups,
                category_indexes,
            )

            ws_totals.cell(row=1, column=index_totals).value = sheet_name

            redFill = PatternFill(
                start_color="C80000", end_color="C80000", fill_type="solid"
            )

            yellowFill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

            blueFill = PatternFill(
                start_color="C5D9F1", end_color="C5D9F1", fill_type="solid"
            )

            max_col = ws_new.max_column
            index_row = 2
            index_ini = 0
            for row_actual in range(1, ws_new.max_row + 1):
                val_a = ws_new["A" + str(row_actual)].value
                val_b = ws_new["B" + str(row_actual)].value
                if val_a:
                    ws_totals.cell(row=index_row, column=index_totals).value = val_a
                elif val_b == "Total":
                    if index_ini == 0:
                        index_ini = row_actual
                    for i in range(2, max_col):
                        ws_totals.cell(
                            row=index_row, column=index_totals + i - 1
                        ).value = ws_new.cell(row=row_actual, column=i + 1).value
                        if ws_new.cell(row=row_actual, column=i + 1).value is np.nan:
                            ws_totals.cell(
                                row=index_row, column=index_totals + i - 1
                            ).fill = redFill
                        elif (
                            not ws_new.cell(row=row_actual, column=i + 1).value
                            == ws_new.cell(row=index_ini, column=i + 1).value
                        ):
                            ws_totals.cell(
                                row=index_row, column=index_totals + i - 1
                            ).fill = yellowFill
                    index_row += 1

            index_totals += max_col
    separators = []

    for col in range(1, ws_totals.max_column + 1):
        column_letter = get_column_letter(col)
        if (
            ws_totals.cell(row=1, column=col).value is None
            and ws_totals.cell(row=1, column=col + 1).value is None
        ):
            ws_totals.column_dimensions[column_letter].width = 3
        if (
            ws_totals.cell(row=1, column=col).value is None
            and ws_totals.cell(row=1, column=col + 1).value is not None
        ):
            ws_totals.column_dimensions[column_letter].width = 4
            separators.append(col)
        if ws_totals.cell(row=1, column=col).value is not None:
            ws_totals.column_dimensions[column_letter].width = 14

    for col in separators:
        for i in range(1, ws_totals.max_row + 1):
            ws_totals.cell(row=i, column=col).fill = blueFill

    return write_temp_excel(wb_new)
