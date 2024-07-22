from io import BytesIO
import tempfile
from itertools import pairwise

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def get_temp_file(file: BytesIO):
    # Save BytesIO object to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        tmp_file.write(file.getvalue())
        temp_file_name = tmp_file.name

    return temp_file_name

def write_temp_excel(wb):
    with tempfile.NamedTemporaryFile() as tmpfile:
        # Write the DataFrame to the temporary SPSS file
        wb.save(tmpfile.name)

        with open(tmpfile.name, 'rb') as f:
            return BytesIO(f.read())

def write_worksheet(result_df: pd.DataFrame, worksheet):
    # Split the DataFrame by unique questions
    unique_questions = result_df['question'].unique().tolist()
    dfs = {question: result_df[result_df['question'] == question].copy() for question in unique_questions}

    # Hide gridlines
    worksheet.sheet_view.showGridLines = False

    # Write each DataFrame to the Excel sheet with formatting
    start_row = 0
    for question, df in dfs.items():
        # Write the headers
        headers = ["Grouped Variable"] + df.columns[2:].to_list()
        for col_num, header in enumerate(headers, 2):
            cell = worksheet.cell(row=start_row + 1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(border_style="thin"))
            cell.alignment = Alignment(horizontal='center')

        # Write the data
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row + 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                if col_num > 2:  # Apply number format to numeric columns
                    cell.number_format = '0.00'
                # Apply thin borders to data cells
                if col_num > 1:
                    cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Write the question in the merged cell
        cell = worksheet.cell(row=start_row + 2, column=1)
        cell.value = question
        cell.alignment = Alignment(vertical='top', wrapText=True)
        worksheet.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 1 + len(df), end_column=1)

        # # Apply borders to the data cells
        # for row in range(start_row + 1, start_row + 1 + len(df) + 1):
        #     for col in range(2, len(df.columns) + 1):
        #         worksheet.cell(row=row, column=col).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        #         if col > 2:  # Apply number format to numeric columns
        #             worksheet.cell(row=row, column=col).number_format = '0.00'



        for col in range(1, len(df.columns) + 1):
            worksheet.cell(row=start_row + 1, column=col).border = Border(top=Side(border_style="thick"))
            worksheet.cell(row=start_row + 1 + len(df), column=col).border = Border(bottom=Side(border_style="thick"))

        worksheet.column_dimensions["A"].width = 400 / 8.43
        worksheet.column_dimensions["B"].width = 250 / 8.43

        # Update start_row for the next table, adding 2 rows of separation
        start_row += len(df) + 3

    # return worksheet

def calculate_penalties(xlsx_file: BytesIO):

    temp_file_name_xlsx = get_temp_file(xlsx_file)

    # Create a new Workbook
    workbook = Workbook()

    # Remove the default sheet created with the new workbook
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sheets_dfs = pd.read_excel(temp_file_name_xlsx, header=None, sheet_name=None)

    # Iterate over all sheets
    for sheet_name, df in sheets_dfs.items():

        questions = df[0].dropna().unique().tolist()
        samples = df.loc[1, 3:].values.tolist()
        df.columns = ['question', 'grouped_variable', 'answer_option'] + samples
        tables_separation_indexes = sorted(list(set([0] + df.index[df.isna().all(axis=1)].values.tolist() + [len(df)])))
        tables_range_indexes = list(pairwise(tables_separation_indexes))

        results_dfs = []

        for question, (start, end) in zip(questions, tables_range_indexes):
            question_df = df.loc[start:end, :]

            # Finding the index of the first occurrence
            first_occurrence_index = question_df[question_df['grouped_variable'].str.contains('Total', na=False)].index[0]

            grouped_variables = question_df.loc[:first_occurrence_index - 1].dropna(subset='grouped_variable')['grouped_variable'].to_list()

            results_calculations = (
                grouped_variables +
                [f'MEAN {grouped_variable} VS. IC' for grouped_variable in grouped_variables] +
                [f'PENALTY {grouped_variable}' for grouped_variable in grouped_variables if 'just' not in grouped_variable.lower()] +
                ['TOTAL']
            )

            base_inner_df = pd.DataFrame(index=results_calculations)

            for sample in samples:
                for grouped_variable in grouped_variables:
                    sub_df = question_df.loc[first_occurrence_index:]
                    grouped_variable_index = sub_df[sub_df['grouped_variable'] == grouped_variable].index[0]
                    grouped_variable_df = sub_df.loc[grouped_variable_index:grouped_variable_index + 4]

                    if grouped_variable_df[sample].sum() == 0:
                        continue

                    total = sub_df[:1][sample].values[0]

                    base_inner_df.loc[grouped_variable, sample] = (
                        grouped_variable_df[sample].sum() / total
                    ) # Percentage

                    base_inner_df.loc[f'MEAN {grouped_variable} VS. IC', sample] = (
                        (grouped_variable_df[sample] * np.array(list(range(0, 101, 25)))).sum() / grouped_variable_df[sample].sum()
                    ) # Mean

                for grouped_variable in grouped_variables:

                    if 'just' not in grouped_variable.lower():
                        percentage = base_inner_df.loc[grouped_variable, sample]
                        mean = base_inner_df.loc[f'MEAN {grouped_variable} VS. IC', sample]
                        jr_mean = base_inner_df.loc[f'MEAN {grouped_variables[1]} VS. IC', sample]

                        base_inner_df.loc[f'PENALTY {grouped_variable}', sample] = (mean - jr_mean) * (percentage / total) * 100 # Penalty

                base_inner_df.loc['TOTAL', sample] = total # Total

            base_inner_df = base_inner_df.reset_index(names='grouped_variable')
            base_inner_df.insert(0, 'question', question)

            results_dfs.append(base_inner_df)

        result_df = pd.concat(results_dfs).reset_index(drop=True)

        worksheet = workbook.create_sheet(title=sheet_name)
        write_worksheet(result_df, worksheet)

    return write_temp_excel(workbook)
